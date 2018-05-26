__author__ = 'vincent'
# -*- coding:utf-8 -*-
from openpyxl import Workbook
import urllib
from pyquery import PyQuery as Pq
from selenium import webdriver
import time
import StrUtil

w= Workbook()
#创建excel工作薄
ws= w.create_sheet(0)
#创建Excel工作表

def get_page_content_str( url):
    time.sleep(1)

    try:
            print("现在开始抓取" + url)
            headers = {'User-Agent': 'Mozilla/5.0 (Windows NT 6.1; WOW64; rv:23.0) Gecko/20100101 Firefox/23.0'}
            #伪装浏览器
            request = urllib.request.Request(url=url, headers=headers)
            #构造请求
            m_fp = urllib.request.urlopen(request, timeout=500)
            #访问网站获取源码
            html_str = m_fp.read().decode('utf-8')
            #读取源码，该网站使用的编码方式是utf-8
            if html_str == None:
                html_str= m_fp.read().decode('gbk')
                if html_str == None:
                    html_str = m_fp.read().decode('gb2312')
            m_fp.close()
            return html_str
    except Exception:
            print(Exception)
#定义抓取网页源码函数

baseurl="http://list.yhd.com/c33618-0-98836?tc=0.0.16.CatMenu_Search_100000024_137709.148&tp=52.20947.157.9.1.LC6piUL-10-CVluh&ti=UF9X#page={}&sort=1"
#目标网站URL
j=1
ws.cell(row=1,column=1).value="名称"
ws.cell(row=1,column=2).value="品牌"
ws.cell(row=1,column=3).value="产地"
ws.cell(row=1,column=4).value="规格"
ws.cell(row=1,column=5).value="价格"
ws.cell(row=1,column=6).value="销量"
ws.cell(row=1,column=7).value="好评率"
#设置excel单元格表头
for i in range (1,51):
    url=baseurl.format(i)
    print("现在开始抓取第{}页".format(i+1))
    html_str=get_page_content_str(url)
    doc=Pq(html_str)
    fruits=doc("#itemSearchList>div")
    for fruit in fruits:
        j=j+1
        fruitname=Pq(fruit)("div>p:eq(1)>a").text()
        fruiturl=Pq(fruit)("div>p:eq(1)>a").attr("href")
        fruitpri=Pq(fruit)("div>p:eq(0)>em").text()
        fruitbrand=Pq(fruit)(" p.storeName > a").text()
        detailhtml=get_page_content_str(fruiturl)
        detail=Pq(detailhtml)
        #fruitbrand=detail("#brand_relevance").text()
        fruitarea=detail("#prodDetailCotentDiv > dl > dd").attr("title","产地").text()
        fruitspec=detail("#prodDetailCotentDiv > dl > dd:nth-child(2)").text()
        fruitspec=StrUtil.substr(fruitspec,'：',"'")
        fruitsale=detail("#mod_salesvolume > strong").text()
        fruitfavr=detail("#skuGoodCommentRate > span.hpl.paise > a > strong").text()
        ws.cell(row=j,column=1).value=fruitname
        ws.cell(row=j,column=2).value=fruitbrand
        ws.cell(row=j,column=3).value=fruitarea
        ws.cell(row=j,column=4).value=fruitspec
        ws.cell(row=j,column=5).value=fruitpri
        ws.cell(row=j,column=6).value=fruitsale
        ws.cell(row=j,column=7).value=fruitfavr
        #写入excel单元格
        print(fruitname+"-"+"-"+fruitbrand+"-"+fruitarea+"-"+fruitspec+"-"+fruitpri+"-"+fruitsale+"-"+fruitfavr)
        #在控制台输出结果
w.save(r'E:\pythontest\电商\onefruits.xlsx')
#保存到本地excel文件