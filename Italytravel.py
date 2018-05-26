__author__ = 'vincent'
# -*- coding:utf-8 -*-
from openpyxl import Workbook
import urllib
from pyquery import PyQuery as Pq
from selenium import webdriver
import time
import StrUtil

w= Workbook()
#创建excel工作薄
ws= w.create_sheet(0)
#创建Excel工作表

def get_page_content_str( url):
    time.sleep(1)

    try:
            print("现在开始抓取" + url)
            headers = {'User-Agent': 'Mozilla/5.0 (Windows NT 6.1; WOW64; rv:23.0) Gecko/20100101 Firefox/23.0'}
            #伪装浏览器
            request = urllib.request.Request(url=url, headers=headers)
            #构造请求
            m_fp = urllib.request.urlopen(request, timeout=500)
            #访问网站获取源码
            html_str = m_fp.read().decode('utf-8')
            #读取源码，该网站使用的编码方式是utf-8
            if html_str == None:
                html_str= m_fp.read().decode('gbk')
                if html_str == None:
                    html_str = m_fp.read().decode('gb2312')
            m_fp.close()
            return html_str
    except Exception:
            print(Exception)
#定义抓取网页源码函数

baseurl="http://www.huarenjie.com/forum.php?mod=forumdisplay&fid=512&filter=sortid&sortid=53&searchsort=1&it_adresse=5&page={}"
#目标网站URL
j=1
ws.cell(row=1,column=1).value="名称"
ws.cell(row=1,column=2).value="地址"
ws.cell(row=1,column=3).value="电话"
#设置excel单元格表头
for i in range (1,2):
    url=baseurl.format(i)
    print("现在开始抓取第{}页".format(i))
    html_str=get_page_content_str(url)
    doc=Pq(html_str)
    sellers=doc("#threadlisttableid>tbody")
    for seller in sellers:
        j=j+1
        name=Pq(seller)(" tr > td > div.taobao_title > span.taobao_titlemsg > a").text()
        if name==None:
            continue
        phone=Pq(seller)(" tr > td > div.taobao_message > p:nth-child(2)").text()
        phone=StrUtil.substr(phone,"电话： ","'")
        adress=Pq(seller)(" tr > td> div.taobao_message > p:nth-child(3)").text()
        adress=StrUtil.substr(adress,"： ","[")
        ws.cell(row=j,column=1).value=name
        ws.cell(row=j,column=2).value=adress
        ws.cell(row=j,column=3).value=phone

        #写入excel单元格
        print(name+"-"+"-"+adress+"-"+phone)
        #在控制台输出结果
w.save(r'D:\pythontest\都灵周边旅行社.xlsx')
#保存到本地excel文件