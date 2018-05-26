__author__ = 'vincent'
# -*- coding:utf-8 -*-
from openpyxl import Workbook
import urllib
from pyquery import PyQuery as Pq
from selenium import webdriver
import time
import StrUtil

w= Workbook()
#创建excel工作薄
sheetnames = w.get_sheet_names()
ws = w.get_sheet_by_name(sheetnames[0])

def get_page_content_str( url):
    time.sleep(1)

    try:
            print("现在开始抓取" + url)
            headers = {'User-Agent': 'Mozilla/5.0 (Windows NT 6.1; WOW64; rv:23.0) Gecko/20100101 Firefox/23.0'}
            #伪装浏览器
            request = urllib.request.Request(url=url, headers=headers)
            #构造请求
            m_fp = urllib.request.urlopen(request, timeout=500)
            #访问网站获取源码
            html_str = m_fp.read().decode('utf-8')
            #读取源码，该网站使用的编码方式是utf-8

            return html_str
    except Exception:
            print(Exception)
            try:
                 request = urllib.request.Request(url=url, headers=headers)
                 m_fp = urllib.request.urlopen(request, timeout=500)
                 html_str= m_fp.read().decode('gbk')
                 return html_str
            except Exception:
             print(Exception)
             try:
                 request = urllib.request.Request(url=url, headers=headers)
                 m_fp = urllib.request.urlopen(request, timeout=500)
                 html_str= m_fp.read().decode('gb2312')
                 return html_str
             except Exception:
                      print(Exception)
            m_fp.close()
#定义抓取网页源码函数

baseurl="http://price.pcauto.com.cn/cars/"
#目标网站URL
j=1
ws.cell(row=1,column=1).value="品牌"
ws.cell(row=1,column=2).value="车系"
ws.cell(row=1,column=3).value="车型"
#设置excel单元格表头
url=baseurl
html_str=get_page_content_str(url)

doc=Pq(html_str)
brands=doc("body > div.wrap.iContent > div")
for brand in brands:

    brandname=Pq(brand)("  div.layA.w88 > div.dFix > a > p").text()
    audis=Pq(brand)("  div.layB.w899.listC > div")
    for audi in audis:
        audiname=Pq(audi)("div.thA > a").text()
        cartypes=Pq(audi)("div.tbA > dl > dd > p.pTitle > a")

        for cartype in cartypes:
            j=j+1
            cartypename=str(Pq(cartype).text())
            newcartypename=cartypename
            if str(brandname) in cartypename:
                   newcartypename=StrUtil.substr(cartypename,brandname,"'")
                   if newcartypename=="":
                       newcartypename=StrUtil.substr(cartypename,"'",brandname)
                       if newcartypename=="":
                          newcartypename=cartypename
            ws.cell(row=j,column=1).value=brandname
            ws.cell(row=j,column=2).value=audiname
            ws.cell(row=j,column=3).value=newcartypename
    #写入excel单元格
            print(brandname+"-"+"-"+audiname+"-"+newcartypename)
    #在控制台输出结果
w.save(r'D:\pythontest\汽车品牌.xlsx')
#保存到本地excel文件