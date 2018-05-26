__author__ = 'vincent'
# -*- coding:utf-8 -*-
from openpyxl import Workbook
import urllib
from pyquery import PyQuery as Pq
from selenium import webdriver
import time
import StrUtil





w= Workbook()
#创建excel工作薄
ws= w.create_sheet(0)
#创建Excel工作表

def get_page_content_str( url):
    time.sleep(1)

    try:
            print("现在开始抓取" + url)
            headers = {'User-Agent': 'Mozilla/5.0 (Windows NT 6.1; WOW64; rv:23.0) Gecko/20100101 Firefox/23.0'}
            #伪装浏览器
            request = urllib.request.Request(url=url, headers=headers)
            #构造请求
            m_fp = urllib.request.urlopen(request, timeout=500)
            #访问网站获取源码
            html_str = m_fp.read().decode('utf-8')
            #读取源码，该网站使用的编码方式是utf-8
            if html_str == None:
                html_str= m_fp.read().decode('gbk')
                if html_str == None:
                    html_str = m_fp.read().decode('gb2312')
            m_fp.close()
            return html_str
    except Exception:
            print(Exception)
#定义抓取网页源码函数

baseurl="http://www.benlai.com/huanan/list-33.html?cty=246"
#目标网站URL
driver=webdriver.Chrome("D://chromedriver.exe")
driver.get(baseurl)
time.sleep(2)
a=driver.switch_to_alert()
a.accept()
j=1
ws.cell(row=1,column=1).value="名称"
ws.cell(row=1,column=2).value="规格"
ws.cell(row=1,column=3).value="价格"
#设置excel单元格表头
for i in range (1,6):
    print("现在开始抓取第{}页".format(i))
    for k in range (1,29):
        j=j+1
        fruitname="#Content > dl:nth-child({}) > dd > div > p.name > a > font".format(k)
        fruitname=driver.find_element_by_css_selector(fruitname).text
        fruiturl="#Content > dl:nth-child({}) > dd > div > p.name > a".format(k)
        fruiturl="http://www.benlai.com/"+driver.find_element_by_css_selector(fruiturl).get_attribute("href")
        detailhtml=get_page_content_str(fruiturl)
        doc=Pq(detailhtml)
        fruitpri=doc("#intro_price > p.price").text()
        fruitspec=doc("#_ProductDetails > div.good15_intro > div.good15_norm > dl:nth-child(1) > dd").text()
        ws.cell(row=j,column=1).value=fruitname
        ws.cell(row=j,column=2).value=fruitspec
        ws.cell(row=j,column=3).value=fruitpri
        #写入excel单元格
        print(fruitname+"-"+fruitspec+"-"+fruitpri)
        #在控制台输出结果
    nextpage=driver.find_element_by_link_text("下一页 >")
    nextpage.click()
w.save(r'E:\pythontest\电商\本来生活\blfruits.xlsx')
#保存到本地excel文件