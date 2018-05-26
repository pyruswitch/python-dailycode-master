__author__ = 'vincent'
# -*- coding:utf-8 -*-
import urllib
from pyquery import PyQuery as Pq
import StrUtil
import time
import os
from openpyxl import Workbook



def get_page_content_str( url):
    time.sleep(1)

    try:
            print("现在开始抓取" + url)
            headers = {'User-Agent': 'Mozilla/5.0 (Windows NT 6.1; WOW64; rv:23.0) Gecko/20100101 Firefox/23.0'}
            #伪装浏览器
            request = urllib.request.Request(url=url, headers=headers)
            #构造请求
            m_fp = urllib.request.urlopen(request, timeout=500)
            #访问网站获取源码
            html_str = m_fp.read().decode('utf-8')
            #读取源码，该网站使用的编码方式是utf-8

            return html_str
    except Exception:
            print(Exception)
            try:
                 request = urllib.request.Request(url=url, headers=headers)
                 m_fp = urllib.request.urlopen(request, timeout=500)
                 html_str= m_fp.read().decode('gbk')
                 return html_str
            except Exception:
             print(Exception)
             try:
                 request = urllib.request.Request(url=url, headers=headers)
                 m_fp = urllib.request.urlopen(request, timeout=500)
                 html_str= m_fp.read().decode('gb2312')
                 return html_str
             except Exception:
                      print(Exception)
            m_fp.close()
#定义抓取网页源码函数

url="http://www.qiban.com/category/701/content.jhtml#"
w= Workbook()
#创建excel工作薄
sheetnames = w.get_sheet_names()
ws= w.get_sheet_by_name(sheetnames[0])
#创建Excel工作表
ws.cell(row=1,column=1).value="一级分类"
ws.cell(row=1,column=2).value="二级分类"
ws.cell(row=1,column=3).value="三级分类"
j=1
html_str=get_page_content_str(url)
doc=Pq(html_str)
firgroup=doc("#menu_sidebar>li")
for fir  in firgroup:

        firname=Pq(fir)("a > div").text()

        thirgroup=Pq(fir)(" div > div.clearfix.pull-left.hovsubmeitemwra>div")

        for thir in  thirgroup:

            thirname=Pq(thir)("div>a").text()

            threegroup=Pq(thir)("ul > li")
            if len(threegroup)==0:
                j=j+1
                ws.cell(row=j,column=1).value=firname
                ws.cell(row=j,column=2).value=thirname
            for three in threegroup:
                j=j+1
                ws.cell(row=j,column=1).value=firname
                ws.cell(row=j,column=2).value=thirname
                threename=Pq(three)("a").text()



                ws.cell(row=j,column=3).value=threename

                print(firname+"-"+thirname+'-'+threename)
w.save(r'H:\upupup\电商后台\品牌需求\企办分类.xlsx')