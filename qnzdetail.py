__author__ = 'vincent'
# -*- coding:utf-8 -*-
from openpyxl import Workbook
import urllib
from pyquery import PyQuery as Pq
from selenium import webdriver
import time

w= Workbook()
#创建excel工作薄
ws= w.create_sheet(0)
#创建Excel工作表

def get_page_content_str( url):
    time.sleep(1)

    try:
            print("现在开始抓取" + url)
            headers = {'User-Agent': 'Mozilla/5.0 (Windows NT 6.1; WOW64; rv:23.0) Gecko/20100101 Firefox/23.0'}
            #伪装浏览器
            request = urllib.request.Request(url=url, headers=headers)
            #构造请求
            m_fp = urllib.request.urlopen(request, timeout=500)
            #访问网站获取源码
            html_str = m_fp.read().decode('utf-8')
            #读取源码，该网站使用的编码方式是utf-8
            if html_str == None:
                html_str= m_fp.read().decode('gbk')
                if html_str == None:
                    html_str = m_fp.read().decode('gb2312')
            m_fp.close()
            return html_str
    except Exception:
            print(Exception)
#定义抓取网页源码函数

baseurl="http://www.qianneizhu.com/tl/select.html"
#目标网站URL
driver=webdriver.Chrome("E://chromedriver.exe")
#模拟打开浏览器（因为1-90页每页网址一样，所以需要通过webdrive模拟浏览器的点击行为去访问）
driver.get(baseurl)
#模拟浏览器访问目标网站
j=1
ws.cell(row=1,column=1).value="项目名称"
ws.cell(row=1,column=2).value="收益率"
ws.cell(row=1,column=3).value="还款期限"
ws.cell(row=1,column=4).value="还款金额"
#设置excel单元格表头
for i in range (1,91):
  print("现在开始执行第{}页".format(i))
  for k in range(1,13):
     j=j+1
     targetname="#porCont > li:nth-child({}) > div.tit > a".format(k)
     #设置项目名称的css path （css地址）
     targetyieldint="#porCont > li:nth-child({}) > div.rate > div.conts > span.integer".format(k)
     #设置项目收益率整数位的css path （css地址）
     targetyielddec="#porCont > li:nth-child({}) > div.rate > div.conts > span.decimal".format(k)
     #设置项目收益率分数位的css path （css地址）
     targeturl="#porCont > li:nth-child({}) > div.tit > a".format(k)
     #设置项目详情页URL的css path （css地址）
     Proname=driver.find_element_by_css_selector(targetname).text
     #根据css地址取内容
     Proyield=str(driver.find_element_by_css_selector(targetyieldint).text)+str(driver.find_element_by_css_selector(targetyielddec).text)+"%"
     Prourl=driver.find_element_by_css_selector(targeturl).get_attribute("href")
     html=get_page_content_str(Prourl)
     #调用前面定义的抓取网页源码函数来度详情页URL
     doc=Pq(html)
     deadline=doc("#time > span").text()
     loan=doc("body > section.bd.m_b_20.clearfix > div.loan_profile.mid.clearfix > div.bid_info > div.main.clearfix > div.con.con2.b_r_0 > div.conts > span").text()
     #通过pyquery 来过滤获得的网页源码，根据内容的css Path 来获取指定内容
     ws.cell(row=j,column=1).value=Proname
     ws.cell(row=j,column=2).value=Proyield
     ws.cell(row=j,column=3).value=deadline
     ws.cell(row=j,column=4).value=loan
     #写入excel单元格
     print(Proname+"-"+Proyield+"-"+deadline+"-"+loan)
     #在控制台输出结果
  nextpage=driver.find_element_by_link_text("下一页")
  #该页抓取完成模拟浏览器点击下一页
w.save('E:\pythontest\qnzdetail.xlsx')
#保存到本地excel文件