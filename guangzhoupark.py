__author__ = 'vincent'
# -*- coding:utf-8 -*-
from openpyxl import Workbook
import urllib
from pyquery import PyQuery as Pq
from selenium import webdriver
import time
import StrUtil

w= Workbook()
#创建excel工作薄
sheetnames = w.get_sheet_names()
ws= w.get_sheet_by_name(sheetnames[0])
#创建Excel工作表

def get_page_content_str( url):
    time.sleep(1)

    try:
            print("现在开始抓取" + url)
            headers = {'User-Agent': 'Mozilla/5.0 (Windows NT 6.1; WOW64; rv:23.0) Gecko/20100101 Firefox/23.0'}
            #伪装浏览器
            request = urllib.request.Request(url=url, headers=headers)
            #构造请求
            m_fp = urllib.request.urlopen(request, timeout=500)
            #访问网站获取源码
            html_str = m_fp.read().decode('utf-8')
            #读取源码，该网站使用的编码方式是utf-8

            return html_str
    except Exception:
            print(Exception)
            try:
                 request = urllib.request.Request(url=url, headers=headers)
                 m_fp = urllib.request.urlopen(request, timeout=500)
                 html_str= m_fp.read().decode('gbk')
                 return html_str
            except Exception:
             print(Exception)
             try:
                 request = urllib.request.Request(url=url, headers=headers)
                 m_fp = urllib.request.urlopen(request, timeout=500)
                 html_str= m_fp.read().decode('gb2312')
                 return html_str
             except Exception:
                      print(Exception)
            m_fp.close()
#定义抓取网页源码函数

baseurl="http://xzl.zhaoshang800.com/offices-0--0-4-p-{}.html"
#目标网站URL
j=1
ws.cell(row=1,column=1).value="区域"
ws.cell(row=1,column=2).value="地址"
ws.cell(row=1,column=3).value="面积"
ws.cell(row=1,column=4).value="级别"
ws.cell(row=1,column=5).value="园区名"
ws.cell(row=1,column=6).value="简介"
#设置excel单元格表头
for i in range (1,29):
    url=baseurl.format(i)
    print("现在开始抓取第{}页".format(i))
    html_str=get_page_content_str(url)
    if html_str==None:
         try:
            headers = {'User-Agent': 'Mozilla/5.0 (Windows NT 6.1; WOW64; rv:23.0) Gecko/20100101 Firefox/23.0'}
            #伪装浏览器
            request = urllib.request.Request(url=url, headers=headers)
            #构造请求
            m_fp = urllib.request.urlopen(request, timeout=500)
            #访问网站获取源码
            html_str = m_fp.read().decode('utf-8')
         except Exception:
            print(Exception)
            if html_str==None:
             try:
              headers = {'User-Agent': 'Mozilla/5.0 (Windows NT 6.1; WOW64; rv:23.0) Gecko/20100101 Firefox/23.0'}
              #伪装浏览器
              request = urllib.request.Request(url=url, headers=headers)
              #构造请求
              m_fp = urllib.request.urlopen(request, timeout=500)
              #访问网站获取源码
              html_str = m_fp.read().decode('gbk')
             except Exception:
              print(Exception)
    doc=Pq(html_str)
    parks=doc("body > div.W1000.rexzl > div.FL > ul > li")
    for park in parks:
        j=j+1
        parkname=Pq(park)("dl > dt > a").text()
        area=Pq(park)("dl > dd:nth-child(3) > div > span:nth-child(1)").text()
        area=StrUtil.substr(area,"：","'")
        adress=Pq(park)(" dl > dd:nth-child(3) > div > span.pla").text()
        adress=StrUtil.substr(adress,"：","'")
        href="http://xzl.zhaoshang800.com"+Pq(park)(" dl > dt > a").attr("href")
        html_detail=get_page_content_str(href)
        doc_detail=Pq(html_detail)
        acr=doc_detail("body > div.W1000.rexzlxx > div.FL > div.txt > table >  tr:nth-child(6) > td:nth-child(2)").text()
        acr=StrUtil.substr(acr,"：","'")
        level=doc_detail("body > div.W1000.rexzlxx > div.FL > div.txt > table >  tr:nth-child(3) > td:nth-child(1)").text()
        level=StrUtil.substr(level,"：","'")
        parkdetail=doc_detail("body > div.W1000.rexzlxx > div.FL > div.txt>div").text()
        if parkdetail=="":
            parkdetail=doc_detail("body > div.W1000.rexzlxx > div.FL > div.txt").text()
            substr=doc_detail("body > div.W1000.rexzlxx > div.FL > div.txt > table").text()
            parkdetail=StrUtil.substr(parkdetail,substr,"'")
        ws.cell(row=j,column=1).value=area
        ws.cell(row=j,column=2).value=adress
        ws.cell(row=j,column=3).value=acr
        ws.cell(row=j,column=4).value=level
        ws.cell(row=j,column=5).value=parkname
        ws.cell(row=j,column=6).value=parkdetail
        #写入excel单元格
        print(area+"-"+adress+"-"+acr+"-"+level+"-"+parkname)
        #在控制台输出结果
w.save(r'D:\pythontest\园区\广州-写字楼.xlsx')
#保存到本地excel文件