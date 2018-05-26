__author__ = 'vincent'
# -*- coding:utf-8 -*-
from openpyxl import Workbook
import urllib
from pyquery import PyQuery as Pq
from selenium import webdriver
import time
import StrUtil

w= Workbook()
#创建excel工作薄
ws= w.create_sheet()
#创建Excel工作表

def get_page_content_str( url):
    time.sleep(1)

    try:
            print("现在开始抓取" + url)
            headers = {'User-Agent': 'Mozilla/5.0 (Windows NT 6.1; WOW64; rv:23.0) Gecko/20100101 Firefox/23.0'}
            #伪装浏览器
            request = urllib.request.Request(url=url, headers=headers)
            #构造请求
            m_fp = urllib.request.urlopen(request, timeout=500)
            #访问网站获取源码
            html_str = m_fp.read().decode('utf-8')
            #读取源码，该网站使用的编码方式是utf-8
            if html_str == None:
                html_str= m_fp.read().decode('gbk')
                if html_str == None:
                    html_str = m_fp.read().decode('utf-8')
            m_fp.close()
            return html_str
    except Exception:
            print(Exception)
#定义抓取网页源码函数

baseurl="http://www.ebeitech.com/index.php?f=news&p={}"
#目标网站URL
j=1
ws.cell(row=1,column=1).value="标题"
ws.cell(row=1,column=2).value="合作方"
ws.cell(row=1,column=3).value="时间"
ws.cell(row=1,column=4).value="项目内容"
#设置excel单元格表头
for i in range (1,10):
    url=baseurl.format(i)
    print("现在开始抓取第{}页".format(i))
    html_str=get_page_content_str(url)
    if html_str==None:
         try:
            headers = {'User-Agent': 'Mozilla/5.0 (Windows NT 6.1; WOW64; rv:23.0) Gecko/20100101 Firefox/23.0'}
            #伪装浏览器
            request = urllib.request.Request(url=url, headers=headers)
            #构造请求
            m_fp = urllib.request.urlopen(request, timeout=500)
            #访问网站获取源码
            html_str = m_fp.read().decode('utf-8')
         except Exception:
            print(Exception)
            if html_str==None:
             try:
              headers = {'User-Agent': 'Mozilla/5.0 (Windows NT 6.1; WOW64; rv:23.0) Gecko/20100101 Firefox/23.0'}
              #伪装浏览器
              request = urllib.request.Request(url=url, headers=headers)
              #构造请求
              m_fp = urllib.request.urlopen(request, timeout=500)
              #访问网站获取源码
              html_str = m_fp.read().decode('gbk')
             except Exception:
              print(Exception)
    doc=Pq(html_str)
    parks=doc("#serve_one")
    for park in parks:
        j=j+1
        parktime=Pq(park)(" span:nth-child(1)").text()
        parkurl=str(Pq(park)(" #serve_one_title > a").attr("href"))
        parkurl="http://www.ebeitech.com/"+parkurl
        detailhtml=get_page_content_str(parkurl)

        if detailhtml==None:
         try:
            headers = {'User-Agent': 'Mozilla/5.0 (Windows NT 6.1; WOW64; rv:23.0) Gecko/20100101 Firefox/23.0'}
            #伪装浏览器
            request = urllib.request.Request(url=parkurl, headers=headers)
            #构造请求
            m_fp = urllib.request.urlopen(request, timeout=500)
            #访问网站获取源码
            detailhtml = m_fp.read().decode('utf-8')
         except Exception:
            print(Exception)
            if detailhtml==None:
             try:
              headers = {'User-Agent': 'Mozilla/5.0 (Windows NT 6.1; WOW64; rv:23.0) Gecko/20100101 Firefox/23.0'}
              #伪装浏览器
              request = urllib.request.Request(url=parkurl, headers=headers)
              #构造请求
              m_fp = urllib.request.urlopen(request, timeout=500)
              #访问网站获取源码
              detailhtml = m_fp.read().decode('iso8859-1').encode("iso8859-1")
             except Exception:
              print(Exception)
        detail=Pq(detailhtml)
        proname=detail("#wz_title").text()
        prourl="http://www.ebeitech.com/"+detail("#detail_text > iframe").attr("src")
        prodetail=get_page_content_str(prourl)
        prodetail=Pq(prodetail)
        adress=prodetail("body ").text()
        busname=prodetail("body > section> section > span > span.tn-page-bg-color.ng-scope").text()
        ws.cell(row=j,column=1).value=proname
        ws.cell(row=j,column=2).value=busname
        ws.cell(row=j,column=3).value=parktime
        ws.cell(row=j,column=4).value=adress
        #写入excel单元格
       # print(proname+"-"+"-"+busname+"-"+time+"-"+adress)
        #在控制台输出结果
w.save(r'D:\pythontest\园区\一碑科技新闻.xlsx')
#保存到本地excel文件