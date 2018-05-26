__author__ = 'vincent'
# -*- coding:utf-8 -*-
from openpyxl import Workbook
import urllib
from pyquery import PyQuery as Pq
from selenium import webdriver
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.common.keys import Keys
import time
import StrUtil

w= Workbook()
#创建excel工作薄
ws= w.create_sheet(0)
#创建Excel工作表

def get_page_content_str( url):
    time.sleep(1)

    try:
            print("现在开始抓取" + url)
            headers = {'User-Agent': 'Mozilla/5.0 (Windows NT 6.1; WOW64; rv:23.0) Gecko/20100101 Firefox/23.0'}
            #伪装浏览器
            request = urllib.request.Request(url=url, headers=headers)
            #构造请求
            m_fp = urllib.request.urlopen(request, timeout=500)
            #访问网站获取源码
            html_str = m_fp.read().decode('utf-8')
            #读取源码，该网站使用的编码方式是utf-8
            if html_str == None:
                html_str= m_fp.read().decode('gbk')
                if html_str == None:
                    html_str = m_fp.read().decode('gb2312')
            m_fp.close()
            return html_str
    except Exception:
            print(Exception)
#定义抓取网页源码函数

j=1
ws.cell(row=1,column=1).value="商品"
ws.cell(row=1,column=2).value="价格"
ws.cell(row=1,column=3).value="销量"
ws.cell(row=1,column=3).value="传送"
#设置excel单元格表头
driver=webdriver.Chrome("D://chromedriver.exe")

#baseurl="https://chi.taobao.com/itemlist/huichi2014.htm?cat=50002766%2C50035978%2C50008825%2C50042258%2C50103282%2C50103280%2C50106154%2C50108542&user_type=0&at=45634&viewIndex=1&as=0&spm=a219e.7769888.a2151qd.127.nb3TZj&atype=b&style=grid&q=%E4%BC%91%E9%97%B2%E9%9B%B6%E9%A3%9F&same_info=1&tid=0&isnew=2&_input_charset=utf-8"
#baseurl="https://www.taobao.com/"
baseurl="https://s.taobao.com/search?q=%E8%9C%9C%E9%A5%AF%E6%9E%9C%E5%B9%B2&imgfile=&js=1&stats_click=search_radio_all%3A1&initiative_id=staobaoz_20160408&ie=utf8"
driver.get(baseurl)
time.sleep(2)
driver.find_element_by_css_selector("#J_relative > div.sort-row > div > ul > li:nth-child(3) > a").click()
#driver.find_element_by_css_selector("body > div.cup.J_Cup > div.screen.J_Screen > div.sa-sub > div > div > ul > li:nth-child(9) > span > a:nth-child(3)").click()
#time.sleep(3)
#all_handles = driver.window_handles
#driver.switch_to.window(all_handles[1])
time.sleep(3)

for page in range(1,20):
    for k in range(1,45):
        j=j+1
        snackname="#mainsrp-itemlist > div > div > div:nth-child(1) > div:nth-child({})>div:nth-child(2)>div:nth-child(2)>a".format(k)
        name=driver.find_element_by_css_selector(snackname).text
        snname=str(name)
        snackprice="#mainsrp-itemlist > div > div > div:nth-child(1) > div:nth-child({})>div:nth-child(2)>div:nth-child(1)>div:nth-child(1)>strong".format(k)
        price=driver.find_element_by_css_selector(snackprice).text
        snacksales="#mainsrp-itemlist > div > div > div:nth-child(1) > div:nth-child({})>div:nth-child(2)>div:nth-child(1)>div:nth-child(2)".format(k)
        sales=driver.find_element_by_css_selector(snacksales).text
        snackhref="#mainsrp-itemlist > div > div > div:nth-child(1) > div:nth-child({}) > div.ctx-box.J_MouseEneterLeave.J_IconMoreNew > div.row.row-2.title>a"
        href=driver.find_element_by_css_selector(snackname).get_attribute("href")

        ws.cell(row=j,column=1).value=name
        ws.cell(row=j,column=2).value=price
        ws.cell(row=j,column=3).value=sales
        ws.cell(row=j,column=4).value=href

        print(name+"-"+price+"-"+sales+"-"+href)

    driver.find_element_by_css_selector("#mainsrp-pager > div > div > div > ul > li.item.next > a > span:nth-child(1)").click()
    time.sleep(2)
        #写入excel单元格

        #在控制台输出结果
w.save('D:\pythontest\电商\淘宝\蜜饯果干.xlsx')
#保存到本地excel文件