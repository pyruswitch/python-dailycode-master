__author__ = 'vincent'
from openpyxl import Workbook
from openpyxl.reader.excel import load_workbook
import mysql.connector
import operator
import  datetime
from time import sleep
from tqdm import tqdm
import time
import xlrd





w= Workbook()
#创建excel工作薄
sheetnames = w.get_sheet_names()
ws = w.get_sheet_by_name(sheetnames[0])
wsnew = w.create_sheet(0)
wsnew.title="新用户原始数据"
ws.title="原始数据"
ws.cell(row=1,column=1).value="用户ID"
ws.cell(row=1,column=2).value="联系方式"
ws.cell(row=1,column=3).value="店铺名"
ws.cell(row=1,column=4).value="渠道"
ws.cell(row=1,column=5).value="频次"
ws.cell(row=1,column=6).value="总额"
ws.cell(row=1,column=7).value="均单价"

user="ning.wei12"
pwd="wn3633"
host="bizdb.zuolin.com"
db="ehbiz"
port="18306"

cnx=mysql.connector.connect(user=user,password=pwd,database=db,host=host,port=port)
cursor=cnx.cursor()
userweek_sql='''SELECT
 a.`buyer_no`,  a.`buyer_phone`,c.`shop_name`,a.`pay_type`,COUNT(a.`buyer_no`),SUM(a.`paid_total`),AVG(a.`paid_total`)
 FROM
 `tbl_order`a
 LEFT JOIN
 `tbl_shop_info`c
 ON c.`shop_no`=a.`shop_no`
 WHERE
 a.`payment_time`>"2016-11-21"
 and  a.`payment_time`<"2016-11-26"
 AND a.`shop_no`IN("14622779991551986272","14477417463124576784","14667588471072254231","14721933954264944406")
 GROUP BY a.`buyer_no`
 ORDER BY COUNT(a.`buyer_no`)DESC'''
cursor.execute(userweek_sql)
alldata=cursor.fetchall()
alldata.sort(key=operator.itemgetter(5),reverse=True)
Arows=1
paytype=[]
print("读取购买记录数据 ……")
for basicdata in alldata:
    Arows=Arows+1
    for uvalue in range(0,7):
        uservalue=basicdata[uvalue]
        if uvalue==3:
            if uservalue==1:
                uservalue="线上"
            else:uservalue="线下"
        ws.cell(row=Arows,column=uvalue+1).value=uservalue
#写入Excel表
wt = w.create_sheet(0)
wt.title="总体情况"
allcount=len(alldata)
wt.cell(row=1,column=1).value="总消费用户 "
wt.cell(row=1,column=2).value=allcount
paytypeli=[]
for type in ws.columns[3]:
    type=type.value
    if type=="渠道":
        continue
    paytypeli.append(type)
paytypes=set(paytypeli)
wtcolumn=2
for paytype in paytypes:
    wt.cell(row=wtcolumn,column=1).value=paytype
    wt.cell(row=wtcolumn,column=2).value=paytypeli.count(paytype)
    wtcolumn=wtcolumn+1
shopuserli=[]
for Suser in ws.columns[2]:
    Suser=Suser.value
    if  Suser=="店铺名":
        continue
    shopuserli.append(Suser)
shopusers=set(shopuserli)
for shopuser in shopusers:
    wt.cell(row=wtcolumn,column=1).value=shopuser
    wt.cell(row=wtcolumn,column=2).value=shopuserli.count(shopuser)
    wtcolumn=wtcolumn+1
#--------------新用户-----------------------------------------------------#

wsnew.cell(row=1,column=1).value="用户ID"
wsnew.cell(row=1,column=2).value="联系方式"
wsnew.cell(row=1,column=3).value="店铺名"
wsnew.cell(row=1,column=4).value="渠道"
wsnew.cell(row=1,column=5).value="频次"
wsnew.cell(row=1,column=6).value="总额"
wsnew.cell(row=1,column=7).value="均单价"
wsnew.cell(row=1,column=8).value="首次消费时间 "
newuser_sql='''SELECT
 a.`buyer_no`,  a.`buyer_phone`,c.`shop_name`,a.`pay_type`,COUNT(a.`buyer_no`),SUM(a.`paid_total`),AVG(a.`paid_total`),MIN(DATE_FORMAT(a.`payment_time`,'%Y-%m-%d'))
 FROM
 `tbl_order`a
 LEFT JOIN
 `tbl_shop_info`c
 ON c.`shop_no`=a.`shop_no`
 WHERE
 a.`payment_time`>"2016-11-20"
 and  a.`payment_time`<"2016-11-27"
 AND a.`shop_no`IN("14622779991551986272","14477417463124576784","14667588471072254231","14721933954264944406")
 AND a.`buyer_no`NOT IN (SELECT
 a.`buyer_no`
 FROM
 `tbl_order`a
 LEFT JOIN
 `tbl_shop_info`c
 ON c.`shop_no`=a.`shop_no`
 WHERE
 a.`payment_time`<"2016-11-20"
 AND a.`shop_no`IN("14622779991551986272","14477417463124576784","14667588471072254231","14721933954264944406"))
 GROUP BY a.`buyer_no`
 ORDER BY COUNT(a.`buyer_no`)DESC'''
cursor.execute(newuser_sql)
allnewdata=cursor.fetchall()
allnewdata.sort(key=operator.itemgetter(5),reverse=True)
Arows=1
paytype=[]
print("读取新用户购买记录数据 ……")
for basicdata in allnewdata:
    Arows=Arows+1
    for uvalue in range(0,8):
        uservalue=basicdata[uvalue]
        if uvalue==3:
            if uservalue==1:
                uservalue="线上"
            else:uservalue="线下"
        wsnew.cell(row=Arows,column=uvalue+1).value=uservalue
#写入Excel表
allnewcount=len(allnewdata)
wt.cell(row=1,column=4).value="新消费用户 "
wt.cell(row=1,column=5).value=allnewcount
paytypeli=[]
for type in wsnew.columns[3]:
    type=type.value
    if type=="渠道":
        continue
    paytypeli.append(type)
paytypes=set(paytypeli)
wtcolumn=2
for paytype in paytypes:
    wt.cell(row=wtcolumn,column=4).value=paytype
    wt.cell(row=wtcolumn,column=5).value=paytypeli.count(paytype)
    wtcolumn=wtcolumn+1
shopuserli=[]
for Suser in wsnew.columns[2]:
    Suser=Suser.value
    if  Suser=="店铺名":
        continue
    shopuserli.append(Suser)
shopusers=set(shopuserli)
for shopuser in shopusers:
    wt.cell(row=wtcolumn,column=4).value=shopuser
    wt.cell(row=wtcolumn,column=5).value=shopuserli.count(shopuser)
    wtcolumn=wtcolumn+1
#--------------周趋势-----------------------------------------------------#
wuser=w.create_sheet(0)
wuser.title="总体用户趋势"
wuser.cell(row=1,column=1).value="日期"
dateuser_sql='''SELECT
DATE_FORMAT(a.`payment_time`,'%Y-%m-%d'),b.`shop_name`,a.`buyer_no`,a.`pay_type`,COUNT(*)
FROM
`tbl_order` a
LEFT JOIN
`tbl_shop_info`b
ON b.`shop_no`=a.`shop_no`
WHERE
a.`payment_time`>"2016-11-21"
AND a.`shop_no`IN("14622779991551986272","14477417463124576784","14667588471072254231","14721933954264944406")
GROUP BY DATE_FORMAT(a.`payment_time`,'%Y-%m-%d'),a.`buyer_no`
'''
cursor.execute(dateuser_sql)
alldausdata=cursor.fetchall()
userdatelist=[t[0]for t in alldausdata]
userdlist=set(userdatelist)
userdlist=list(userdlist)
userdlist.sort()
#日期列表
shopdatelist=[t[1] for t in alldausdata ]
shopdlist=set(shopdatelist)
#店铺列表
typedatelist=[t[3]for t in alldausdata]
typedlist=set(typedatelist)
#渠道列表
daterow=1
#每日渠道用户
for usdate in userdlist:
     daterow=daterow+1
     wuser.cell(row=daterow,column=1).value=usdate
     typecol=1
     shopcol=3
     for ustype in typedlist:
         typecol=typecol+1
         if ustype==1:
             wuser.cell(row=1,column=typecol).value="线上"
         elif ustype==3:
             wuser.cell(row=1,column=typecol).value="线下"
         else:
             continue

         ustypecount=0
         for typeele in alldausdata:
             if typeele[0]==usdate:
                 if typeele[3]==ustype:
                    ustypecount=ustypecount+1
         wuser.cell(row=daterow,column=typecol).value=ustypecount
#每日店铺用户
     for usshop in shopdlist:
        shopcol=shopcol+1
        wuser.cell(row=1,column=shopcol).value=usshop
        usshopcount=0
        for shopele in alldausdata:
             if shopele[0]==usdate:
                 if shopele[1]==usshop:
                     usshopcount=usshopcount+1
        wuser.cell(row=daterow,column=shopcol).value=usshopcount
#------------------------线上站点用户趋势------------------------------------
wuseronline=w.create_sheet(0)
wuseronline.title="线上站点用户趋势"
wuseronline.cell(row=1,column=1).value="日期"
onlinedaterow=1
for usdate in userdlist:
     onlinedaterow=onlinedaterow+1
     wuseronline.cell(row=onlinedaterow,column=1).value=usdate
     shopcol=1
#每日店铺用户
     for usshop in shopdlist:
        shopcol=shopcol+1
        wuseronline.cell(row=1,column=shopcol).value=usshop
        usshopcount=0
        for shopele in alldausdata:
             if shopele[3]!=1:
                 continue
             if shopele[0]==usdate:
                 if shopele[1]==usshop:
                     usshopcount=usshopcount+1
        wuseronline.cell(row=onlinedaterow,column=shopcol).value=usshopcount

#-----------------------线下总体用户趋势------------------------------------
wuserline=w.create_sheet(0)
wuserline.title="线下站点用户趋势"
wuserline.cell(row=1,column=1).value="日期"
linedaterow=1
for usdate in userdlist:
     linedaterow=linedaterow+1
     wuserline.cell(row=linedaterow,column=1).value=usdate
     shopcol=1
#每日店铺用户
     for usshop in shopdlist:
        shopcol=shopcol+1
        wuserline.cell(row=1,column=shopcol).value=usshop
        usshopcount=0
        for shopele in alldausdata:
             if shopele[3]!=3:
                 continue
             if shopele[0]==usdate:
                 if shopele[1]==usshop:
                     usshopcount=usshopcount+1
        wuserline.cell(row=linedaterow,column=shopcol).value=usshopcount
#----------------------------新增用户趋势———————————————————
wsdaynew=w.create_sheet(0)
wsdaynew.title="用户新增趋势"
wsdaynew.cell(row=1,column=1).value="日期"
newdaylist=[day[7]for day in allnewdata]
newdaylist=set(newdaylist)
newdaylist=list(newdaylist)
newdaylist.sort()
newdayshops=[d[2]for d in allnewdata]
newdayshops=set(newdayshops)
newdaytype=[d[3]for d in allnewdata]
newdaytype=set(newdaytype)
dayrow=1
for nday in newdaylist:
     dayrow=dayrow+1
     wsdaynew.cell(row=dayrow,column=1).value=nday
     newtycol=1
     newshopcol=3
     for newtype in newdaytype:
         newtycol=newtycol+1
         if newtype==1:
             wsdaynew.cell(row=1,column=newtycol).value="线上"
         elif newtype==3:
             wsdaynew.cell(row=1,column=newtycol).value="线下"
         else:
             continue

         newtypecount=0
         for nwwtypeele in allnewdata:
             if nwwtypeele[7]==nday:
                 if nwwtypeele[3]==newtype:
                    newtypecount=newtypecount+1
         wsdaynew.cell(row=dayrow,column=newtycol).value=newtypecount
#每日店铺用户
     for newdayshop in newdayshops:
        newshopcol=newshopcol+1
        wsdaynew.cell(row=1,column=newshopcol).value=newdayshop
        newshopcount=0
        for newshopele in allnewdata:
             if newshopele[7]==nday:
                 if newshopele[2]==newdayshop:
                     newshopcount=newshopcount+1
        wsdaynew.cell(row=dayrow,column=newshopcol).value=newshopcount
#----------------------------线上站点新增用户趋势———————————————————
wsdaynewonline=w.create_sheet(0)
wsdaynewonline.title="线上站点用户新增趋势"
wsdaynewonline.cell(row=1,column=1).value="日期"
onlinedayrow=1
for nday in newdaylist:
     onlinedayrow=onlinedayrow+1
     wsdaynewonline.cell(row=onlinedayrow,column=1).value=nday
     newtycol=1
     newshopcol=1
#每日店铺用户
     for newdayshop in newdayshops:
        newshopcol=newshopcol+1
        wsdaynewonline.cell(row=1,column=newshopcol).value=newdayshop
        newshopcount=0
        for newshopele in allnewdata:
             if newshopele[3]!=1:
                 continue
             if newshopele[7]==nday:
                 if newshopele[2]==newdayshop:
                     newshopcount=newshopcount+1
        wsdaynewonline.cell(row=onlinedayrow,column=newshopcol).value=newshopcount

#----------------------------线下站点新增用户趋势———————————————————
wsdaynewline=w.create_sheet(0)
wsdaynewline.title="线下站点用户新增趋势"
wsdaynewline.cell(row=1,column=1).value="日期"
linedayrow=1
for nday in newdaylist:
     linedayrow=linedayrow+1
     wsdaynewline.cell(row=linedayrow,column=1).value=nday
     newtycol=1
     newshopcol=1
#每日店铺用户
     for newdayshop in newdayshops:
        newshopcol=newshopcol+1
        wsdaynewline.cell(row=1,column=newshopcol).value=newdayshop
        newshopcount=0
        for newshopele in allnewdata:
             if newshopele[3]!=3:
                 continue
             if newshopele[7]==nday:
                 if newshopele[2]==newdayshop:
                     newshopcount=newshopcount+1
        wsdaynewline.cell(row=linedayrow,column=newshopcol).value=newshopcount

w.save(r"H:\upupup\电商后台\自动化统计\某时段用户1107-1113.xlsx")
cursor.close()
cnx.commit()
cnx.close()