__author__ = 'vincent'
import jieba
from operator import itemgetter
from math import log
import os
import os.path
import jieba.posseg as pseg
from snownlp import  SnowNLP as snlp
from openpyxl import Workbook



rootdir=r"H:\upupup\电商后台\主题挖掘\文本挖掘\黄伟文"
TF={}
IDF={}
count=0.0
i=0
countwords=0

w= Workbook()
#创建excel工作薄
sheetnames = w.get_sheet_names()
ws= w.get_sheet_by_name(sheetnames[0])
j=1
#创建Excel工作表
for root,dirs, filenames in os.walk(rootdir):
    for filename in filenames:
        i=i+1
        filepath=r"H:/upupup/电商后台/主题挖掘/文本挖掘/黄伟文/{}".format(filename)
        nowdic={}
        sumwords=0
        linecount=0
        emotion=0
        try :
            with open(filepath) as fr:
                for line in fr.readlines():
                    line =line.strip()
                    if str(line)=="":
                        continue
                    #print(filename,line)
                    try:
                     str_line=snlp(str(line))
                    except IOError as err:
                      continue
                    emotion=str_line.sentiments+emotion
                    linecount=linecount+1
                if linecount==0 :
                    continue
                lyricemotion=float(emotion/linecount)
                print(filename+":"+str(lyricemotion))
                ws.cell(row=j,column=1).value=lyricemotion
                ws.cell(row=j,column=2).value=filename
                j=j+1
        except IOError as err:
            pass
w.save(r'H:\upupup\电商后台\主题挖掘\文本挖掘\黄伟文歌词文本情绪.xlsx')