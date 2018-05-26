__author__ = 'Administrator'

# -*- coding:utf-8 -*-
from openpyxl import Workbook
import urllib
from pyquery import PyQuery as Pq
from selenium import webdriver
import time
import StrUtil
from openpyxl.reader.excel import load_workbook
import win32ui
import  os
from lxml import etree

dlg = win32ui.CreateFileDialog(1) # 1表示打开文件对话框
dlg.SetOFNInitialDir('E:/Python') # 设置打开文件对话框中的初始显示目录
dlg.DoModal()

filename = dlg.GetPathName() # 获取选择的文件名称
print(filename)

#wb = load_workbook(r'H:\upupup\企业新闻\关键词.xlsx')
wb = load_workbook(filename)
sheetnames = wb.get_sheet_names()
wo = wb.get_sheet_by_name(sheetnames[0])
w= Workbook()
#创建excel工作薄
ws= w.create_sheet(0)

#创建Excel工作表
baseurl="http://news.baidu.com/ns?ct=1&rn=20&ie=utf-8&bs="
sufurl="&rsv_bp=1&sr=0&cl=2&f=8&prevct=no&tn=news&word="
keywords=[]
for ele in range(2,wo.get_highest_row()+1):
    keywords.append(wo.cell(row=ele,column=1).value)
weblist=["163.com/",
"xiaofei.china.com.cn/",
"ifeng.com/",
"news.dichan.sina.com.cn/",
"sina.com.cn/",
"finance.qq.com/",

"epaper.gmw.cn/",
"iof.hexun.com/",
"caijing.com.cn/",
"szsb.sznews.com/",
"news.hexun.com/",
"gb.cri.cn/",
"tech.qq.com/",
"tech.hnr.cn/",
"finance.qq.com/",
"gmw.cn/",
"sasac.gov.cn/",
"chinabyte.com/300",
"sohu.com/",
"cnstock.com/"]
idlist=[
"endText",
"Article",
"main_content",
 "divContent",
"artibody",
"Cnt-Main-Article-QQ",

"articleContent",
"art_contextBox",
"the_content",
"ozoom",
"artibody",
"ccontent",
"Cnt-Main-Article-QQ",
"txt_left",
"Cnt-Main-Article-QQ",
"contentMain",
"con_con",
"logincontent",
"contentText",
"qmt_content_div"]
def get_page_content_str( url):
    time.sleep(1)

    try:
            print("现在开始抓取" + url)
            headers = {'User-Agent': 'Mozilla/5.0 (Windows NT 6.1; WOW64; rv:23.0) Gecko/20100101 Firefox/23.0'}
            #伪装浏览器
            request = urllib.request.Request(url=url, headers=headers)
            #构造请求
            m_fp = urllib.request.urlopen(request, timeout=500)
            #访问网站获取源码
            html_str = m_fp.read().decode('utf-8')
            #读取源码，该网站使用的编码方式是utf-8

            return html_str
    except Exception:
            print(Exception)
            try:
                 request = urllib.request.Request(url=url, headers=headers)
                 m_fp = urllib.request.urlopen(request, timeout=500)
                 html_str= m_fp.read().decode('gbk')
                 return html_str
            except Exception:
             print(Exception)
             try:
                 request = urllib.request.Request(url=url, headers=headers)
                 m_fp = urllib.request.urlopen(request, timeout=500)
                 html_str= m_fp.read().decode('gb2312')
                 return html_str
             except Exception:
                      print(Exception)
            m_fp.close()
#定义抓取网页源码函数


#目标网站URL
j=1
ws.cell(row=1,column=1).value="关键词"
ws.cell(row=1,column=2).value="标题"
ws.cell(row=1,column=3).value="网站"
ws.cell(row=1,column=4).value="日期"
ws.cell(row=1,column=5).value="摘要"
ws.cell(row=1,column=6).value="链接"
ws.cell(row=1,column=7).value="原文"
ws.cell(row=1,column=8).value="源码"
for keyword in keywords:
    word=keyword.encode("utf-8")
    word=str(word)
    word=StrUtil.substr(word,"b'","'")
    word=word.replace(r"\x","%")
    url=baseurl+word+sufurl+word
    print("现在开始抓取"+keyword+"的新闻")
    html=get_page_content_str(url)
    doc=Pq(html)
    news=doc("#content_left > div:nth-child(3)>div")
    if len(news) ==0:
        j=j+1
        ws.cell(row=j,column=1).value=keyword
        ws.cell(row=j,column=2).value="无相关内容"
        continue
    else:
     for new in news:

        title=Pq(new)("h3>a").text()
        link=Pq(new)("h3>a").attr("href")
        source=Pq(new)("div>div.c-span18.c-span-last >p").text()
        date=StrUtil.substr(source,' ','"')
        if date=="":
            source=Pq(new)("div>p").text()
            date=StrUtil.substr(source,' ','"')
            web=StrUtil.substr(source,'"',' ')
            remark=Pq(new)("div ").text()
            remark=StrUtil.substr(remark,date,'...')
        else:
         web=StrUtil.substr(source,'"',' ')
         remark=Pq(new)("div > div.c-span18.c-span-last").text()
         remark=StrUtil.substr(remark,date,'...')
        for i in range(0, len(weblist)):
            if weblist[i] in link :
                j=j+1
                webhtml=get_page_content_str(link)
                webdoc=Pq(webhtml)
                textid=idlist[i]
                id="#"+textid

                #print(id)
                webcode=webdoc(id)
                #print(str(webcode))
                webtext=webcode.text()
            else:
                continue
            ws.cell(row=j,column=1).value=keyword
            ws.cell(row=j,column=2).value=title
            ws.cell(row=j,column=3).value=web
            ws.cell(row=j,column=4).value=date
            ws.cell(row=j,column=5).value=remark
            ws.cell(row=j,column=6).value=link
            ws.cell(row=j,column=7).value=webtext
            ws.cell(row=j,column=8).value=str(webcode)
            print(keyword+"-"+title+"-"+web+"-"+date)
            break
#w.save(r'H:\upupup\企业新闻\园区企业.xlsx')
file_path = os.path.split(filename)
filepath=file_path[0]
w.save(str(filepath)+r"\news.xlsx")

