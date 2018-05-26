__author__ = 'vincent'
# -*- coding:utf-8 -*-
from openpyxl import Workbook
import urllib
from pyquery import PyQuery as Pq
from selenium import webdriver
import time
import StrUtil

w= Workbook()
#创建excel工作薄
ws= w.create_sheet()
#创建Excel工作表

def get_page_content_str( url):
    time.sleep(1)

    try:
            print("现在开始抓取" + url)
            headers = {'User-Agent': 'Mozilla/5.0 (Windows NT 6.1; WOW64; rv:23.0) Gecko/20100101 Firefox/23.0'}
            #伪装浏览器
            request = urllib.request.Request(url=url, headers=headers)
            #构造请求
            m_fp = urllib.request.urlopen(request, timeout=500)
            #访问网站获取源码
            html_str = m_fp.read().decode('utf-8')
            #读取源码，该网站使用的编码方式是utf-8

            return html_str
    except Exception:
            print(Exception)
            try:
                 request = urllib.request.Request(url=url, headers=headers)
                 m_fp = urllib.request.urlopen(request, timeout=500)
                 html_str= m_fp.read().decode('gbk')
                 return html_str
            except Exception:
             print(Exception)
             try:
                 request = urllib.request.Request(url=url, headers=headers)
                 m_fp = urllib.request.urlopen(request, timeout=500)
                 html_str= m_fp.read().decode('gb2312')
                 return html_str
             except Exception:
                      print(Exception)
            m_fp.close()
#定义抓取网页源码函数

baseurl="http://gd.cnipai.com/park/?gyyclass=0&keyword=&CY_id=0&city=440100&page={}"
#目标网站URL
j=1
ws.cell(row=1,column=1).value="城市"
ws.cell(row=1,column=2).value="园区名"
ws.cell(row=1,column=3).value="地址"
ws.cell(row=1,column=4).value="类型"
ws.cell(row=1,column=5).value="面积"
ws.cell(row=1,column=6).value="价格"
ws.cell(row=1,column=7).value="简介"
#设置excel单元格表头
for i in range (1,3):
    url=baseurl.format(i)
    print("现在开始抓取第{}页".format(i))
    html_str=get_page_content_str(url)

    doc=Pq(html_str)
    parks=doc("body > div.park_list > div.park_listL > div > div.fact_list")
    for park in parks:
        j=j+1
        city="广州"
        parkname=Pq(park)("  div.fact_incon2 > div.fact_tit > a").text()
        parkurl="http://gd.cnipai.com/"+Pq(park)(" div.fact_incon2 > div.fact_tit > a").attr("href")+"detail/"
        adress=Pq(park)(" div.fact_incon2 > div.fact_add").text()
        #type=Pq(park)("div.fact_incon2 > div.fact_date").text()
        area=Pq(park)(" div.pk_price > div.fact_area").text()
        price=Pq(park)(" div.pk_price > div.fact_inprice > b").text()
        detailhtml=get_page_content_str(parkurl)

        if detailhtml==None:
         try:
            headers = {'User-Agent': 'Mozilla/5.0 (Windows NT 6.1; WOW64; rv:23.0) Gecko/20100101 Firefox/23.0'}
            #伪装浏览器
            request = urllib.request.Request(url=parkurl, headers=headers)
            #构造请求
            m_fp = urllib.request.urlopen(request, timeout=500)
            #访问网站获取源码
            detailhtml = m_fp.read().decode('utf-8')
         except Exception:
            print(Exception)
            if detailhtml==None:
             try:
              headers = {'User-Agent': 'Mozilla/5.0 (Windows NT 6.1; WOW64; rv:23.0) Gecko/20100101 Firefox/23.0'}
              #伪装浏览器
              request = urllib.request.Request(url=parkurl, headers=headers)
              #构造请求
              m_fp = urllib.request.urlopen(request, timeout=500)
              #访问网站获取源码
              detailhtml = m_fp.read().decode('gbk')
             except Exception:
              print(Exception)
        detail=Pq(detailhtml)
        abs=detail("#info_body").text()
        ws.cell(row=j,column=1).value=city
        ws.cell(row=j,column=2).value=parkname
        ws.cell(row=j,column=3).value=adress
        #ws.cell(row=j,column=4).value=type
        ws.cell(row=j,column=5).value=area
        ws.cell(row=j,column=6).value=price
        ws.cell(row=j,column=7).value=abs
        #写入excel单元格
        print(city+"-"+"-"+parkname+"-"+area+"-"+"-"+abs)
        #在控制台输出结果
w.save(r'D:\pythontest\园区\广州2 .xlsx')
#保存到本地excel文件