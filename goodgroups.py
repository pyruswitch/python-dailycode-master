__author__ = 'Administrator'
from openpyxl import Workbook
from openpyxl.reader.excel import load_workbook

wb = load_workbook(r'D:\pythontest\电商\数据处理\商品关联度分析.xlsx')
sheetnames = wb.get_sheet_names()
wo = wb.get_sheet_by_name(sheetnames[0])
w= Workbook()
ws= w.create_sheet(0)
li=[]
ws.cell(row=1,column=1).value ="组合名"
ws.cell(row=1,column=2).value ="组合频次"
i=2
while (i<wo.get_highest_row()):
    print("现在执行到第"+str(i)+"行")
    if wo.cell(row = i,column =1).value==wo.cell(row = i+1,column =1).value:
        group= wo.cell(row = i,column =2).value+"&"+wo.cell(row = i+1,column =2).value
        if i+2<wo.get_highest_row()+1:
          for j in range(i+2,wo.get_highest_row()+1):
            if wo.cell(row = i,column =1).value==wo.cell(row = j,column =1).value:
                group=group+"&"+wo.cell(row = j,column =2).value
            else:
                i=j
                li.append(str(group))
                break
        else:
            i=i+1
            continue
    else:
        i=i+1
        continue
groups=set(li)
r=2
for item in groups:
    ws.cell(row=r,column=1).value =item
    ws.cell(row=r,column=2).value =li.count(item)
    r=r+1
    print(str(item)+":"+str(li.count(item)))
w.save('D:\pythontest\电商\数据处理\goodgroups.xlsx')
