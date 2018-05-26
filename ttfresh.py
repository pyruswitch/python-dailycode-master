__author__ = 'vincent'
# -*- coding:utf-8 -*-
from openpyxl import Workbook
import urllib
from pyquery import PyQuery as Pq
from selenium import webdriver
import time
import StrUtil

w= Workbook()
#创建excel工作薄
ws= w.create_sheet(0)
#创建Excel工作表

def get_page_content_str( url):
    time.sleep(1)

    try:
            print("现在开始抓取" + url)
            headers = {'User-Agent': 'Mozilla/5.0 (Windows NT 6.1; WOW64; rv:23.0) Gecko/20100101 Firefox/23.0'}
            #伪装浏览器
            request = urllib.request.Request(url=url, headers=headers)
            #构造请求
            m_fp = urllib.request.urlopen(request, timeout=500)
            #访问网站获取源码
            html_str = m_fp.read().decode('utf-8')
            #读取源码，该网站使用的编码方式是utf-8
            if html_str == None:
                html_str= m_fp.read().decode('gbk')
                if html_str == None:
                    html_str = m_fp.read().decode('gb2312')
            m_fp.close()
            return html_str
    except Exception:
            print(Exception)
#定义抓取网页源码函数

baseurl="http://www.fruitday.com/prolist/index/277"
#目标网站URL
j=1
ws.cell(row=1,column=1).value="名称"
ws.cell(row=1,column=2).value="规格"
ws.cell(row=1,column=3).value="价格"
#设置excel单元格表头
html_str=get_page_content_str(baseurl)
doc=Pq(html_str)
fruits=doc("body > div.content.typepage.typepage-pad > div.f-list.clearfix > div.leftpart.pull-left > ul > li")
for fruit in fruits:
     j=j+1
     fruitname=Pq(fruit)("div > div.s-info.clearfix").text()
     fruitspec=Pq(fruit)(" div > div.p-operate.clearfix > div.s-kg.clearfix.pull-left > span").text()
     fruitpri=Pq(fruit)(" div > div.s-info.clearfix > span.s-unit.pull-right.font-color").text()
     ws.cell(row=j,column=1).value=fruitname
     ws.cell(row=j,column=2).value=fruitspec
     ws.cell(row=j,column=3).value=fruitpri
     #写入excel单元格
     print(fruitname+"-"+fruitspec+"-"+fruitpri)
     #在控制台输出结果
w.save(r'E:\pythontest\电商\ttfresh.xlsx')
#保存到本地excel文件