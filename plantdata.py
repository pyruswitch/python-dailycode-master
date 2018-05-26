__author__ = 'vincent'

# -*- coding:utf-8 -*-
import urllib
from pyquery import PyQuery as Pq
import StrUtil
import time
import os
from openpyxl import Workbook



def get_page_content_str( url):
    time.sleep(1)

    try:
            print("现在开始抓取" + url)
            headers = {'User-Agent': 'Mozilla/5.0 (Windows NT 6.1; WOW64; rv:23.0) Gecko/20100101 Firefox/23.0'}
            #伪装浏览器
            request = urllib.request.Request(url=url, headers=headers)
            #构造请求
            m_fp = urllib.request.urlopen(request, timeout=500)
            #访问网站获取源码
            html_str = m_fp.read().decode('utf-8')
            #读取源码，该网站使用的编码方式是utf-8

            return html_str
    except Exception:
            print(Exception)
            try:
                 request = urllib.request.Request(url=url, headers=headers)
                 m_fp = urllib.request.urlopen(request, timeout=500)
                 html_str= m_fp.read().decode('gbk')
                 return html_str
            except Exception:
             print(Exception)
             try:
                 request = urllib.request.Request(url=url, headers=headers)
                 m_fp = urllib.request.urlopen(request, timeout=500)
                 html_str= m_fp.read().decode('gb2312')
                 return html_str
             except Exception:
                      print(Exception)
            m_fp.close()
#定义抓取网页源码函数
w= Workbook()
#创建excel工作薄
sheetnames = w.get_sheet_names()
ws= w.get_sheet_by_name(sheetnames[0])
#创建Excel工作表

baseurl="http://www.huamu.cn/pinzhong/search.aspx?page={}"
count=0
namelist=[]
j=1
for i in range(1,691):
    url=baseurl.format(i)
    print("现在开始抓取第{}页".format(i))
    print(url)

    html_str=get_page_content_str(url)
    doc=Pq(html_str)

    names=doc("#searchForm2 > div.bd.tab > table >  tr")
    for name in names:

        plantname=Pq(name)("td:nth-child(2) > a").text()
        if plantname=='':
            continue
        ws.cell(row=j,column=1).value=plantname
        j=j+1
w.save(r'H:\upupup\电商后台\主题挖掘\文本挖掘\植物.xlsx')
