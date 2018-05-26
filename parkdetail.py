__author__ = 'vincent'
# -*- coding:utf-8 -*-
from openpyxl import Workbook
import urllib
from pyquery import PyQuery as Pq
from selenium import webdriver
import time
import StrUtil

w= Workbook()
#创建excel工作薄
ws= w.create_sheet()
#创建Excel工作表

def get_page_content_str( url):
    time.sleep(1)

    try:
            print("现在开始抓取" + url)
            headers = {'User-Agent': 'Mozilla/5.0 (Windows NT 6.1; WOW64; rv:23.0) Gecko/20100101 Firefox/23.0'}
            #伪装浏览器
            request = urllib.request.Request(url=url, headers=headers)
            #构造请求
            m_fp = urllib.request.urlopen(request, timeout=500)
            #访问网站获取源码
            html_str = m_fp.read().decode('utf-8')
            #读取源码，该网站使用的编码方式是utf-8

            return html_str
    except Exception:
            print(Exception)
            try:
                 request = urllib.request.Request(url=url, headers=headers)
                 m_fp = urllib.request.urlopen(request, timeout=500)
                 html_str= m_fp.read().decode('gbk')
                 return html_str
            except Exception:
             print(Exception)
             try:
                 request = urllib.request.Request(url=url, headers=headers)
                 m_fp = urllib.request.urlopen(request, timeout=500)
                 html_str= m_fp.read().decode('gb2312')
                 return html_str
             except Exception:
                      print(Exception)
            m_fp.close()
#定义抓取网页源码函数

baseurl="http://chanye.focus.cn/tlist/c19_t0_r0_s0_p{}.html"
#目标网站URL
j=1
ws.cell(row=1,column=1).value="省份"
ws.cell(row=1,column=2).value="类型"
ws.cell(row=1,column=3).value="所属园区"
ws.cell(row=1,column=4).value="园区名"
ws.cell(row=1,column=5).value="地址"
ws.cell(row=1,column=6).value="面积"
ws.cell(row=1,column=7).value="价格"
ws.cell(row=1,column=8).value="发展商"
ws.cell(row=1,column=9).value="联系电话"
#设置excel单元格表头
for i in range (1,12):
    url=baseurl.format(i)
    print("现在开始抓取第{}页".format(i))
    html_str=get_page_content_str(url)
    if html_str==None:
         try:
            headers = {'User-Agent': 'Mozilla/5.0 (Windows NT 6.1; WOW64; rv:23.0) Gecko/20100101 Firefox/23.0'}
            #伪装浏览器
            request = urllib.request.Request(url=url, headers=headers)
            #构造请求
            m_fp = urllib.request.urlopen(request, timeout=500)
            #访问网站获取源码
            html_str = m_fp.read().decode('utf-8')
         except Exception:
            print(Exception)
            if html_str==None:
             try:
              headers = {'User-Agent': 'Mozilla/5.0 (Windows NT 6.1; WOW64; rv:23.0) Gecko/20100101 Firefox/23.0'}
              #伪装浏览器
              request = urllib.request.Request(url=url, headers=headers)
              #构造请求
              m_fp = urllib.request.urlopen(request, timeout=500)
              #访问网站获取源码
              html_str = m_fp.read().decode('gbk')
             except Exception:
              print(Exception)
    doc=Pq(html_str)
    parks=doc("body > div.comm-list-img > ul > li")
    for park in parks:
        j=j+1
        pro="上海"
        parkname=Pq(park)(" div > strong > a").text()
        parkurl="http://chanye.focus.cn/"+Pq(park)("div > strong > a").attr("href")
        price=Pq(park)(" div > div.txt > span > em").text()
        detailhtml=get_page_content_str(parkurl)

        if detailhtml==None:
         try:
            headers = {'User-Agent': 'Mozilla/5.0 (Windows NT 6.1; WOW64; rv:23.0) Gecko/20100101 Firefox/23.0'}
            #伪装浏览器
            request = urllib.request.Request(url=parkurl, headers=headers)
            #构造请求
            m_fp = urllib.request.urlopen(request, timeout=500)
            #访问网站获取源码
            detailhtml = m_fp.read().decode('utf-8')
         except Exception:
            print(Exception)
            if detailhtml==None:
             try:
              headers = {'User-Agent': 'Mozilla/5.0 (Windows NT 6.1; WOW64; rv:23.0) Gecko/20100101 Firefox/23.0'}
              #伪装浏览器
              request = urllib.request.Request(url=parkurl, headers=headers)
              #构造请求
              m_fp = urllib.request.urlopen(request, timeout=500)
              #访问网站获取源码
              detailhtml = m_fp.read().decode('iso8859-1').encode("iso8859-1")
             except Exception:
              print(Exception)
        detail=Pq(detailhtml)
        type=detail("#contentB > div.l > div:nth-child(1) > ul > li:nth-child(2) ").text()
        adress=detail("#contentB > div.l > div:nth-child(1) > ul > li:nth-child(4) ").text()
        belong=detail("#contentB > div.l > div:nth-child(1) > ul > li:nth-child(3) ").text()
        area=detail("#contentB > div.l > div.blockLA.blockLAA > ul:nth-child(4) > li:nth-child(2) ").text()
        developer=detail("#contentB > div.l > div.blockLA.blockLAA > ul:nth-child(1) > li:nth-child(1) ").text()
        phone=detail("#contentB > div.l > div.blockLA.blockLAA > div.more.clear > em ").text()
        type=StrUtil.substr(type,'：',"'")
        adress=StrUtil.substr(adress,'：',"'")
        belong=StrUtil.substr(belong,'：',"'")
        area=StrUtil.substr(area,'：',"'")
        developer=StrUtil.substr(developer,'：',"'")
        phone=StrUtil.substr(phone,'：',"'")
        ws.cell(row=j,column=1).value=pro
        ws.cell(row=j,column=2).value=type
        ws.cell(row=j,column=3).value=belong
        ws.cell(row=j,column=4).value=parkname
        ws.cell(row=j,column=5).value=adress
        ws.cell(row=j,column=6).value=area
        ws.cell(row=j,column=7).value=price
        ws.cell(row=j,column=8).value=developer
        ws.cell(row=j,column=9).value=phone
        #写入excel单元格
        print(pro+"-"+"-"+type+"-"+belong+"-"+parkname+"-"+adress+"-"+area+"-"+price+"-"+developer+"-"+phone)
        #在控制台输出结果
w.save(r'D:\pythontest\园区\上海.xlsx')
#保存到本地excel文件