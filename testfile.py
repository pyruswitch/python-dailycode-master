__author__ = 'Administrator'
import win32ui
import os
from openpyxl import Workbook
import win32api,win32con

dlg = win32ui.CreateFileDialog(1) # 1表示打开文件对话框
dlg.SetOFNInitialDir('E:/Python') # 设置打开文件对话框中的初始显示目录
dlg.DoModal()

filename = dlg.GetPathName() # 获取选择的文件名称
file_path = os.path.split(filename)
filepath=file_path[0]
print(filename)
print(file_path)
print(filepath)
w= Workbook()
w= Workbook()
#创建excel工作薄
ws= w.create_sheet()
w.save(str(filepath)+r"\news.xlsx")