__author__ = 'vincent'
from openpyxl import Workbook
from openpyxl.reader.excel import load_workbook
import numpy as np

def getdata(path):#读数据
    wb = load_workbook(filename = path )
    sheetnames = wb.get_sheet_names()
    ws = wb.get_sheet_by_name(sheetnames[0])
    dataarr=[]
    resultarr=[]
    for i in range(2,ws.max_row+1):
        dataarr.append([ws.cell(row=i,column=2).value,ws.cell(row=i,column=3).value,ws.cell(row=i,column=4).value])
        resultarr.append([ws.cell(row=i,column=1).value])
    x=np.array(dataarr)#输入值
    y=np.array(resultarr)#输出值
    return x,y

def derivat(x,deriv=False):#求导
    if(deriv==True):
        return x*(1-x)
    return 1/(1+np.exp(-x))

def relatalg(x,y):
     np.random.seed(1)
     syn0 = 2*np.random.random((3,1)) - 1#维度为3,1的权重矩阵
     for iter in range(10000):#训练数据
        l0 = x
        l1 = derivat(np.dot(l0,syn0))
        l1_error = y - l1
        l1_delta = l1_error * derivat(l1,True)
        syn0 += np.dot(l0.T,l1_delta)
     print("Output After Training:")
     print (l1)

if __name__ == '__main__':
    path=r"H:\upupup\功课\Python\binary.xlsx"
    x,y=getdata(path)
    relatalg(x,y)