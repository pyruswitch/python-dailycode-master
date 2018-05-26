__author__ = 'vincent'
# -*- coding:utf-8 -*-
from openpyxl import Workbook
from openpyxl.reader.excel import load_workbook
from openpyxl.writer.excel import ExcelWriter
from openpyxl.cell import get_column_letter
import urllib
import time
from pyquery import PyQuery as Pq


w= Workbook()
ws= w.create_sheet(0)
baseurl="http://61.144.226.82/ghweb/main/zwgk/yssz/ysszResultViewAction.do?method=jsydXkzQuery&&order=0&sort=desc&pageNo={}"
headers = {'User-Agent': 'Mozilla/5.0 (Windows NT 6.1; WOW64; rv:23.0) Gecko/20100101 Firefox/23.0'}
j=1
ws.cell(row=1,column=1).value="单位名称"
ws.cell(row=1,column=2).value="项目名称"
ws.cell(row=1,column=3).value="用地位置"
ws.cell(row=1,column=4).value="发证时间"
ws.cell(row=1,column=5).value="证号"
for i in range (1,200):
    url=baseurl.format(i)
    print("现在开始执行第{}页".format(i))
    print("现在开始抓取"+url)
    request=urllib.request.Request(url=url,headers=headers)
    m_fp=urllib.request.urlopen(request,timeout=500)
    html=m_fp.read().decode("gbk")
    m_fp.close()
    doc=Pq(html)
    list=doc("body > form > table >  tr:nth-child(2) > td > table.tabmaincss > tr")
    for a in list:

        Uname=Pq(a)("td:eq(0)>a").text()
        if Uname=="单位名称":
            continue
        j=j+1
        ws.cell(row=j,column=1).value=Uname
        Pname=Pq(a)("td:eq(1)>a").text()
        ws.cell(row=j,column=2).value=Pname
        adress=Pq(a)("td:eq(2)>a").text()
        ws.cell(row=j,column=3).value=adress
        creattime=Pq(a)("td:eq(3)>a").text()
        ws.cell(row=j,column=4).value=creattime
        number=Pq(a)("td:eq(4)>a").text()
        ws.cell(row=j,column=5).value=number
        print(Pname+"-"+Uname+"-"+adress+"-"+creattime+"-"+number)
w.save('E:\pythontest\yuanqudetail.xlsx')