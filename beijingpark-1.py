__author__ = 'vincent'
# -*- coding:utf-8 -*-
from openpyxl import Workbook
import urllib
from pyquery import PyQuery as Pq
from selenium import webdriver
import time
import StrUtil

w= Workbook()
#创建excel工作薄
ws= w.create_sheet()
#创建Excel工作表

def get_page_content_str( url):
    time.sleep(1)

    try:
            print("现在开始抓取" + url)
            headers = {'User-Agent': 'Mozilla/5.0 (Windows NT 6.1; WOW64; rv:23.0) Gecko/20100101 Firefox/23.0'}
            #伪装浏览器
            request = urllib.request.Request(url=url, headers=headers)
            #构造请求
            m_fp = urllib.request.urlopen(request, timeout=500)
            #访问网站获取源码
            html_str = m_fp.read().decode('gb2312')
            #读取源码，该网站使用的编码方式是utf-8
            if html_str == None:
                html_str= m_fp.read().decode('gbk')
                if html_str == None:
                    html_str = m_fp.read().decode('utf-8')
            m_fp.close()
            return html_str
    except Exception:
            print(Exception)
#定义抓取网页源码函数

baseurl="http://www.guangxinhongye.com/xq.asp?list=-------{}-----------2---"
#目标网站URL
j=1
ws.cell(row=1,column=1).value="城市"
ws.cell(row=1,column=2).value="园区名"
ws.cell(row=1,column=3).value="级别"
ws.cell(row=1,column=4).value="价格"
ws.cell(row=1,column=5).value="简介"
#设置excel单元格表头
for i in range (1,8):
    url=baseurl.format(i)
    print("现在开始抓取第{}页".format(i))
    html_str=get_page_content_str(url)
    if html_str==None:
         try:
            headers = {'User-Agent': 'Mozilla/5.0 (Windows NT 6.1; WOW64; rv:23.0) Gecko/20100101 Firefox/23.0'}
            #伪装浏览器
            request = urllib.request.Request(url=url, headers=headers)
            #构造请求
            m_fp = urllib.request.urlopen(request, timeout=500)
            #访问网站获取源码
            html_str = m_fp.read().decode('utf-8')
         except Exception:
            print(Exception)
            if html_str==None:
             try:
              headers = {'User-Agent': 'Mozilla/5.0 (Windows NT 6.1; WOW64; rv:23.0) Gecko/20100101 Firefox/23.0'}
              #伪装浏览器
              request = urllib.request.Request(url=url, headers=headers)
              #构造请求
              m_fp = urllib.request.urlopen(request, timeout=500)
              #访问网站获取源码
              html_str = m_fp.read().decode('gbk')
             except Exception:
              print(Exception)
    doc=Pq(html_str)
    parks=doc("body > div > div.con > div.con_right > div.con_right_xiangmu_bj.xiangmu_qiehuan > ul>li")
    for park in parks:
        j=j+1
        city="北京"
        parkname=Pq(park)("  div > a > div > h3").text()
        parkurl="http://www.guangxinhongye.com/"+Pq(park)(" div > a").attr("href")

        detailhtml=get_page_content_str(parkurl)

        if detailhtml==None:
         try:
            headers = {'User-Agent': 'Mozilla/5.0 (Windows NT 6.1; WOW64; rv:23.0) Gecko/20100101 Firefox/23.0'}
            #伪装浏览器
            request = urllib.request.Request(url=parkurl, headers=headers)
            #构造请求
            m_fp = urllib.request.urlopen(request, timeout=500)
            #访问网站获取源码
            detailhtml = m_fp.read().decode('utf-8')
         except Exception:
            print(Exception)
            if detailhtml==None:
             try:
              headers = {'User-Agent': 'Mozilla/5.0 (Windows NT 6.1; WOW64; rv:23.0) Gecko/20100101 Firefox/23.0'}
              #伪装浏览器
              request = urllib.request.Request(url=parkurl, headers=headers)
              #构造请求
              m_fp = urllib.request.urlopen(request, timeout=500)
              #访问网站获取源码
              detailhtml = m_fp.read().decode('gbk').encode("iso8859-1")
             except Exception:
              print(Exception)
        detail=Pq(detailhtml)
        price=detail("body > div > div.fangyuan_list>ul>li:nth-child(1)>dl > dd > div.fangyuan_list_top > span").text()+detail(" body > div > div.fangyuan_list>ul>li:nth-child(1)> dl > dd > div.fangyuan_list_top > b").text()
        type=detail("body > div > div.louyuxinxi_bj > div > ul > li:nth-child(1) > div > p:nth-child(1)").text()
        type=StrUtil.substr(type,"：",'"')
        phone=detail("body > div > div.fangyuan_list>ul>li:nth-child(1)> dl > dd > div.xiangmu_tel").text()
        abs=detail("body > div > div.louyuxinxi_bj > div > span").text()
        ws.cell(row=j,column=1).value=city
        ws.cell(row=j,column=2).value=parkname
        ws.cell(row=j,column=3).value=type
        ws.cell(row=j,column=4).value=price
        ws.cell(row=j,column=5).value=abs
        #写入excel单元格
        print(city+"-"+"-"+parkname+"-"+type+"-"+price+"-"+abs)
        #在控制台输出结果
w.save(r'D:\pythontest\园区\北京创意园.xlsx')
#保存到本地excel文件