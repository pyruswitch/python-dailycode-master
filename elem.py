__author__ = 'vincent'
# -*- coding:utf-8 -*-
from openpyxl import Workbook
import urllib
from pyquery import PyQuery as Pq
from selenium import webdriver
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.common.keys import Keys
import os,time

import StrUtil

w= Workbook()
#创建excel工作薄
ws= w.create_sheet(0)
#创建Excel工作表

def get_page_content_str( url):
    time.sleep(1)

    try:
            print("现在开始抓取" + url)
            headers = {'User-Agent': 'Mozilla/5.0 (Windows NT 6.1; WOW64; rv:23.0) Gecko/20100101 Firefox/23.0'}
            #伪装浏览器
            request = urllib.request.Request(url=url, headers=headers)
            #构造请求
            m_fp = urllib.request.urlopen(request, timeout=500)
            #访问网站获取源码
            html_str = m_fp.read().decode('utf-8')
            #读取源码，该网站使用的编码方式是utf-8
            if html_str == None:
                html_str= m_fp.read().decode('gbk')
                if html_str == None:
                    html_str = m_fp.read().decode('gb2312')
            m_fp.close()
            return html_str
    except Exception:
            print(Exception)
#定义抓取网页源码函数

baseurl="https://www.ele.me/place/ws100vs46jc"
#目标网站URL
j=1
ws.cell(row=1,column=1).value="名称"
ws.cell(row=1,column=2).value="销量"
ws.cell(row=1,column=3).value="电话"
ws.cell(row=1,column=4).value="地址"
#设置excel单元格表头
driver=webdriver.Chrome("D://chromedriver.exe")
driver.get(baseurl)
print("现在开始抓取"+baseurl)
driver.find_element_by_link_text("小吃零食").click()
time.sleep(2)
driver.find_element_by_link_text("销量高").click()
for down in range (1,8):
 ActionChains(driver).key_down(Keys.PAGE_DOWN).perform()
time.sleep(2)
for i in range(1,100):
        j=j+1
        try:
         name=driver.find_element_by_css_selector("body > div.wrapper > div.page-wrap > div > div > div.rest-list > ul > li:nth-child({}) > div > div > a.rest-atag > div.top-content > div.content > div.name > span".format(i)).text
        except  Exception:
            print(Exception)
            if i ==99:
             break
            else :continue
        try:
         proceeds=driver.find_element_by_css_selector("body > div.wrapper > div.page-wrap > div > div > div.rest-list > ul > li:nth-child({}) > div > div > a.rest-atag > div.top-content > div.content > div.rank.clearfix > span.total.cc-lightred-new.fr".format(i)).text
        except  Exception:
            print(Exception)
            proceeds="none"
        #detailurl=driver.find_element_by_css_selector("body > div.wrapper > div.page-wrap > div > div > div.rest-list > ul > li:nth-child({}) > div > div > a.rest-atag".format(i))
        #detailurl.click()

        #adress=driver.find_element_by_css_selector("body > div.wrapper > div.page-wrap > div > div.rest-info > div.details > div.rest-info-down-wrap > div.location.fl > span.fl.info-detail").text
        #phone=driver.find_element_by_css_selector("body > div.wrapper > div.page-wrap > div > div.rest-info > div.details > div.rest-info-down-wrap > div.telephone.fl > span.fl.info-detail").text
        #driver.close()
        ws.cell(row=j,column=1).value=name
        ws.cell(row=j,column=2).value=proceeds

        #写入excel单元格
        print(name+"-"+"-"+ proceeds)
        #在控制台输出结果
w.save(r'D:\pythontest\电商\elm餐饮商家.xlsx')
#保存到本地excel文件