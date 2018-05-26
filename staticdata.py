__author__ = 'vincent'
# -*- coding: utf-8 -*-
from openpyxl.reader.excel import load_workbook
import redisdemo
import redis
import datetime


def importstaticdata():

    inwb = load_workbook(filename ='昌智汇领导驾驶舱数据.xlsx' )
    wbdict={}

    namespaceidlist=[]
    for sheetName in inwb.get_sheet_names():
       sheet = inwb[sheetName]
       for row in range(2,sheet.max_row+1):
           #namespaceid=sheet.cell(row=row,column=1).value
           #namespaceidlist.append(namespaceid)
           #for col in range(1,sheet.max_colum+1):
               #key=sheet.cell(row=row,column=col).value
           if sheet.cell(row=row,column=1).value==None:
               continue
           ns=sheet.cell(row=row,column=1).value
           page=sheet.cell(row=row,column=2).value
           module=sheet.cell(row=row,column=3).value
           fieldfir=sheet.cell(row=row,column=4).value
           fieldsec=sheet.cell(row=row,column=5).value

           if ns not in wbdict.keys():
              wbdict[ns]={}
           if page not in wbdict[ns].keys():
              wbdict[ns][page]={}
           if  module not in wbdict[ns][page].keys():
               wbdict[ns][page][module]=[]
           excdict={}
           if module=="最新订单":
              #excdict={}
              excdict["ordertype"]=sheet.cell(row=row,column=6).value
              excdict["paydate"]=(sheet.cell(row=row,column=11).value).strftime('%m/%d %H:%M')
              excdict["name"]=sheet.cell(row=row,column=7).value
              excdict["paychannel"]=sheet.cell(row=row,column=10).value
              excdict["payamount"]=sheet.cell(row=row,column=9).value
              wbdict[ns][page][module].append(excdict)
           elif module in ["最新发布-活动","最新发布-帖子"]:
              excdict["name"]=sheet.cell(row=row,column=6).value
              excdict["value"]=sheet.cell(row=row,column=9).value
              excdict["time"]=(sheet.cell(row=row,column=8).value).strftime('%m/%d %H:%M')
              excdict["type"]=sheet.cell(row=row,column=7).value
              wbdict[ns][page][module].append(excdict)
           elif module=="话题排行榜":
              excdict["name"]=sheet.cell(row=row,column=6).value
              excdict["value"]=sheet.cell(row=row,column=8).value
              excdict["type"]=sheet.cell(row=row,column=7).value
              wbdict[ns][page][module].append(excdict)
           elif module=="企业名册":
              excdict["name"]=sheet.cell(row=row,column=6).value
              excdict["scale"]=sheet.cell(row=row,column=8).value
              excdict["type"]=sheet.cell(row=row,column=7).value
              excdict["adress"]=sheet.cell(row=row,column=9).value
              wbdict[ns][page][module].append(excdict)
           elif module=="最新待办任务":
              excdict["name"]=sheet.cell(row=row,column=6).value
              excdict["date"]=sheet.cell(row=row,column=8).value
              excdict["content"]=sheet.cell(row=row,column=7).value
              wbdict[ns][page][module].append(excdict)
           elif module in["水表实时读数", "电表实时读数"]:
              excdict["name"]=str(sheet.cell(row=row,column=6).value)+"***"+str((sheet.cell(row=row,column=7).value))[-2:]
              excdict["date"]=sheet.cell(row=row,column=8).value
              #excdict["content"]=sheet.cell(row=row,column=7).value
              wbdict[ns][page][module].append(excdict)
           elif module =="招租趋势图":
              excdict["month"]=sheet.cell(row=row,column=6).value
              excdict["rentarea"]=sheet.cell(row=row,column=8).value
              excdict["leasedarea"]=sheet.cell(row=row,column=7).value
              wbdict[ns][page][module].append(excdict)
           else:
               if fieldsec==None:

                  excdict["name"]=fieldfir
                  excdict["value"]=sheet.cell(row=row,column=6).value
                  wbdict[ns][page][module].append(excdict)
               else:
                  excdict["classify"]=fieldfir
                  excdict["name"]=fieldsec
                  excdict["value"]=sheet.cell(row=row,column=6).value
                  wbdict[ns][page][module].append(excdict)
       print("已读取完"+str(page)+"数据")
    redisdemo.dictredis("staticdata",wbdict,"")
    print("静态数据已读入缓存")

def parkorder(ns):
   # importstaticdata(path)
    r=redis.Redis(host='127.0.0.1',port=6379,decode_responses=True)
    data=r.hmget("staticdata",ns)
    data=eval(data[0])
    data=data["园区收入"]["最新订单"]
    return list(data)

def ordertypecount(ns):
    r=redis.Redis(host='127.0.0.1',port=6379,decode_responses=True)
    data=r.hmget("staticdata",ns)
    data=eval(data[0])
    data=data["园区收入"]["订单类型"]
    return list(data)

def ordertypeamount(ns):
    r=redis.Redis(host='127.0.0.1',port=6379,decode_responses=True)
    data=r.hmget("staticdata",ns)
    data=eval(data[0])
    data=data["园区收入"]["消费分布"]
    return list(data)

def orderchannelcount(ns):
    r=redis.Redis(host='127.0.0.1',port=6379,decode_responses=True)
    data=r.hmget("staticdata",ns)
    data=eval(data[0])
    data=data["园区收入"]["支付方式"]
    data=list(data)
    datadict={}
    for idata in data:
        idata=str(idata)
        idata=eval(idata)
        if idata["name"]=="微信":
           datadict["wechat"]=int(idata["value"])
        else:
            datadict["alipay"]=int(idata["value"])
    return datadict

def orderchannelamount(ns):
    r=redis.Redis(host='127.0.0.1',port=6379,decode_responses=True)
    data=r.hmget("staticdata",ns)
    data=eval(data[0])
    data=data["园区收入"]["支付方式金额"]
    data=list(data)
    datadict={}
    for idata in data:
        idata=str(idata)
        idata=eval(idata)
        if idata["name"]=="微信":
           datadict["wechat"]=idata["value"]
        else:
            datadict["alipay"]=idata["value"]
    return datadict

def monthamount(ns):
    r=redis.Redis(host='127.0.0.1',port=6379,decode_responses=True)
    data=r.hmget("staticdata",ns)
    data=eval(data[0])
    data=data["园区收入"]["园区收入"]
    datadict={}
    datalist=[]
    i=0
    for idata in data:
        datadict[i]={}
        idata=str(idata)
        idata=eval(idata)
        if "classify"in idata.keys():

            if idata["value"]==None:
                idata["value"]=0
            else:
             datadict[i]["value"]=idata["value"]
            datal=[str(idata["name"]),idata["value"]]
            datalist.append(datal)
            i=i+1
    return datalist

def monthcount(ns):
    r=redis.Redis(host='127.0.0.1',port=6379,decode_responses=True)
    data=r.hmget("staticdata",ns)
    data=eval(data[0])
    data=data["园区收入"]["园区订单"]
    datadict={}
    datalist=[]
    i=0
    for idata in data:
        datadict[i]={}
        idata=str(idata)
        idata=eval(idata)
        if "classify"in idata.keys():

            if idata["value"]==None:
                idata["value"]=0
            else:
             datadict[i]["value"]=idata["value"]
            datal=[str(idata["name"]),idata["value"]]
            datalist.append(datal)
            i=i+1
    return datalist

def orderamount(ns):
    r=redis.Redis(host='127.0.0.1',port=6379,decode_responses=True)
    data=r.hmget("staticdata",ns)
    data=eval(data[0])
    data=data["园区收入"]["消费总额"]
    data=str(data[0])
    data=eval(data)
    data=data["value"]
    return data

def ordercount(ns):
    r=redis.Redis(host='127.0.0.1',port=6379,decode_responses=True)
    data=r.hmget("staticdata",ns)
    data=eval(data[0])
    data=data["园区收入"]["订单总数"]
    data=str(data[0])
    data=eval(data)
    data=data["value"]
    return data

def tasktypecount(ns):
    r=redis.Redis(host='127.0.0.1',port=6379,decode_responses=True)
    data=r.hmget("staticdata",ns)
    data=eval(data[0])
    data=data["客户服务"]["任务类型分布"]
    return list(data)

def boardvita(ns):
    r=redis.Redis(host='127.0.0.1',port=6379,decode_responses=True)
    data=r.hmget("staticdata",ns)
    data=eval(data[0])
    data=data["客户服务"]["服务模块活跃度"]
    return list(data)

def taskclose(ns):
    r=redis.Redis(host='127.0.0.1',port=6379,decode_responses=True)
    data=r.hmget("staticdata",ns)
    data=eval(data[0])
    data=data["客户服务"]["概况"]
    data=list(data)
    datadict={}
    for idata in data:
        idata=str(idata)
        idata=eval(idata)
        if idata["name"]=="任务总数":
           datadict["tasktotal"]=idata["value"]
        else:
            datadict["closetasktotal"]=int(idata["value"])
    return datadict

def daytaskorder(ns):
    r=redis.Redis(host='127.0.0.1',port=6379,decode_responses=True)
    data=r.hmget("staticdata",ns)
    data=eval(data[0])
    data=data["客户服务"]["整体趋势图"]
    data=list(data)
    datadict={}
    datalist=[]
    i=0
    for idata in data:
        datadict[i]={}
        idata=str(idata)
        idata=eval(idata)
        d=idata["name"]
        datadict[i]["date"]=d.month
        datadict[i]["value"]=idata["value"]
        datalist.append(datadict[i])
        i=i+1
    return datalist
def tasktyperespon(ns):
    r=redis.Redis(host='127.0.0.1',port=6379,decode_responses=True)
    data=r.hmget("staticdata",ns)
    data=eval(data[0])
    data=data["客户服务"]["平均响应时间"]
    return list(data)

def pengdingtask(ns):
   # importstaticdata(path)
    r=redis.Redis(host='127.0.0.1',port=6379,decode_responses=True)
    data=r.hmget("staticdata",ns)
    data=eval(data[0])
    data=data["客户服务"]["最新待办任务"]
    data=list(data)
    datadict={}
    datalist=[]
    i=0
    for idata in data:
        datadict[i]={}
        idata=str(idata)
        idata=eval(idata)
        d=idata["date"]
        datadict[i]["date"]=d.strftime('%m/%d')
        datadict[i]["name"]=idata["name"]
        datadict[i]["content"]=idata["content"]
        datalist.append(datadict[i])
        i=i+1
    return list(datalist)

def taskstatus(ns):
    r=redis.Redis(host='127.0.0.1',port=6379,decode_responses=True)
    data=r.hmget("staticdata",ns)
    data=eval(data[0])
    data=data["客户服务"]["任务分布"]
    data=list(data)
    datadict={}
    datalist=[]
    i=0
    for idata in data:
        datadict[i]={}
        idata=str(idata)
        idata=eval(idata)
        datadict[i]["name"]=idata["classify"]
        datadict[i]["value"]=idata["value"]
        datadict[i]["status"]=idata["name"]
        datalist.append(datadict[i])
        i=i+1
    return list(datalist)

def totalassets(ns):
    r=redis.Redis(host='127.0.0.1',port=6379,decode_responses=True)
    data=r.hmget("staticdata",ns)
    data=eval(data[0])
    data=data["招商租赁"]["产值概况"]
    return list(data)


def Companyespon(ns):
    r=redis.Redis(host='127.0.0.1',port=6379,decode_responses=True)
    data=r.hmget("staticdata",ns)
    data=eval(data[0])
    data=data["招商租赁"]["企业名册"]

    return list(data)

def settledenter(ns):
    r=redis.Redis(host='127.0.0.1',port=6379,decode_responses=True)
    data=r.hmget("staticdata",ns)
    data=eval(data[0])
    data=data["招商租赁"]["企业概况"]
    data=list(data)
    datadict={}
    datalist=[]
    i=0
    for idata in data:
        idata=str(idata)
        idata=eval(idata)
        if idata["name"]=="今年到期数":
            datadict.setdefault("expiration",idata["value"])
        elif idata["name"]=="今年签约数":
            datadict.setdefault("contract",idata["value"])
        elif idata["name"]=="入驻企业数":
            datadict.setdefault("enterprisestotal",idata["value"])
    return datadict

def InviteBusiness(ns):
    r=redis.Redis(host='127.0.0.1',port=6379,decode_responses=True)
    data=r.hmget("staticdata",ns)
    data=eval(data[0])
    data=data["招商租赁"]["招商引资"]

    return list(data)
def Companytype(ns):
    r=redis.Redis(host='127.0.0.1',port=6379,decode_responses=True)
    data=r.hmget("staticdata",ns)
    data=eval(data[0])
    data=data["招商租赁"]["行业分析"]

    return list(data)
def occupancyrate(ns):
    r=redis.Redis(host='127.0.0.1',port=6379,decode_responses=True)
    data=r.hmget("staticdata",ns)
    data=eval(data[0])
    data=data["招商租赁"]["入住率"]

    return list(data)

def rentalamount(ns):
    r=redis.Redis(host='127.0.0.1',port=6379,decode_responses=True)
    data=r.hmget("staticdata",ns)
    data=eval(data[0])
    data=data["招商租赁"]["租赁概况"]
    data=list(data)
    datadict={}
    datalist=[]
    i=0
    for idata in data:
        idata=str(idata)
        idata=eval(idata)
        if idata["name"]=="年租金收入":
            datadict.setdefault("yearrental",idata["value"])
        elif idata["name"]=="在租面积":
            datadict.setdefault("leasedarea",idata["value"])
        elif idata["name"]=="单位租金":
            datadict.setdefault("unitrental",idata["value"])
    return datadict

def montharea(ns):
    r=redis.Redis(host='127.0.0.1',port=6379,decode_responses=True)
    data=r.hmget("staticdata",ns)
    data=eval(data[0])
    data=data["招商租赁"]["招租趋势图"]
    return list(data)

def activityenrollment(ns):
    r=redis.Redis(host='127.0.0.1',port=6379,decode_responses=True)
    data=r.hmget("staticdata",ns)
    data=eval(data[0])
    data=data["园区社交"]["活动报名"]
    data=list(data)
    datadict={}
    datalist=[]
    i=0
    for idata in data:
        idata=str(idata)
        idata=eval(idata)
        if idata["name"]=="活动报名总人次":
            datadict.setdefault("enroll",idata["value"])
        elif idata["name"]=="活动发布总数":
            datadict.setdefault("acttotal",idata["value"])
        elif idata["name"]=="平均报名人次/场":
            datadict.setdefault("aveper",idata["value"])
    return datadict
def forumthreads(ns):
    r=redis.Redis(host='127.0.0.1',port=6379,decode_responses=True)
    data=r.hmget("staticdata",ns)
    data=eval(data[0])
    data=data["园区社交"]["帖子"]
    data=list(data)
    datadict={}
    datalist=[]
    i=0
    for idata in data:
        idata=str(idata)
        idata=eval(idata)
        if idata["name"]=="帖子发布总数":
            datadict.setdefault("post",idata["value"])
        elif idata["name"]=="总回复数":
            datadict.setdefault("reply",idata["value"])
        elif idata["name"]=="总点赞数":
            datadict.setdefault("likecount",idata["value"])
        elif idata["name"]=="总阅读量":
            datadict.setdefault("viewcount",idata["value"])
    return datadict

def activitytype(ns):
    r=redis.Redis(host='127.0.0.1',port=6379,decode_responses=True)
    data=r.hmget("staticdata",ns)
    data=eval(data[0])
    data=data["园区社交"]["活动标签"]

    return list(data)
def activityattention(ns):
    r=redis.Redis(host='127.0.0.1',port=6379,decode_responses=True)
    data=r.hmget("staticdata",ns)
    data=eval(data[0])
    data=data["园区社交"]["活动热度"]

    return list(data)

def trendingtopics(ns):
    r=redis.Redis(host='127.0.0.1',port=6379,decode_responses=True)
    data=r.hmget("staticdata",ns)
    data=eval(data[0])
    data=data["园区社交"]["话题排行榜"]

    return list(data)

def newactivity(ns):
    r=redis.Redis(host='127.0.0.1',port=6379,decode_responses=True)
    data=r.hmget("staticdata",ns)
    data=eval(data[0])
    data=data["园区社交"]["最新发布-活动"]

    return list(data)
def newforum(ns):
    r=redis.Redis(host='127.0.0.1',port=6379,decode_responses=True)
    data=r.hmget("staticdata",ns)
    data=eval(data[0])
    data=data["园区社交"]["最新发布-帖子"]

    return list(data)

def communityuseract(ns):
    r=redis.Redis(host='127.0.0.1',port=6379,decode_responses=True)
    data=r.hmget("staticdata",ns)
    data=eval(data[0])
    data=data["园区社交"]["用户概况"]
    data=list(data)
    datadict={}
    datalist=[]
    i=0
    for idata in data:
        idata=str(idata)
        idata=eval(idata)
        if idata["name"]=="总注册用户":
            datadict.setdefault("totaluser",idata["value"])
        elif idata["name"]=="月新增用户":
            datadict.setdefault("thirtynewuser",idata["value"])
        elif idata["name"]=="周新增用户":
            datadict.setdefault("sevenneweuser",idata["value"])
        elif idata["name"]=="月活跃用户":
            datadict.setdefault("thirtyactiveuser",idata["value"])
        elif idata["name"]=="周活跃用户":
            datadict.setdefault("sevenactiveuser",idata["value"])
        elif idata["name"]=="日活跃用户":
            datadict.setdefault("activeuser",idata["value"])
    return datadict

def daynewusertrend(ns):
    r=redis.Redis(host='127.0.0.1',port=6379,decode_responses=True)
    data=r.hmget("staticdata",ns)
    data=eval(data[0])
    data=data["园区社交"]["新增用户"]

    return list(data)
def dayactusertrend(ns):
    r=redis.Redis(host='127.0.0.1',port=6379,decode_responses=True)
    data=r.hmget("staticdata",ns)
    data=eval(data[0])
    data=data["园区社交"]["活跃用户"]

    return list(data)
def daytotalusertrend(ns):
    r=redis.Redis(host='127.0.0.1',port=6379,decode_responses=True)
    data=r.hmget("staticdata",ns)
    data=eval(data[0])
    data=data["园区社交"]["累计用户"]

    return list(data)
def usergender(ns):
    r=redis.Redis(host='127.0.0.1',port=6379,decode_responses=True)
    data=r.hmget("staticdata",ns)
    data=eval(data[0])
    data=data["园区社交"]["性别比例"]
    data=list(data)
    datadict={}
    for idata in data:
        idata=str(idata)
        idata=eval(idata)
        if idata["name"]=="男性":
           datadict["male"]=idata["value"]
        else:
            datadict["female"]=idata["value"]
    return datadict

def totalenergy(ns):
    r=redis.Redis(host='127.0.0.1',port=6379,decode_responses=True)
    data=r.hmget("staticdata",ns)
    data=eval(data[0])
    data=data["智慧能源"]["能耗概况"]

    return list(data)
def buildmonthwatermeter(ns):
    r=redis.Redis(host='127.0.0.1',port=6379,decode_responses=True)
    data=r.hmget("staticdata",ns)
    data=eval(data[0])
    data=data["智慧能源"]["月度用水量"]

    return list(data)
def buildmonthelectr(ns):
    r=redis.Redis(host='127.0.0.1',port=6379,decode_responses=True)
    data=r.hmget("staticdata",ns)
    data=eval(data[0])
    data=data["智慧能源"]["月度用电量"]

    return list(data)

def watermetercureading(ns):
    r=redis.Redis(host='127.0.0.1',port=6379,decode_responses=True)
    data=r.hmget("staticdata",ns)
    data=eval(data[0])
    data=data["智慧能源"]["水表实时读数"]

    return list(data)

def electrmetercureading(ns):
    r=redis.Redis(host='127.0.0.1',port=6379,decode_responses=True)
    data=r.hmget("staticdata",ns)
    data=eval(data[0])
    data=data["智慧能源"]["电表实时读数"]

    return list(data)


if __name__ == '__main__':
     r=redis.Redis(host='127.0.0.1',port=6379,decode_responses=True)
     #path= r'H:\upupup\项目\功能清单\数据舱\昌智汇领导驾驶舱数据.xlsx'
     importstaticdata()
     #ns="999955"
     #print(electrmetercureading(ns))