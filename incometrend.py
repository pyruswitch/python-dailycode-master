__author__ = 'vincent'

from openpyxl import Workbook
from openpyxl.reader.excel import load_workbook
import mysql.connector
import operator
import  datetime


w= Workbook()
#创建excel工作薄
sheetnames = w.get_sheet_names()
ws = w.get_sheet_by_name(sheetnames[0])
ws.title="数据源"
#ws= w.create_sheet(0)
ws.cell(row=1,column=1).value="日期"
ws.cell(row=1,column=2).value="店铺名"
ws.cell(row=1,column=3).value="商户号"
ws.cell(row=1,column=4).value="支付渠道"
ws.cell(row=1,column=5).value="订单号"
ws.cell(row=1,column=6).value="订单类型 "
ws.cell(row=1,column=7).value="金额"
ws.cell(row=1,column=8).value="支付方式"
ws.cell(row=1,column=9).value="来源 "

user="ning.wei12"
pwd="wn3633"
host="bizdb.zuolin.com"
db="ehbiz"
port="18306"

cnx=mysql.connector.connect(user=user,password=pwd,database=db,host=host,port=port)
cursor=cnx.cursor()

select_sql='''SELECT
a.`pay_date`,c.`shop_name`,a.`pay_no`,a.`pay_type`,a.`order_no`,a.`order_type`,a.`pay_amount`,a.`online_pay_style_no`,a.`realm`
FROM
`pay_info_record`a
LEFT JOIN
`tbl_order`b
ON b.`order_no`=a.`order_no`
LEFT JOIN
`tbl_shop_info`c
ON c.`shop_no`=b.`shop_no`
WHERE
a.`pay_date`>"2016-11-01"
AND c.`shop_no`IN("14622779991551986272","14477417463124576784","14667588471072254231","14721933954264944406")
AND c.`shop_name` IS NOT NULL
AND a.`pay_status`="success"'''

cursor.execute(select_sql)
alldata=cursor.fetchall()
alldata.sort(key=operator.itemgetter(0))
rows=1
for incomdata in alldata:
    rows=rows+1
    for invalue in range(0,9):
        incomvalue=incomdata[invalue]
        ws.cell(row=rows,column=invalue+1).value=incomvalue
        if invalue==3:
            if incomvalue==1:
                incomvalue="线上"
            elif incomvalue==3:
                 incomvalue="线下"
        ws.cell(row=rows,column=invalue+1).value=incomvalue
    print(incomdata)



w.save(r"H:\upupup\电商后台\自动化统计\流水.xlsx")

cursor.close()
cnx.commit()
cnx.close()