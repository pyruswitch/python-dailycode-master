__author__ = 'vincent'
from openpyxl import Workbook
import urllib
from pyquery import PyQuery as Pq
from selenium import webdriver
import time
import StrUtil

w= Workbook()
#创建excel工作薄
sheetnames = w.get_sheet_names()
ws= w.get_sheet_by_name(sheetnames[0])
#创建Excel工作表

def get_page_content_str( url):
    time.sleep(1)

    try:
            print("现在开始抓取" + url)
            headers = {'User-Agent': 'Mozilla/5.0 (Windows NT 6.1; WOW64; rv:23.0) Gecko/20100101 Firefox/23.0'}
            #伪装浏览器
            request = urllib.request.Request(url=url, headers=headers)
            #构造请求
            m_fp = urllib.request.urlopen(request, timeout=500)
            #访问网站获取源码
            html_str = m_fp.read().decode('utf-8')
            #读取源码，该网站使用的编码方式是utf-8

            return html_str
    except Exception:
            print(Exception)
            try:
                 request = urllib.request.Request(url=url, headers=headers)
                 m_fp = urllib.request.urlopen(request, timeout=500)
                 html_str= m_fp.read().decode('gbk')
                 return html_str
            except Exception:
             print(Exception)
             try:
                 request = urllib.request.Request(url=url, headers=headers)
                 m_fp = urllib.request.urlopen(request, timeout=500)
                 html_str= m_fp.read().decode('gb2312')
                 return html_str
             except Exception:
                      print(Exception)
            m_fp.close()
#定义抓取网页源码函数
urllist=["http://dgzs.wincn.com/html/zxcq/{}.html","http://dgzs.wincn.com/html/zxcz/{}.html","http://dgzs.wincn.com/html/dbcz/{}.html","http://dgzs.wincn.com/html/xbcz/{}.html"]

#目标网站URL
j=1
ws.cell(row=1,column=1).value="类型"
ws.cell(row=1,column=2).value="面积"
ws.cell(row=1,column=3).value="销售价格"
ws.cell(row=1,column=4).value="租赁价格"
ws.cell(row=1,column=5).value="园区名"
ws.cell(row=1,column=6).value="地址"
ws.cell(row=1,column=7).value="简介"
#设置excel单元格表头
for baseurl in urllist:
  for i in range (1,4):
    if i==1:
       i="index"
    else:i==i
    url=baseurl.format(i)
    print("现在开始抓取第{}页".format(i))
    html_str=get_page_content_str(url)
    doc=Pq(html_str)
    parks=doc("#page > main > div.news > div > section > div.L.jmlist > div.newslist > article")
    for park in parks:
        j=j+1
        parkname=Pq(park)("div.newsbox > h2 > a").attr("title")
        parkhref=Pq(park)("div.newsbox > h2 > a").attr("href")
        parkdetail=get_page_content_str(parkhref)
        parkdoc=Pq(parkdetail)
        kind=parkdoc("#page > main > div > div > section > div.infoshow > div.info > div.infoclass > ul > li:nth-child(1)").text()
        area=parkdoc("#page > main > div > div > section > div.infoshow > div.info > div.infoclass > ul > li:nth-child(2)").text()
        sprice=parkdoc("#page > main > div > div > section > div.infoshow > div.info > div.infoclass > ul > li:nth-child(3) > b").text()
        zprice=parkdoc("#page > main > div > div > section > div.infoshow > div.info > div.infoclass > ul > li:nth-child(4) > b").text()
        adress=parkdoc("#page > main > div > div > section > div.infoshow > div.info > div.infoclass > ul > li:nth-child(6)").text()
        abs=parkdoc("#myTab1_Content0 > div:nth-child(1) > p").text()
        ws.cell(row=j,column=1).value=kind
        ws.cell(row=j,column=2).value=area
        ws.cell(row=j,column=3).value=sprice
        ws.cell(row=j,column=4).value=zprice
        ws.cell(row=j,column=5).value=parkname
        ws.cell(row=j,column=6).value=adress
        ws.cell(row=j,column=7).value=abs
        #写入excel单元格
        print(area+"-"+parkname+"-"+kind)
        #在控制台输出结果
w.save(r'D:\pythontest\园区\东莞招商网2.xlsx')
#保存到本地excel文件

