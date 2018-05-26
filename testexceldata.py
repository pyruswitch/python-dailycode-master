__author__ = 'vincent'
# -*- coding:utf-8 -*-
from openpyxl import Workbook
from openpyxl.reader.excel import load_workbook
from openpyxl.writer.excel import ExcelWriter
from openpyxl.cell import get_column_letter


wb = load_workbook('E:\Desktop\testexcel\KEY MESSAGE.xlsx')
sheetnames = wb.get_sheet_names()
wo = wb.get_sheet_by_name(sheetnames[0])
#w= Workbook()
#ws= w.create_sheet(0)
wo.cell(row=1,column=5).value ="小区名"
wo.cell(row=1,column=6).value ="楼栋号"
wo.cell(row=1,column=7).value="单元号"
wo.cell(row=1,column=8).value ="门牌号"
print("please wait a……")
for i in range(2,wo.get_highest_row()+1):
    endstr=wo.cell(row=i,column=2).value
    mainstr=wo.cell(row=i,column=1).value
    index=mainstr.find(endstr,0)
    if index==0:
       wo.cell(row=i,column=5).value =mainstr
       wo.cell(row=i,column=6).value =''
    else:
     name=mainstr[:index]
     wo.cell(row=i,column=5).value=name
     if "号"in endstr:
        Bindex=endstr.find("号",0)
        build=endstr[:Bindex]
        wo.cell(row=i,column=6).value=build
     elif "#"in  endstr:
        Bindex=endstr.find("#",0)
        build=endstr[:Bindex]
        wo.cell(row=i,column=6).value=build
     else:
        wo.cell(row=i,column=6).value=endstr
    apt=wo.cell(row=i,column=3).value
    if len(apt)>4:
        unit=apt[:2]
        wo.cell(row=i,column=7).value=unit
        newapt=apt[2:]
        wo.cell(row=i,column=8).value=newapt
    else:
        wo.cell(row=i,column=8).value=apt
    print(wo.cell(row=i,column=5).value,"-",wo.cell(row=i,column=6).value,"-",wo.cell(row=i,column=7).value,"-",wo.cell(row=i,column=8).value)
wb.save('E:\Desktop\testexcel\KEY.xlsx')