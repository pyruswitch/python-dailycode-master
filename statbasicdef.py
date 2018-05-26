__author__ = 'vincent'

from openpyxl import Workbook
from openpyxl.reader.excel import load_workbook
import mysql.connector
import operator
import  datetime





w= Workbook()
#创建excel工作薄
sheetnames = w.get_sheet_names()
ws = w.get_sheet_by_name(sheetnames[0])
ws.title="数据源"
wswastage=w.create_sheet(0)
wswastage.title="损耗数据"
#ws= w.create_sheet(0)
ws.cell(row=1,column=1).value="店铺"
ws.cell(row=1,column=2).value="订单号"
ws.cell(row=1,column=3).value="支付方式"
ws.cell(row=1,column=4).value="品类"
ws.cell(row=1,column=5).value="供应商 "
ws.cell(row=1,column=6).value="商品名"
ws.cell(row=1,column=7).value="规格"
ws.cell(row=1,column=8).value="成本价"
ws.cell(row=1,column=9).value="售价"
ws.cell(row=1,column=10).value="数量 "
ws.cell(row=1,column=11).value="总金额 "
ws.cell(row=1,column=12).value="折扣金额 "
ws.cell(row=1,column=13).value="支付金额 "
ws.cell(row=1,column=14).value="活动优惠 "
ws.cell(row=1,column=15).value="活动描述 "
ws.cell(row=1,column=16).value="支付时间"
ws.cell(row=1,column=17).value="下单时间段"

user="ning.wei12"
pwd="wn3633"
host="bizdb.zuolin.com"
db="ehbiz"
port="18306"

cnx=mysql.connector.connect(user=user,password=pwd,database=db,host=host,port=port)
cursor=cnx.cursor()

select_sql='''SELECT DISTINCT
c.`shop_name`,a.`order_no`,a.`pay_type`,e.`cat_name`,e.`supplier_name`,b.`prod_name`,e.`model`,d.`current_prime_price`,b.`price`,b.`quantity`,a.`price_total`,a.`discount_total`,a.`paid_total`,a.`activity_benefit_amount`,f.`activity_name`,a.`payment_time`
FROM
`tbl_order`a
LEFT JOIN `tbl_shop_info`c
ON a.`shop_no`=c.`shop_no`
LEFT JOIN  `tbl_order_item`b
ON b.`order_no`=a.`order_no`
LEFT JOIN `tbl_model`d
ON b.`prod_no`=d.`model_no`
LEFT JOIN `tbl_commodity`e
ON d.`commo_no`=e.`commo_no`
LEFT JOIN `tbl_order_activity_contact`f
ON a.`order_no`=f.`order_no`
LEFT JOIN `tbl_activity`g
ON g.activity_no=f.activity_no
WHERE
 a.`payment_time`>"2016-11-09 00:00:00"
AND a.`payment_time`<"2016-12-01 00:00:00"
AND a.`basic_state`!=1'''
cursor.execute(select_sql)
alldata=cursor.fetchall()
alldata.sort(key=operator.itemgetter(1))
rows=1
lunchseller=["房咚","乐盒饭","鲸鱼外卖","星宝外卖","鼎丰圣麦","佳味外卖"]
for proddata in alldata:
    rows=rows+1
    for Provalue in range(0,16):
        provalue=proddata[Provalue]
        ws.cell(row=rows,column=Provalue+1).value=provalue
        if Provalue==15:
            ostime=provalue.strftime("%H")
            times=str(ostime)+"—"+str(int(ostime)+1)
            ws.cell(row=rows,column=17).value=times
    print(proddata)
for i in range(2,ws.max_row+1):
    if "+"in ws.cell(row=i,column=6).value:
        if ws.cell(row=i,column=5).value in lunchseller:
            ws.cell(row=i,column=4).value="套餐"
#####################损耗数据############################################################
wastage_sql='''SELECT DISTINCT
c.`shop_name`,d.`cat_name`,d.`supplier_name`,b.`commo_name`,a.`reduce_stock_type`,a.`description`,a.`stock`,a.`price`,DATE_FORMAT(a.`create_time`,'%Y-%m-%d')
FROM
`tbl_shop_commodity_log`a
LEFT JOIN
`tbl_shop_model`b
ON a.`commo_no`=b.`commo_no`
LEFT JOIN
`tbl_shop_info`c
ON a.`shop_no`=c.`shop_no`
LEFT JOIN
`tbl_commodity`d
ON a.`commo_no`=d.`commo_no`
WHERE
 a.`operate_type`=2
AND a.`reduce_stock_type`=2
AND a.`create_time`>"2016-11-01 " '''
cursor.execute(wastage_sql)
wastagedata=cursor.fetchall()
wastagedata.sort(key=operator.itemgetter(8))
print("写入损耗数据")
wswastage.cell(row=1,column=1).value="店铺"
wswastage.cell(row=1,column=2).value="品类"
wswastage.cell(row=1,column=3).value="供应商"
wswastage.cell(row=1,column=4).value="商品名"
wswastage.cell(row=1,column=5).value="出库类型"
wswastage.cell(row=1,column=6).value="原因"
wswastage.cell(row=1,column=7).value="数量"
wswastage.cell(row=1,column=8).value="成本价"
wswastage.cell(row=1,column=9).value="日期"
wasrow=1
for wasbadate in wastagedata:
    wasrow=wasrow+1
    for wasvalue in range(0,9):
        wastagevalue=wasbadate[wasvalue]
        wswastage.cell(row=wasrow,column=wasvalue+1).value=wastagevalue
#写入excel表
#------购买倾向--------------------------------------
wg=w.create_sheet(0)
wg.title="购买行为"
wg.cell(row=1,column=1).value="购买行为"
wg.cell(row=1,column=2).value="次数"
wg.cell(row=1,column=3).value="组合详情"
wg.cell(row=1,column=4).value="次数"
groupli=[]
portion=0
group=0
package=0
j=2
maxrow=ws.max_row+1
while( j <maxrow):
    if ws.cell(row=j,column=2).value==ws.cell(row=j+1,column=2).value:
        groups=str(ws.cell(row=j,column=4).value)+"&"+str(ws.cell(row=j+1,column=4).value)
        regroups=str(ws.cell(row=j+1,column=4).value)+"&"+str(ws.cell(row=j,column=4).value)
        if regroups in groupli:
            groups=regroups
        if j+1<maxrow-1:
            for j1 in range (j+2,maxrow):
                if ws.cell(row=j,column=2).value==ws.cell(row=j1,column=2).value:
                    catname=ws.cell(row=j1,column=4).value
                    if str(catname)  in groups:
                       j=j1
                       if j==maxrow-1:
                          j=maxrow
                       else:
                        continue
                    else:
                        groups=groups+"&"+ws.cell(row=j1,column=4).value
                        j=j1

                else:
                    if "套餐"in groups:
                        package=package+1
                    else:
                        groupli.append(groups)
                        group=group+1
                    j=j1
                    break
        else :
            if "套餐"in groups:
                package=package+1
                j=j+1
            else:
                groupli.append(groups)
                group=group+1
                j=j+1
        continue
    else:
        if  ws.cell(row=j,column=4).value=="套餐":
             package=package+1
        else:portion=portion+1
        j=j+1
        #print(j)
        #print("分类执行到第"+str(j)+"行啦")
        continue
print("分类结束")
wg.cell(row=2,column=1).value="单独"
wg.cell(row=2,column=2).value=portion
wg.cell(row=3,column=1).value="组合"
wg.cell(row=3,column=2).value=group
wg.cell(row=4,column=1).value="套餐"
wg.cell(row=4,column=2).value=package
grouprow=2
groups=set(groupli)
for item in groups:

    wg.cell(row=grouprow,column=3).value=item
    wg.cell(row=grouprow,column=4).value=groupli.count(item)
    grouprow=grouprow+1
    #print("组合详情执行到第"+str(grouprow)+"行啦")
##############总体##############################################################
wsale=w.create_sheet(0)
wsale.title="总体销量统计"
salelist=[]
supplierlist=["套餐-其他商品"]
wsale.cell(row=1,column=1).value="商品名"
wsale.cell(row=1,column=2).value="销量"
wsale.cell(row=1,column=3).value="毛利"
wsale.cell(row=1,column=4).value="供应商"
wsale.cell(row=1,column=5).value="结算价"
wsale.cell(row=1,column=6).value="销售额"
wsale.cell(row=1,column=7).value="销量"
togoldrow=2
tosupperow=3
wsale.cell(row=2,column=4).value="套餐-其他商品"
wsale.cell(row=tosupperow,column=5).value=0
wsale.cell(row=2,column=5).value=0
wsale.cell(row=2,column=7).value=0
onlinelunch_dict={"房咚":"14","乐盒饭":"12.82","鲸鱼外卖":"14","星宝外卖":"12.5","船说":"14.5","鼎丰圣麦":"13","佳味外卖":"13.5"}
linelunch_dict={"房咚":"14","乐盒饭":"13.5","鲸鱼外卖":"14","星宝外卖":"12.5","船说":"14.5","鼎丰圣麦":"13","佳味外卖":"13.5"}
for goldrow in range(2,maxrow):
    goldname=ws.cell(row=goldrow,column=6).value
    goldsale=ws.cell(row=goldrow,column=10).value
    goldcost=ws.cell(row=goldrow,column=8).value
    goldprice=ws.cell(row=goldrow,column=9).value
    suppliername=ws.cell(row=goldrow,column=5).value
    goldsamount=goldprice*goldsale
    if ws.cell(row=goldrow,column=4).value=="套餐":
       if suppliername in onlinelunch_dict:
          if ws.cell(row=goldrow,column=3).value==1:
              lungoldcost=float(onlinelunch_dict[suppliername])
          else:
              lungoldcost=float(linelunch_dict[suppliername])
          if suppliername not in supplierlist:
              wsale.cell(row=tosupperow,column=4).value=suppliername
              wsale.cell(row=tosupperow,column=5).value=lungoldcost*goldsale
              wsale.cell(row=tosupperow,column=6).value=goldsamount
              wsale.cell(row=tosupperow,column=7).value=goldsale
              tosupperow=tosupperow+1
              supplierlist.append(suppliername)
          else:
              indexsupplier=supplierlist.index(suppliername)+2
              wsale.cell(row=indexsupplier,column=5).value=wsale.cell(row=indexsupplier,column=5).value+lungoldcost*goldsale
              wsale.cell(row=indexsupplier,column=6).value=wsale.cell(row=indexsupplier,column=6).value+goldsamount
              wsale.cell(row=indexsupplier,column=7).value=wsale.cell(row=indexsupplier,column=7).value+goldsale
          wsale.cell(row=2,column=5).value=wsale.cell(row=2,column=5).value+(goldcost-lungoldcost)*goldsale
          wsale.cell(row=2,column=7).value=wsale.cell(row=2,column=7).value+goldsale
       else:
           if suppliername not in supplierlist:
              wsale.cell(row=tosupperow,column=4).value=suppliername
              wsale.cell(row=tosupperow,column=5).value=goldcost*goldsale
              wsale.cell(row=tosupperow,column=6).value=goldsamount
              wsale.cell(row=tosupperow,column=7).value=goldsale
              tosupperow=tosupperow+1
              supplierlist.append(suppliername)
           else:
              indexsupplier=supplierlist.index(suppliername)+2
              wsale.cell(row=indexsupplier,column=5).value=wsale.cell(row=indexsupplier,column=5).value+goldcost*goldsale
              wsale.cell(row=indexsupplier,column=6).value=wsale.cell(row=indexsupplier,column=6).value+goldsamount
              wsale.cell(row=indexsupplier,column=7).value=wsale.cell(row=indexsupplier,column=7).value+goldsale
    else:
           if suppliername not in supplierlist:
              wsale.cell(row=tosupperow,column=4).value=suppliername
              wsale.cell(row=tosupperow,column=5).value=goldcost*goldsale
              wsale.cell(row=tosupperow,column=6).value=goldsamount
              wsale.cell(row=tosupperow,column=7).value=goldsale
              tosupperow=tosupperow+1
              supplierlist.append(suppliername)
           else:
              indexsupplier=supplierlist.index(suppliername)+2
              wsale.cell(row=indexsupplier,column=5).value=wsale.cell(row=indexsupplier,column=5).value+goldcost*goldsale
              wsale.cell(row=indexsupplier,column=6).value=wsale.cell(row=indexsupplier,column=6).value+goldsamount
              wsale.cell(row=indexsupplier,column=7).value=wsale.cell(row=indexsupplier,column=7).value+goldsale
    if goldname not in salelist:

        wsale.cell(row=togoldrow,column=1).value=goldname
        wsale.cell(row=togoldrow,column=2).value=goldsale
        salelist.append(goldname)
        profit=goldsale*(goldprice-goldcost)
        wsale.cell(row=togoldrow,column=3).value=profit
        togoldrow=togoldrow+1
    else:
        indexrow=salelist.index(goldname)+2
        goldsale=wsale.cell(row=indexrow,column=2).value+ws.cell(row=goldrow,column=10).value
        wsale.cell(row=indexrow,column=2).value=goldsale
        profit=wsale.cell(row=indexrow,column=3).value+ws.cell(row=goldrow,column=10).value*(goldprice-goldcost)
        wsale.cell(row=indexrow,column=3).value=profit
    #print("结算执行到第"+str(togoldrow)+"行啦")
ssortlist=[]
for salesort in range(2,wsale.max_row+1):
    ssortlist.append((wsale.cell(row=salesort,column=1).value,wsale.cell(row=salesort,column=2).value,wsale.cell(row=salesort,column=3).value))
ssortlist.sort(key=operator.itemgetter(1),reverse=True)
sortrow=1
for ssort in ssortlist:
    sortrow=sortrow+1
    for ssortele in range(0,3):
        selement=ssort[ssortele]
        wsale.cell(row=sortrow,column=ssortele+1).value=selement
        #print("销量执行到第"+str(sortrow)+"行啦")
wsuprow=tosupperow+2
#供应商损耗
Wdaylist=[t[8]for t in wastagedata]
Wdaylist=set(Wdaylist)
Wdaylist=list(Wdaylist)
Wdaylist.sort()
Wsuplist=[t[2]for t in wastagedata]
Wsuplist=set(Wsuplist)
Wshoplist=[t[0]for t in wastagedata]
Wshoplist=set(Wshoplist)
wsale.cell(row=wsuprow,column=4).value="损耗-供应商"
wsale.cell(row=wsuprow,column=5).value="金额"
wsale.cell(row=wsuprow,column=6).value="数量"

#wsuprow=tosupperow+2
for wsup in Wsuplist:
    wsuprow=wsuprow+1
    wsale.cell(row=wsuprow,column=4).value=wsup
    wsupcount=0
    wsupamount=0
    for wastageele in wastagedata:
        if wastageele[2]==wsup:
           wsupcount=wsupcount+wastageele[6]
           wsupamount=wsupamount+wastageele[7]*wastageele[6]
    wsale.cell(row=wsuprow,column=6).value=wsupcount
    wsale.cell(row=wsuprow,column=5).value=wsupamount
#站点损耗
wsuprow=wsuprow+2
wsale.cell(row=wsuprow,column=4).value="损耗-站点"
wsale.cell(row=wsuprow,column=5).value="金额"
wsale.cell(row=wsuprow,column=6).value="数量"
for wshop in Wshoplist:
    wsuprow=wsuprow+1
    wsale.cell(row=wsuprow,column=4).value=wshop
    wshopcount=0
    wshopamount=0
    for wastageele in wastagedata:
        if wastageele[0]==wshop:
           wshopcount=wshopcount+wastageele[6]
           wshopamount=wshopamount+wastageele[7]*wastageele[6]
    wsale.cell(row=wsuprow,column=6).value=wshopcount
    wsale.cell(row=wsuprow,column=5).value=wshopamount
#每日站点损耗
wsuprow=wsuprow+2
wsale.cell(row=wsuprow,column=4).value="日期(损耗)"

wshoprow=wsuprow
for wdayele in Wdaylist:
    wsuprow=wsuprow+1
    wsale.cell(row=wsuprow,column=4).value=wdayele
    wdaycol=5
    for wshop in Wshoplist:
        wsale.cell(row=wshoprow,column=wdaycol).value=wshop
        wasdayshopcount=0
        for wdashopele in wastagedata:
            if wdashopele[0]==wshop:
                if wdashopele[8]==wdayele:
                    wasdayshopcount=wasdayshopcount+wdashopele[7]*wdashopele[6]
        wsale.cell(row=wsuprow,column=wdaycol).value=wasdayshopcount
        wdaycol=wdaycol+1
#每日供应商损耗-总额
wsuprow=wsuprow+2
wsale.cell(row=wsuprow,column=4).value="日期/损耗总额"

wsuperrow=wsuprow
for wdayele in Wdaylist:
    wsuprow=wsuprow+1
    wsale.cell(row=wsuprow,column=4).value=wdayele
    wdaycol=5
    for wsuper in Wsuplist:
        wsale.cell(row=wsuperrow,column=wdaycol).value=wsuper
        wasdaysupercount=0
        for wdashopele in wastagedata:
            if wdashopele[2]==wsuper:
                if wdashopele[8]==wdayele:
                    wasdaysupercount=wasdaysupercount+wdashopele[7]*wdashopele[6]
        wsale.cell(row=wsuprow,column=wdaycol).value=wasdaysupercount
        wdaycol=wdaycol+1
#每日供应商损耗-数量
wsuprow=wsuprow+2
wsale.cell(row=wsuprow,column=4).value="日期/损耗数量"

wsuperrow=wsuprow
for wdayele in Wdaylist:
    wsuprow=wsuprow+1
    wsale.cell(row=wsuprow,column=4).value=wdayele
    wdaycol=5
    for wsuper in Wsuplist:
        wsale.cell(row=wsuperrow,column=wdaycol).value=wsuper
        wasdaysupercount=0
        for wdashopele in wastagedata:
            if wdashopele[2]==wsuper:
                if wdashopele[8]==wdayele:
                    wasdaysupercount=wasdaysupercount+wdashopele[6]
        wsale.cell(row=wsuprow,column=wdaycol).value=wasdaysupercount
        wdaycol=wdaycol+1
#----------------------品类销量-------------------------#
wsale.cell(row=1,column=9).value="日期"
sdatelist=[]
scatelist=[]
for sdate in range(2,ws.max_row+1):
    datevalue=ws.cell(row=sdate,column=16).value
    if datevalue==None:
        continue
    datevalue=datetime.datetime.strptime(str(datevalue),"%Y-%m-%d %H:%M:%S")
    datevalue=datevalue.strftime("%Y-%m-%d")
    #print("执行到"+str(sdate))
    sdatelist.append(datevalue)
for scate in range(2,ws.max_row+1):
    catevalue=ws.cell(row=scate,column=4).value
    if catevalue==None:
        continue
    scatelist.append(catevalue)
saledates=set(sdatelist)
saledates=list(saledates)
saledates.sort()
salecates=set(scatelist)

daterow=1


for saledate in saledates:

    daterow=daterow+1
    wsale.cell(row=daterow,column=9).value=saledate
    catecol=10
    for salecate in salecates:
        wsale.cell(row=1,column=catecol).value=salecate

        catedcount=0
        for catecount in range(2,ws.max_row+1):
            datevalue=ws.cell(row=catecount,column=16).value
            if datevalue==None:
                continue
            catevalue=ws.cell(row=catecount,column=4).value
            datevalue=datetime.datetime.strptime(str(datevalue),"%Y-%m-%d %H:%M:%S")
            datevalue=datevalue.strftime("%Y-%m-%d")
            if  catevalue==salecate:
                if datevalue==saledate:
                    catedcount=ws.cell(row=catecount,column=10).value+catedcount
                    wsale.cell(row=daterow,column=catecol).value=catedcount
                else:
                    wsale.cell(row=daterow,column=catecol).value=catedcount
            else:
                wsale.cell(row=daterow,column=catecol).value=catedcount
        catecol=catecol+1
    print("品类销量执行到"+str(saledate)+"号")
#----------------------------------供应商销量----------------------------------------------------------------------------#
wsale.cell(row=1,column=catecol+2).value="日期"
sellerlist=[]
for sseller in range(2,ws.max_row+1):
    sellervalue=ws.cell(row=sseller,column=5).value
    if sellervalue==None:
        continue
    sellerlist.append(sellervalue)
salesellers=set(sellerlist)

daterow=1
for saledate in saledates:
    daterow=daterow+1
    wsale.cell(row=daterow,column=catecol+2).value=saledate
    sellercol=catecol+3
    for saleseller in salesellers:
        wsale.cell(row=1,column=sellercol).value=saleseller

        sellerdcount=0
        for sellercount in range(2,ws.max_row+1):
            datevalue=ws.cell(row=sellercount,column=16).value
            if datevalue==None:
                continue
            sellcountvalue=ws.cell(row=sellercount,column=5).value
            datevalue=datetime.datetime.strptime(str(datevalue),"%Y-%m-%d %H:%M:%S")
            datevalue=datevalue.strftime("%Y-%m-%d")
            if  sellcountvalue==saleseller:
                if datevalue==saledate:
                    sellerdcount=ws.cell(row=sellercount,column=10).value+sellerdcount
                    wsale.cell(row=daterow,column=sellercol).value=sellerdcount
                else:
                    wsale.cell(row=daterow,column=sellercol).value=sellerdcount
            else:
                wsale.cell(row=daterow,column=sellercol).value=sellerdcount
        sellercol=sellercol+1
    print("供应商销量执行到"+str(saledate)+"号")

#-----------------------午餐情况----------------------------------------------------
wt=w.create_sheet(0)
wt.title="午餐情况"
wt.cell(row=1,column=1).value="时间段"
wt.cell(row=1,column=2).value="午餐下单量"
orderli=[]
timesli=[]
timerow=1
for timele in range(2,ws.max_row+1):
    if ws.cell(row=timele,column=5).value in lunchseller:
        if ws.cell(row=timele,column=2)not in orderli:
            timesli.append(ws.cell(row=timele,column=17).value)
            orderli.append(ws.cell(row=timele,column=2).value)
        else:continue
    else:continue
    #print("时间段执行到第"+str(timerow)+"行啦")
timegroup=set(timesli)
timegroup=list(timegroup)
timegroup.sort()
for itemtime in timegroup:
   timerow=timerow+1
   wt.cell(row=timerow,column=1).value=itemtime
   wt.cell(row=timerow,column=2).value=timesli.count(itemtime)
#---------------------午餐各供应商销量-------------------------------------------------
wt.cell(row=1,column=4).value="日期"
wt.cell(row=1,column=5).value="总计"
dayrow=1
for saleday in  saledates:
    dayrow=dayrow+1
    wt.cell(row=dayrow,column=4).value=saleday
    lunchsupcol=6
    wt.cell(row=dayrow,column=5).value=0
    for lunchsup  in lunchseller:
        wt.cell(row=1,column=lunchsupcol).value=lunchsup
        lunchsupcount=0
        for lunchele in range(2,ws.max_row):
            datevalue=ws.cell(row=lunchele,column=16).value
            if datevalue==None:
                continue
            lunchsupname=ws.cell(row=lunchele,column=5).value
            datevalue=datetime.datetime.strptime(str(datevalue),"%Y-%m-%d %H:%M:%S")
            datevalue=datevalue.strftime("%Y-%m-%d")
            if datevalue==saleday:
                if lunchsup==lunchsupname:
                    lunchsupcount=ws.cell(row=lunchele,column=10).value+lunchsupcount

        wt.cell(row=dayrow,column=lunchsupcol).value=lunchsupcount
        lunchsupcol=lunchsupcol+1
        wt.cell(row=dayrow,column=5).value=wt.cell(row=dayrow,column=5).value+lunchsupcount

#---------------------午餐各站点销量-------------------------------------------------
dayrow=dayrow+2
wt.cell(row=dayrow,column=4).value="日期"
wt.cell(row=dayrow,column=5).value="总计"
dayshoprow=dayrow
Wlunchshoplist=["UFine左邻小站－中国储能大厦","左邻小站—科技工业园大厦","海岸左邻小站","Volgo Caffe－金融科技大厦"]
for saleday in  saledates:
    dayrow=dayrow+1
    wt.cell(row=dayrow,column=4).value=saleday
    lunchshopcol=6
    wt.cell(row=dayrow,column=5).value=0
    for lunchshop  in Wlunchshoplist:
        wt.cell(row=dayshoprow,column=lunchshopcol).value=lunchshop
        lunchshopcount=0
        for lunchele in range(2,ws.max_row):
            datevalue=ws.cell(row=lunchele,column=16).value
            if datevalue==None:
                continue
            lunchshopname=ws.cell(row=lunchele,column=1).value
            datevalue=datetime.datetime.strptime(str(datevalue),"%Y-%m-%d %H:%M:%S")
            datevalue=datevalue.strftime("%Y-%m-%d")
            if datevalue==saleday:
                if lunchshop==lunchshopname:
                    if ws.cell(row=lunchele,column=5).value in lunchseller:
                       lunchshopcount=ws.cell(row=lunchele,column=10).value+lunchshopcount

        wt.cell(row=dayrow,column=lunchshopcol).value=lunchshopcount
        lunchshopcol=lunchshopcol+1
        wt.cell(row=dayrow,column=5).value=wt.cell(row=dayrow,column=5).value+lunchshopcount
#############储能#######################################################################
wshopa=w.create_sheet(0)
wshopa.title="UFine左邻小站－中国储能大厦销量统计"
salelist=[]
supplierlist=["套餐-其他商品"]
wshopa.cell(row=1,column=1).value="商品名"
wshopa.cell(row=1,column=2).value="销量"
wshopa.cell(row=1,column=3).value="毛利"
wshopa.cell(row=1,column=4).value="供应商"
wshopa.cell(row=1,column=5).value="结算价"
wshopa.cell(row=1,column=6).value="销售额"
wshopa.cell(row=1,column=7).value="销量"
togoldrow=2
tosupperow=3
wshopa.cell(row=2,column=4).value="套餐-其他商品"
wshopa.cell(row=tosupperow,column=5).value=0
wshopa.cell(row=2,column=5).value=0
wshopa.cell(row=2,column=7).value=0
onlinelunch_dict={"房咚":"14","乐盒饭":"12.82","鲸鱼外卖":"14","星宝外卖":"12.5","船说":"14.5","鼎丰圣麦":"13","佳味外卖":"13.5"}
linelunch_dict={"房咚":"14","乐盒饭":"13.5","鲸鱼外卖":"14","星宝外卖":"12.5","船说":"14.5","鼎丰圣麦":"13","佳味外卖":"13.5"}
for goldrow in range(2,maxrow):
    if ws.cell(row=goldrow,column=1).value!="UFine左邻小站－中国储能大厦":
        continue
    goldname=ws.cell(row=goldrow,column=6).value
    goldsale=ws.cell(row=goldrow,column=10).value
    goldcost=ws.cell(row=goldrow,column=8).value
    goldprice=ws.cell(row=goldrow,column=9).value
    suppliername=ws.cell(row=goldrow,column=5).value
    goldsamount=goldprice*goldsale
    if ws.cell(row=goldrow,column=4).value=="套餐":
       if suppliername in onlinelunch_dict:
          if ws.cell(row=goldrow,column=3).value==1:
              lungoldcost=float(onlinelunch_dict[suppliername])
          else:
              lungoldcost=float(linelunch_dict[suppliername])
          if suppliername not in supplierlist:
              wshopa.cell(row=tosupperow,column=4).value=suppliername
              wshopa.cell(row=tosupperow,column=5).value=lungoldcost*goldsale
              wshopa.cell(row=tosupperow,column=6).value=goldsamount
              wshopa.cell(row=tosupperow,column=7).value=goldsale
              tosupperow=tosupperow+1
              supplierlist.append(suppliername)
          else:
              indexsupplier=supplierlist.index(suppliername)+2
              wshopa.cell(row=indexsupplier,column=5).value=wshopa.cell(row=indexsupplier,column=5).value+lungoldcost*goldsale
              wshopa.cell(row=indexsupplier,column=6).value=wshopa.cell(row=indexsupplier,column=6).value+goldsamount
              wshopa.cell(row=indexsupplier,column=7).value=wshopa.cell(row=indexsupplier,column=7).value+goldsale
          wshopa.cell(row=2,column=5).value=wshopa.cell(row=2,column=5).value+(goldcost-lungoldcost)*goldsale
          wshopa.cell(row=2,column=7).value=wshopa.cell(row=2,column=7).value+goldsale
       else:
           if suppliername not in supplierlist:
              wshopa.cell(row=tosupperow,column=4).value=suppliername
              wshopa.cell(row=tosupperow,column=5).value=goldcost*goldsale
              wshopa.cell(row=tosupperow,column=6).value=goldsamount
              wshopa.cell(row=tosupperow,column=7).value=goldsale
              tosupperow=tosupperow+1
              supplierlist.append(suppliername)
           else:
              indexsupplier=supplierlist.index(suppliername)+2
              wshopa.cell(row=indexsupplier,column=5).value=wshopa.cell(row=indexsupplier,column=5).value+goldcost*goldsale
              wshopa.cell(row=indexsupplier,column=6).value=wshopa.cell(row=indexsupplier,column=6).value+goldsamount
              wshopa.cell(row=indexsupplier,column=7).value=wshopa.cell(row=indexsupplier,column=7).value+goldsale
    else:
           if suppliername not in supplierlist:
              wshopa.cell(row=tosupperow,column=4).value=suppliername
              wshopa.cell(row=tosupperow,column=5).value=goldcost*goldsale
              wshopa.cell(row=tosupperow,column=6).value=goldsamount
              wshopa.cell(row=tosupperow,column=7).value=goldsale
              tosupperow=tosupperow+1
              supplierlist.append(suppliername)
           else:
              indexsupplier=supplierlist.index(suppliername)+2
              wshopa.cell(row=indexsupplier,column=5).value=wshopa.cell(row=indexsupplier,column=5).value+goldcost*goldsale
              wshopa.cell(row=indexsupplier,column=6).value=wshopa.cell(row=indexsupplier,column=6).value+goldsamount
              wshopa.cell(row=indexsupplier,column=7).value=wshopa.cell(row=indexsupplier,column=7).value+goldsale
    if goldname not in salelist:

        wshopa.cell(row=togoldrow,column=1).value=goldname
        wshopa.cell(row=togoldrow,column=2).value=goldsale
        salelist.append(goldname)
        profit=goldsale*(goldprice-goldcost)
        wshopa.cell(row=togoldrow,column=3).value=profit
        togoldrow=togoldrow+1
    else:
        indexrow=salelist.index(goldname)+2
        goldsale=wshopa.cell(row=indexrow,column=2).value+ws.cell(row=goldrow,column=10).value
        wshopa.cell(row=indexrow,column=2).value=goldsale
        profit=wshopa.cell(row=indexrow,column=3).value+ws.cell(row=goldrow,column=10).value*(goldprice-goldcost)
        wshopa.cell(row=indexrow,column=3).value=profit
    #print("结算执行到第"+str(togoldrow)+"行啦")
ssortlist=[]
for salesort in range(2,wshopa.max_row+1):
    ssortlist.append((wshopa.cell(row=salesort,column=1).value,wshopa.cell(row=salesort,column=2).value,wshopa.cell(row=salesort,column=3).value))
ssortlist.sort(key=operator.itemgetter(1),reverse=True)
sortrow=1
for ssort in ssortlist:
    sortrow=sortrow+1
    for ssortele in range(0,3):
        selement=ssort[ssortele]
        wshopa.cell(row=sortrow,column=ssortele+1).value=selement
        #print("销量执行到第"+str(sortrow)+"行啦")
wsuprow=tosupperow+2
#供应商损耗
Wdaylist=[t[8]for t in wastagedata]
Wdaylist=set(Wdaylist)
Wsuplist=[t[2]for t in wastagedata]
Wsuplist=set(Wsuplist)
#Wshoplist=[t[0]for t in wastagedata]
#Wshoplist=set(Wshoplist)
wshopa.cell(row=wsuprow,column=4).value="损耗-供应商"
wshopa.cell(row=wsuprow,column=5).value="金额"
wshopa.cell(row=wsuprow,column=6).value="数量"

#wsuprow=tosupperow+2
for wsup in Wsuplist:
    wsuprow=wsuprow+1
    wshopa.cell(row=wsuprow,column=4).value=wsup
    wsupcount=0
    wsupamount=0
    for wastageele in wastagedata:
        if wastageele[0]!="UFine左邻小站－中国储能大厦":
         continue
        if wastageele[2]==wsup:
           wsupcount=wsupcount+wastageele[6]
           wsupamount=wsupamount+wastageele[7]*wastageele[6]
    wshopa.cell(row=wsuprow,column=6).value=wsupcount
    wshopa.cell(row=wsuprow,column=5).value=wsupamount

#----------------------品类销量-------------------------#
wshopa.cell(row=1,column=9).value="日期"
sdatelist=[]
scatelist=[]
for sdate in range(2,ws.max_row+1):
    datevalue=ws.cell(row=sdate,column=16).value
    if datevalue==None:
        continue
    datevalue=datetime.datetime.strptime(str(datevalue),"%Y-%m-%d %H:%M:%S")
    datevalue=datevalue.strftime("%Y-%m-%d")
    #print("储能执行到"+str(sdate))
    sdatelist.append(datevalue)
for scate in range(2,ws.max_row+1):
    catevalue=ws.cell(row=scate,column=4).value
    if catevalue==None:
        continue
    scatelist.append(catevalue)
saledates=set(sdatelist)
saledates=list(saledates)
saledates.sort()
salecates=set(scatelist)

daterow=1


for saledate in saledates:

    daterow=daterow+1
    wshopa.cell(row=daterow,column=9).value=saledate
    catecol=10
    for salecate in salecates:
        wshopa.cell(row=1,column=catecol).value=salecate

        catedcount=0
        for catecount in range(2,ws.max_row+1):
            if ws.cell(row=catecount,column=1).value!="UFine左邻小站－中国储能大厦":
                   continue
            datevalue=ws.cell(row=catecount,column=16).value
            if datevalue==None:
                continue
            catevalue=ws.cell(row=catecount,column=4).value
            datevalue=datetime.datetime.strptime(str(datevalue),"%Y-%m-%d %H:%M:%S")
            datevalue=datevalue.strftime("%Y-%m-%d")
            if  catevalue==salecate:
                if datevalue==saledate:
                    catedcount=ws.cell(row=catecount,column=10).value+catedcount
                    wshopa.cell(row=daterow,column=catecol).value=catedcount
                else:
                    wshopa.cell(row=daterow,column=catecol).value=catedcount
            else:
                wshopa.cell(row=daterow,column=catecol).value=catedcount
        catecol=catecol+1
    print("储能品类销量执行到"+str(saledate)+"号")
#----------------------------------供应商销量----------------------------------------------------------------------------#
wshopa.cell(row=1,column=catecol+2).value="日期"
sellerlist=[]
for sseller in range(2,ws.max_row+1):
    sellervalue=ws.cell(row=sseller,column=5).value
    if sellervalue==None:
        continue
    sellerlist.append(sellervalue)
salesellers=set(sellerlist)

daterow=1
for saledate in saledates:
    daterow=daterow+1
    wshopa.cell(row=daterow,column=catecol+2).value=saledate
    sellercol=catecol+3
    for saleseller in salesellers:
        wshopa.cell(row=1,column=sellercol).value=saleseller

        sellerdcount=0
        for sellercount in range(2,ws.max_row+1):
            if ws.cell(row=sellercount,column=1).value!="UFine左邻小站－中国储能大厦":
                   continue
            datevalue=ws.cell(row=sellercount,column=16).value
            if datevalue==None:
                continue
            sellcountvalue=ws.cell(row=sellercount,column=5).value
            datevalue=datetime.datetime.strptime(str(datevalue),"%Y-%m-%d %H:%M:%S")
            datevalue=datevalue.strftime("%Y-%m-%d")
            if  sellcountvalue==saleseller:
                if datevalue==saledate:
                    sellerdcount=ws.cell(row=sellercount,column=10).value+sellerdcount
                    wshopa.cell(row=daterow,column=sellercol).value=sellerdcount
                else:
                    wshopa.cell(row=daterow,column=sellercol).value=sellerdcount
            else:
                wshopa.cell(row=daterow,column=sellercol).value=sellerdcount
        sellercol=sellercol+1
    print("储能供应商销量执行到"+str(saledate)+"号")

###################################################################################
#############工业大厦#######################################################################
wshopb=w.create_sheet(0)
wshopb.title="左邻小站—科技工业园大厦销量统计"
salelist=[]
supplierlist=["套餐-其他商品"]
wshopb.cell(row=1,column=1).value="商品名"
wshopb.cell(row=1,column=2).value="销量"
wshopb.cell(row=1,column=3).value="毛利"
wshopb.cell(row=1,column=4).value="供应商"
wshopb.cell(row=1,column=5).value="结算价"
wshopb.cell(row=1,column=6).value="销售额"
wshopb.cell(row=1,column=7).value="销量"
togoldrow=2
tosupperow=3
wshopb.cell(row=2,column=4).value="套餐-其他商品"
wshopb.cell(row=tosupperow,column=5).value=0
wshopb.cell(row=2,column=5).value=0
wshopb.cell(row=2,column=7).value=0
onlinelunch_dict={"房咚":"14","乐盒饭":"12.82","鲸鱼外卖":"14","星宝外卖":"12.5","船说":"14.5","鼎丰圣麦":"13","佳味外卖":"13.5"}
linelunch_dict={"房咚":"14","乐盒饭":"13.5","鲸鱼外卖":"14","星宝外卖":"12.5","船说":"14.5","鼎丰圣麦":"13","佳味外卖":"13.5"}
for goldrow in range(2,maxrow):
    if ws.cell(row=goldrow,column=1).value!="左邻小站—科技工业园大厦":
        continue
    goldname=ws.cell(row=goldrow,column=6).value
    goldsale=ws.cell(row=goldrow,column=10).value
    goldcost=ws.cell(row=goldrow,column=8).value
    goldprice=ws.cell(row=goldrow,column=9).value
    suppliername=ws.cell(row=goldrow,column=5).value
    goldsamount=goldprice*goldsale
    if ws.cell(row=goldrow,column=4).value=="套餐":
       if suppliername in onlinelunch_dict:
          if ws.cell(row=goldrow,column=3).value==1:
              lungoldcost=float(onlinelunch_dict[suppliername])
          else:
              lungoldcost=float(linelunch_dict[suppliername])
          if suppliername not in supplierlist:
              wshopb.cell(row=tosupperow,column=4).value=suppliername
              wshopb.cell(row=tosupperow,column=5).value=lungoldcost*goldsale
              wshopb.cell(row=tosupperow,column=6).value=goldsamount
              wshopb.cell(row=tosupperow,column=7).value=goldsale
              tosupperow=tosupperow+1
              supplierlist.append(suppliername)
          else:
              indexsupplier=supplierlist.index(suppliername)+2
              wshopb.cell(row=indexsupplier,column=5).value=wshopb.cell(row=indexsupplier,column=5).value+lungoldcost*goldsale
              wshopb.cell(row=indexsupplier,column=6).value=wshopb.cell(row=indexsupplier,column=6).value+goldsamount
              wshopb.cell(row=indexsupplier,column=7).value=wshopb.cell(row=indexsupplier,column=7).value+goldsale
          wshopb.cell(row=2,column=5).value=wshopb.cell(row=2,column=5).value+(goldcost-lungoldcost)*goldsale
          wshopb.cell(row=2,column=7).value=wshopb.cell(row=2,column=7).value+goldsale
       else:
           if suppliername not in supplierlist:
              wshopb.cell(row=tosupperow,column=4).value=suppliername
              wshopb.cell(row=tosupperow,column=5).value=goldcost*goldsale
              wshopb.cell(row=tosupperow,column=6).value=goldsamount
              wshopb.cell(row=tosupperow,column=7).value=goldsale
              tosupperow=tosupperow+1
              supplierlist.append(suppliername)
           else:
              indexsupplier=supplierlist.index(suppliername)+2
              wshopb.cell(row=indexsupplier,column=5).value=wshopb.cell(row=indexsupplier,column=5).value+goldcost*goldsale
              wshopb.cell(row=indexsupplier,column=6).value=wshopb.cell(row=indexsupplier,column=6).value+goldsamount
              wshopb.cell(row=indexsupplier,column=7).value=wshopb.cell(row=indexsupplier,column=7).value+goldsale
    else:
           if suppliername not in supplierlist:
              wshopb.cell(row=tosupperow,column=4).value=suppliername
              wshopb.cell(row=tosupperow,column=5).value=goldcost*goldsale
              wshopb.cell(row=tosupperow,column=6).value=goldsamount
              wshopb.cell(row=tosupperow,column=7).value=goldsale
              tosupperow=tosupperow+1
              supplierlist.append(suppliername)
           else:
              indexsupplier=supplierlist.index(suppliername)+2
              wshopb.cell(row=indexsupplier,column=5).value=wshopb.cell(row=indexsupplier,column=5).value+goldcost*goldsale
              wshopb.cell(row=indexsupplier,column=6).value=wshopb.cell(row=indexsupplier,column=6).value+goldsamount
              wshopb.cell(row=indexsupplier,column=7).value=wshopb.cell(row=indexsupplier,column=7).value+goldsale
    if goldname not in salelist:

        wshopb.cell(row=togoldrow,column=1).value=goldname
        wshopb.cell(row=togoldrow,column=2).value=goldsale
        salelist.append(goldname)
        profit=goldsale*(goldprice-goldcost)
        wshopb.cell(row=togoldrow,column=3).value=profit
        togoldrow=togoldrow+1
    else:
        indexrow=salelist.index(goldname)+2
        goldsale=wshopb.cell(row=indexrow,column=2).value+ws.cell(row=goldrow,column=10).value
        wshopb.cell(row=indexrow,column=2).value=goldsale
        profit=wshopb.cell(row=indexrow,column=3).value+ws.cell(row=goldrow,column=10).value*(goldprice-goldcost)
        wshopb.cell(row=indexrow,column=3).value=profit
    #print("结算执行到第"+str(togoldrow)+"行啦")
ssortlist=[]
for salesort in range(2,wshopb.max_row+1):
    ssortlist.append((wshopb.cell(row=salesort,column=1).value,wshopb.cell(row=salesort,column=2).value,wshopb.cell(row=salesort,column=3).value))
ssortlist.sort(key=operator.itemgetter(1),reverse=True)
sortrow=1
for ssort in ssortlist:
    sortrow=sortrow+1
    for ssortele in range(0,3):
        selement=ssort[ssortele]
        wshopb.cell(row=sortrow,column=ssortele+1).value=selement
        #print("销量执行到第"+str(sortrow)+"行啦")
wsuprow=tosupperow+2
#供应商损耗
Wdaylist=[t[8]for t in wastagedata]
Wdaylist=set(Wdaylist)
Wsuplist=[t[2]for t in wastagedata]
Wsuplist=set(Wsuplist)
#Wshoplist=[t[0]for t in wastagedata]
#Wshoplist=set(Wshoplist)
wshopb.cell(row=wsuprow,column=4).value="损耗-供应商"
wshopb.cell(row=wsuprow,column=5).value="金额"
wshopb.cell(row=wsuprow,column=6).value="数量"

#wsuprow=tosupperow+2
for wsup in Wsuplist:
    wsuprow=wsuprow+1
    wshopb.cell(row=wsuprow,column=4).value=wsup
    wsupcount=0
    wsupamount=0
    for wastageele in wastagedata:
        if wastageele[0]!="左邻小站—科技工业园大厦":
         continue
        if wastageele[2]==wsup:
           wsupcount=wsupcount+wastageele[6]
           wsupamount=wsupamount+wastageele[7]*wastageele[6]
    wshopb.cell(row=wsuprow,column=6).value=wsupcount
    wshopb.cell(row=wsuprow,column=5).value=wsupamount
#----------------------品类销量-------------------------#
wshopb.cell(row=1,column=9).value="日期"
sdatelist=[]
scatelist=[]
for sdate in range(2,ws.max_row+1):
    datevalue=ws.cell(row=sdate,column=16).value
    if datevalue==None:
        continue
    datevalue=datetime.datetime.strptime(str(datevalue),"%Y-%m-%d %H:%M:%S")
    datevalue=datevalue.strftime("%Y-%m-%d")
    #print("工业大堂执行到"+str(sdate))
    sdatelist.append(datevalue)
for scate in range(2,ws.max_row+1):
    catevalue=ws.cell(row=scate,column=4).value
    if catevalue==None:
        continue
    scatelist.append(catevalue)
saledates=set(sdatelist)
saledates=list(saledates)
saledates.sort()
salecates=set(scatelist)

daterow=1


for saledate in saledates:

    daterow=daterow+1
    wshopb.cell(row=daterow,column=9).value=saledate
    catecol=10
    for salecate in salecates:
        wshopb.cell(row=1,column=catecol).value=salecate

        catedcount=0
        for catecount in range(2,ws.max_row+1):
            if ws.cell(row=catecount,column=1).value!="左邻小站—科技工业园大厦":
                   continue
            datevalue=ws.cell(row=catecount,column=16).value
            if datevalue==None:
                continue
            catevalue=ws.cell(row=catecount,column=4).value
            datevalue=datetime.datetime.strptime(str(datevalue),"%Y-%m-%d %H:%M:%S")
            datevalue=datevalue.strftime("%Y-%m-%d")
            if  catevalue==salecate:
                if datevalue==saledate:
                    catedcount=ws.cell(row=catecount,column=10).value+catedcount
                    wshopb.cell(row=daterow,column=catecol).value=catedcount
                else:
                    wshopb.cell(row=daterow,column=catecol).value=catedcount
            else:
                wshopb.cell(row=daterow,column=catecol).value=catedcount
        catecol=catecol+1
    print("工业大堂品类销量执行到"+str(saledate)+"号")
#----------------------------------供应商销量----------------------------------------------------------------------------#
wshopb.cell(row=1,column=catecol+2).value="日期"
sellerlist=[]
for sseller in range(2,ws.max_row+1):
    sellervalue=ws.cell(row=sseller,column=5).value
    if sellervalue==None:
        continue
    sellerlist.append(sellervalue)
salesellers=set(sellerlist)

daterow=1
for saledate in saledates:
    daterow=daterow+1
    wshopb.cell(row=daterow,column=catecol+2).value=saledate
    sellercol=catecol+3
    for saleseller in salesellers:
        wshopb.cell(row=1,column=sellercol).value=saleseller

        sellerdcount=0
        for sellercount in range(2,ws.max_row+1):
            if ws.cell(row=sellercount,column=1).value!="左邻小站—科技工业园大厦":
                   continue
            datevalue=ws.cell(row=sellercount,column=16).value
            if datevalue==None:
                continue
            sellcountvalue=ws.cell(row=sellercount,column=5).value
            datevalue=datetime.datetime.strptime(str(datevalue),"%Y-%m-%d %H:%M:%S")
            datevalue=datevalue.strftime("%Y-%m-%d")
            if  sellcountvalue==saleseller:
                if datevalue==saledate:
                    sellerdcount=ws.cell(row=sellercount,column=10).value+sellerdcount
                    wshopb.cell(row=daterow,column=sellercol).value=sellerdcount
                else:
                    wshopb.cell(row=daterow,column=sellercol).value=sellerdcount
            else:
                wshopb.cell(row=daterow,column=sellercol).value=sellerdcount
        sellercol=sellercol+1
    print("工业大堂供应商销量执行到"+str(saledate)+"号")

###################################################################################
#############金融科技#######################################################################
wshopc=w.create_sheet(0)
wshopc.title="Volgo Caffe－金融科技大厦销量统计"
salelist=[]
supplierlist=["套餐-其他商品"]
wshopc.cell(row=1,column=1).value="商品名"
wshopc.cell(row=1,column=2).value="销量"
wshopc.cell(row=1,column=3).value="毛利"
wshopc.cell(row=1,column=4).value="供应商"
wshopc.cell(row=1,column=5).value="结算价"
wshopc.cell(row=1,column=6).value="销售额"
wshopc.cell(row=1,column=7).value="销量"
togoldrow=2
tosupperow=3
wshopc.cell(row=2,column=4).value="套餐-其他商品"
wshopc.cell(row=tosupperow,column=5).value=0
wshopc.cell(row=2,column=5).value=0
wshopc.cell(row=2,column=7).value=0
onlinelunch_dict={"房咚":"14","乐盒饭":"12.82","鲸鱼外卖":"14","星宝外卖":"12.5","船说":"14.5","鼎丰圣麦":"13","佳味外卖":"13.5"}
linelunch_dict={"房咚":"14","乐盒饭":"13.5","鲸鱼外卖":"14","星宝外卖":"12.5","船说":"14.5","鼎丰圣麦":"13","佳味外卖":"13.5"}
for goldrow in range(2,maxrow):
    if ws.cell(row=goldrow,column=1).value!="Volgo Caffe－金融科技大厦":
        continue
    goldname=ws.cell(row=goldrow,column=6).value
    goldsale=ws.cell(row=goldrow,column=10).value
    goldcost=ws.cell(row=goldrow,column=8).value
    goldprice=ws.cell(row=goldrow,column=9).value
    suppliername=ws.cell(row=goldrow,column=5).value
    goldsamount=goldprice*goldsale
    if ws.cell(row=goldrow,column=4).value=="套餐":
       if suppliername in onlinelunch_dict:
          if ws.cell(row=goldrow,column=3).value==1:
              lungoldcost=float(onlinelunch_dict[suppliername])
          else:
              lungoldcost=float(linelunch_dict[suppliername])
          if suppliername not in supplierlist:
              wshopc.cell(row=tosupperow,column=4).value=suppliername
              wshopc.cell(row=tosupperow,column=5).value=lungoldcost*goldsale
              wshopc.cell(row=tosupperow,column=6).value=goldsamount
              wshopc.cell(row=tosupperow,column=7).value=goldsale
              tosupperow=tosupperow+1
              supplierlist.append(suppliername)
          else:
              indexsupplier=supplierlist.index(suppliername)+2
              wshopc.cell(row=indexsupplier,column=5).value=wshopc.cell(row=indexsupplier,column=5).value+lungoldcost*goldsale
              wshopc.cell(row=indexsupplier,column=6).value=wshopc.cell(row=indexsupplier,column=6).value+goldsamount
              wshopc.cell(row=indexsupplier,column=7).value=wshopc.cell(row=indexsupplier,column=7).value+goldsale
          wshopc.cell(row=2,column=5).value=wshopc.cell(row=2,column=5).value+(goldcost-lungoldcost)*goldsale
          wshopc.cell(row=2,column=7).value=wshopc.cell(row=2,column=7).value+goldsale
       else:
           if suppliername not in supplierlist:
              wshopc.cell(row=tosupperow,column=4).value=suppliername
              wshopc.cell(row=tosupperow,column=5).value=goldcost*goldsale
              wshopc.cell(row=tosupperow,column=6).value=goldsamount
              wshopc.cell(row=tosupperow,column=7).value=goldsale
              tosupperow=tosupperow+1
              supplierlist.append(suppliername)
           else:
              indexsupplier=supplierlist.index(suppliername)+2
              wshopc.cell(row=indexsupplier,column=5).value=wshopc.cell(row=indexsupplier,column=5).value+goldcost*goldsale
              wshopc.cell(row=indexsupplier,column=6).value=wshopc.cell(row=indexsupplier,column=6).value+goldsamount
              wshopc.cell(row=indexsupplier,column=7).value=wshopc.cell(row=indexsupplier,column=7).value+goldsale
    else:
           if suppliername not in supplierlist:
              wshopc.cell(row=tosupperow,column=4).value=suppliername
              wshopc.cell(row=tosupperow,column=5).value=goldcost*goldsale
              wshopc.cell(row=tosupperow,column=6).value=goldsamount
              wshopc.cell(row=tosupperow,column=7).value=goldsale
              tosupperow=tosupperow+1
              supplierlist.append(suppliername)
           else:
              indexsupplier=supplierlist.index(suppliername)+2
              wshopc.cell(row=indexsupplier,column=5).value=wshopc.cell(row=indexsupplier,column=5).value+goldcost*goldsale
              wshopc.cell(row=indexsupplier,column=6).value=wshopc.cell(row=indexsupplier,column=6).value+goldsamount
              wshopc.cell(row=indexsupplier,column=7).value=wshopc.cell(row=indexsupplier,column=7).value+goldsale
    if goldname not in salelist:

        wshopc.cell(row=togoldrow,column=1).value=goldname
        wshopc.cell(row=togoldrow,column=2).value=goldsale
        salelist.append(goldname)
        profit=goldsale*(goldprice-goldcost)
        wshopc.cell(row=togoldrow,column=3).value=profit
        togoldrow=togoldrow+1
    else:
        indexrow=salelist.index(goldname)+2
        goldsale=wshopc.cell(row=indexrow,column=2).value+ws.cell(row=goldrow,column=10).value
        wshopc.cell(row=indexrow,column=2).value=goldsale
        profit=wshopc.cell(row=indexrow,column=3).value+ws.cell(row=goldrow,column=10).value*(goldprice-goldcost)
        wshopc.cell(row=indexrow,column=3).value=profit
    #print("结算执行到第"+str(togoldrow)+"行啦")
ssortlist=[]
for salesort in range(2,wshopc.max_row+1):
    ssortlist.append((wshopc.cell(row=salesort,column=1).value,wshopc.cell(row=salesort,column=2).value,wshopc.cell(row=salesort,column=3).value))
ssortlist.sort(key=operator.itemgetter(1),reverse=True)
sortrow=1
for ssort in ssortlist:
    sortrow=sortrow+1
    for ssortele in range(0,3):
        selement=ssort[ssortele]
        wshopc.cell(row=sortrow,column=ssortele+1).value=selement
        #print("销量执行到第"+str(sortrow)+"行啦")
wsuprow=tosupperow+2
#供应商损耗
Wdaylist=[t[8]for t in wastagedata]
Wdaylist=set(Wdaylist)
Wsuplist=[t[2]for t in wastagedata]
Wsuplist=set(Wsuplist)
#Wshoplist=[t[0]for t in wastagedata]
#Wshoplist=set(Wshoplist)
wshopc.cell(row=wsuprow,column=4).value="损耗-供应商"
wshopc.cell(row=wsuprow,column=5).value="金额"
wshopc.cell(row=wsuprow,column=6).value="数量"

#wsuprow=tosupperow+2
for wsup in Wsuplist:
    wsuprow=wsuprow+1
    wshopc.cell(row=wsuprow,column=4).value=wsup
    wsupcount=0
    wsupamount=0
    for wastageele in wastagedata:
        if wastageele[0]!="Volgo Caffe－金融科技大厦":
         continue
        if wastageele[2]==wsup:
           wsupcount=wsupcount+wastageele[6]
           wsupamount=wsupamount+wastageele[7]*wastageele[6]
    wshopc.cell(row=wsuprow,column=6).value=wsupcount
    wshopc.cell(row=wsuprow,column=5).value=wsupamount
#----------------------品类销量-------------------------#
wshopc.cell(row=1,column=9).value="日期"
sdatelist=[]
scatelist=[]
for sdate in range(2,ws.max_row+1):
    datevalue=ws.cell(row=sdate,column=16).value
    if datevalue==None:
        continue
    datevalue=datetime.datetime.strptime(str(datevalue),"%Y-%m-%d %H:%M:%S")
    datevalue=datevalue.strftime("%Y-%m-%d")
    #print("金融科技执行到"+str(sdate))
    sdatelist.append(datevalue)
for scate in range(2,ws.max_row+1):
    catevalue=ws.cell(row=scate,column=4).value
    if catevalue==None:
        continue
    scatelist.append(catevalue)
saledates=set(sdatelist)
saledates=list(saledates)
saledates.sort()
salecates=set(scatelist)

daterow=1


for saledate in saledates:

    daterow=daterow+1
    wshopc.cell(row=daterow,column=9).value=saledate
    catecol=10
    for salecate in salecates:
        wshopc.cell(row=1,column=catecol).value=salecate

        catedcount=0
        for catecount in range(2,ws.max_row+1):
            if ws.cell(row=catecount,column=1).value!="Volgo Caffe－金融科技大厦":
                   continue
            datevalue=ws.cell(row=catecount,column=16).value
            if datevalue==None:
                continue
            catevalue=ws.cell(row=catecount,column=4).value
            datevalue=datetime.datetime.strptime(str(datevalue),"%Y-%m-%d %H:%M:%S")
            datevalue=datevalue.strftime("%Y-%m-%d")
            if  catevalue==salecate:
                if datevalue==saledate:
                    catedcount=ws.cell(row=catecount,column=10).value+catedcount
                    wshopc.cell(row=daterow,column=catecol).value=catedcount
                else:
                    wshopc.cell(row=daterow,column=catecol).value=catedcount
            else:
                wshopc.cell(row=daterow,column=catecol).value=catedcount
        catecol=catecol+1
    print("金融科技品类销量执行到"+str(saledate)+"号")
#----------------------------------供应商销量----------------------------------------------------------------------------#
wshopc.cell(row=1,column=catecol+2).value="日期"
sellerlist=[]
for sseller in range(2,ws.max_row+1):
    sellervalue=ws.cell(row=sseller,column=5).value
    if sellervalue==None:
        continue
    sellerlist.append(sellervalue)
salesellers=set(sellerlist)

daterow=1
for saledate in saledates:
    daterow=daterow+1
    wshopc.cell(row=daterow,column=catecol+2).value=saledate
    sellercol=catecol+3
    for saleseller in salesellers:
        wshopc.cell(row=1,column=sellercol).value=saleseller

        sellerdcount=0
        for sellercount in range(2,ws.max_row+1):
            if ws.cell(row=sellercount,column=1).value!="Volgo Caffe－金融科技大厦":
                   continue
            datevalue=ws.cell(row=sellercount,column=16).value
            if datevalue==None:
                continue
            sellcountvalue=ws.cell(row=sellercount,column=5).value
            datevalue=datetime.datetime.strptime(str(datevalue),"%Y-%m-%d %H:%M:%S")
            datevalue=datevalue.strftime("%Y-%m-%d")
            if  sellcountvalue==saleseller:
                if datevalue==saledate:
                    sellerdcount=ws.cell(row=sellercount,column=10).value+sellerdcount
                    wshopc.cell(row=daterow,column=sellercol).value=sellerdcount
                else:
                    wshopc.cell(row=daterow,column=sellercol).value=sellerdcount
            else:
                wshopc.cell(row=daterow,column=sellercol).value=sellerdcount
        sellercol=sellercol+1
    print("金融科技供应商销量执行到"+str(saledate)+"号")
#############海岸#######################################################################
wshopd=w.create_sheet(0)
wshopd.title="海岸左邻小站销量统计"
salelist=[]
supplierlist=["套餐-其他商品"]
wshopd.cell(row=1,column=1).value="商品名"
wshopd.cell(row=1,column=2).value="销量"
wshopd.cell(row=1,column=3).value="毛利"
wshopd.cell(row=1,column=4).value="供应商"
wshopd.cell(row=1,column=5).value="结算价"
wshopd.cell(row=1,column=6).value="销售额"
wshopd.cell(row=1,column=7).value="销量"
togoldrow=2
tosupperow=3
wshopd.cell(row=2,column=4).value="套餐-其他商品"
wshopd.cell(row=tosupperow,column=5).value=0
wshopd.cell(row=2,column=5).value=0
wshopd.cell(row=2,column=7).value=0
onlinelunch_dict={"房咚":"14","乐盒饭":"12.82","鲸鱼外卖":"14","星宝外卖":"12.5","船说":"14.5","鼎丰圣麦":"13","佳味外卖":"13.5"}
linelunch_dict={"房咚":"14","乐盒饭":"13.5","鲸鱼外卖":"14","星宝外卖":"12.5","船说":"14.5","鼎丰圣麦":"13","佳味外卖":"13.5"}
for goldrow in range(2,maxrow):
    if ws.cell(row=goldrow,column=1).value!="海岸左邻小站":
        continue
    goldname=ws.cell(row=goldrow,column=6).value
    goldsale=ws.cell(row=goldrow,column=10).value
    goldcost=ws.cell(row=goldrow,column=8).value
    goldprice=ws.cell(row=goldrow,column=9).value
    suppliername=ws.cell(row=goldrow,column=5).value
    goldsamount=goldprice*goldsale
    if ws.cell(row=goldrow,column=4).value=="套餐":
       if suppliername in onlinelunch_dict:
          if ws.cell(row=goldrow,column=3).value==1:
              lungoldcost=float(onlinelunch_dict[suppliername])
          else:
              lungoldcost=float(linelunch_dict[suppliername])
          if suppliername not in supplierlist:
              wshopd.cell(row=tosupperow,column=4).value=suppliername
              wshopd.cell(row=tosupperow,column=5).value=lungoldcost*goldsale
              wshopd.cell(row=tosupperow,column=6).value=goldsamount
              wshopd.cell(row=tosupperow,column=7).value=goldsale
              tosupperow=tosupperow+1
              supplierlist.append(suppliername)
          else:
              indexsupplier=supplierlist.index(suppliername)+2
              wshopd.cell(row=indexsupplier,column=5).value=wshopd.cell(row=indexsupplier,column=5).value+lungoldcost*goldsale
              wshopd.cell(row=indexsupplier,column=6).value=wshopd.cell(row=indexsupplier,column=6).value+goldsamount
              wshopd.cell(row=indexsupplier,column=7).value=wshopd.cell(row=indexsupplier,column=7).value+goldsale
          wshopd.cell(row=2,column=5).value=wshopd.cell(row=2,column=5).value+(goldcost-lungoldcost)*goldsale
          wshopd.cell(row=2,column=7).value=wshopd.cell(row=2,column=7).value+goldsale
       else:
           if suppliername not in supplierlist:
              wshopd.cell(row=tosupperow,column=4).value=suppliername
              wshopd.cell(row=tosupperow,column=5).value=goldcost*goldsale
              wshopd.cell(row=tosupperow,column=6).value=goldsamount
              wshopd.cell(row=tosupperow,column=7).value=goldsale
              tosupperow=tosupperow+1
              supplierlist.append(suppliername)
           else:
              indexsupplier=supplierlist.index(suppliername)+2
              wshopd.cell(row=indexsupplier,column=5).value=wshopd.cell(row=indexsupplier,column=5).value+goldcost*goldsale
              wshopd.cell(row=indexsupplier,column=6).value=wshopd.cell(row=indexsupplier,column=6).value+goldsamount
              wshopd.cell(row=indexsupplier,column=7).value=wshopd.cell(row=indexsupplier,column=7).value+goldsale
    else:
           if suppliername not in supplierlist:
              wshopd.cell(row=tosupperow,column=4).value=suppliername
              wshopd.cell(row=tosupperow,column=5).value=goldcost*goldsale
              wshopd.cell(row=tosupperow,column=6).value=goldsamount
              wshopd.cell(row=tosupperow,column=7).value=goldsale
              tosupperow=tosupperow+1
              supplierlist.append(suppliername)
           else:
              indexsupplier=supplierlist.index(suppliername)+2
              wshopd.cell(row=indexsupplier,column=5).value=wshopd.cell(row=indexsupplier,column=5).value+goldcost*goldsale
              wshopd.cell(row=indexsupplier,column=6).value=wshopd.cell(row=indexsupplier,column=6).value+goldsamount
              wshopd.cell(row=indexsupplier,column=7).value=wshopd.cell(row=indexsupplier,column=7).value+goldsale
    if goldname not in salelist:

        wshopd.cell(row=togoldrow,column=1).value=goldname
        wshopd.cell(row=togoldrow,column=2).value=goldsale
        salelist.append(goldname)
        profit=goldsale*(goldprice-goldcost)
        wshopd.cell(row=togoldrow,column=3).value=profit
        togoldrow=togoldrow+1
    else:
        indexrow=salelist.index(goldname)+2
        goldsale=wshopd.cell(row=indexrow,column=2).value+ws.cell(row=goldrow,column=10).value
        wshopd.cell(row=indexrow,column=2).value=goldsale
        profit=wshopd.cell(row=indexrow,column=3).value+ws.cell(row=goldrow,column=10).value*(goldprice-goldcost)
        wshopd.cell(row=indexrow,column=3).value=profit
    #print("结算执行到第"+str(togoldrow)+"行啦")
ssortlist=[]
for salesort in range(2,wshopd.max_row+1):
    ssortlist.append((wshopd.cell(row=salesort,column=1).value,wshopd.cell(row=salesort,column=2).value,wshopd.cell(row=salesort,column=3).value))
ssortlist.sort(key=operator.itemgetter(1),reverse=True)
sortrow=1
for ssort in ssortlist:
    sortrow=sortrow+1
    for ssortele in range(0,3):
        selement=ssort[ssortele]
        wshopd.cell(row=sortrow,column=ssortele+1).value=selement
        #print("销量执行到第"+str(sortrow)+"行啦")
wsuprow=tosupperow+2
#供应商损耗
Wdaylist=[t[8]for t in wastagedata]
Wdaylist=set(Wdaylist)
Wsuplist=[t[2]for t in wastagedata]
Wsuplist=set(Wsuplist)
#Wshoplist=[t[0]for t in wastagedata]
#Wshoplist=set(Wshoplist)
wshopd.cell(row=wsuprow,column=4).value="损耗-供应商"
wshopd.cell(row=wsuprow,column=5).value="金额"
wshopd.cell(row=wsuprow,column=6).value="数量"

#wsuprow=tosupperow+2
for wsup in Wsuplist:
    wsuprow=wsuprow+1
    wshopd.cell(row=wsuprow,column=4).value=wsup
    wsupcount=0
    wsupamount=0
    for wastageele in wastagedata:
        if wastageele[0]!="海岸左邻小站":
         continue
        if wastageele[2]==wsup:
           wsupcount=wsupcount+wastageele[6]
           wsupamount=wsupamount+wastageele[7]*wastageele[6]
    wshopd.cell(row=wsuprow,column=6).value=wsupcount
    wshopd.cell(row=wsuprow,column=5).value=wsupamount
#----------------------品类销量-------------------------#
wshopd.cell(row=1,column=9).value="日期"
sdatelist=[]
scatelist=[]
for sdate in range(2,ws.max_row+1):
    datevalue=ws.cell(row=sdate,column=16).value
    if datevalue==None:
        continue
    datevalue=datetime.datetime.strptime(str(datevalue),"%Y-%m-%d %H:%M:%S")
    datevalue=datevalue.strftime("%Y-%m-%d")
    #print("海岸执行到"+str(sdate))
    sdatelist.append(datevalue)
for scate in range(2,ws.max_row+1):
    catevalue=ws.cell(row=scate,column=4).value
    if catevalue==None:
        continue
    scatelist.append(catevalue)
saledates=set(sdatelist)
saledates=list(saledates)
saledates.sort()
salecates=set(scatelist)

daterow=1


for saledate in saledates:

    daterow=daterow+1
    wshopd.cell(row=daterow,column=9).value=saledate
    catecol=10
    for salecate in salecates:
        wshopd.cell(row=1,column=catecol).value=salecate

        catedcount=0
        for catecount in range(2,ws.max_row+1):
            if ws.cell(row=catecount,column=1).value!="海岸左邻小站":
                   continue
            datevalue=ws.cell(row=catecount,column=16).value
            if datevalue==None:
                continue
            catevalue=ws.cell(row=catecount,column=4).value
            datevalue=datetime.datetime.strptime(str(datevalue),"%Y-%m-%d %H:%M:%S")
            datevalue=datevalue.strftime("%Y-%m-%d")
            if  catevalue==salecate:
                if datevalue==saledate:
                    catedcount=ws.cell(row=catecount,column=10).value+catedcount
                    wshopd.cell(row=daterow,column=catecol).value=catedcount
                else:
                    wshopd.cell(row=daterow,column=catecol).value=catedcount
            else:
                wshopd.cell(row=daterow,column=catecol).value=catedcount
        catecol=catecol+1
    print("海岸品类销量执行到"+str(saledate)+"号")
#----------------------------------供应商销量----------------------------------------------------------------------------#
wshopd.cell(row=1,column=catecol+2).value="日期"
sellerlist=[]
for sseller in range(2,ws.max_row+1):
    sellervalue=ws.cell(row=sseller,column=5).value
    if sellervalue==None:
        continue
    sellerlist.append(sellervalue)
salesellers=set(sellerlist)

daterow=1
for saledate in saledates:
    daterow=daterow+1
    wshopd.cell(row=daterow,column=catecol+2).value=saledate
    sellercol=catecol+3
    for saleseller in salesellers:
        wshopd.cell(row=1,column=sellercol).value=saleseller

        sellerdcount=0
        for sellercount in range(2,ws.max_row+1):
            if ws.cell(row=sellercount,column=1).value!="海岸左邻小站":
                   continue
            datevalue=ws.cell(row=sellercount,column=16).value
            if datevalue==None:
                continue
            sellcountvalue=ws.cell(row=sellercount,column=5).value
            datevalue=datetime.datetime.strptime(str(datevalue),"%Y-%m-%d %H:%M:%S")
            datevalue=datevalue.strftime("%Y-%m-%d")
            if  sellcountvalue==saleseller:
                if datevalue==saledate:
                    sellerdcount=ws.cell(row=sellercount,column=10).value+sellerdcount
                    wshopd.cell(row=daterow,column=sellercol).value=sellerdcount
                else:
                    wshopd.cell(row=daterow,column=sellercol).value=sellerdcount
            else:
                wshopd.cell(row=daterow,column=sellercol).value=sellerdcount
        sellercol=sellercol+1
    print("海岸供应商销量执行到"+str(saledate)+"号")



w.save(r"H:\upupup\电商后台\自动化统计\0825.xlsx")

cursor.close()
cnx.commit()
cnx.close()