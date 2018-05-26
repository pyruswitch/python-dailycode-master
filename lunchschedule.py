__author__ = 'vincent'
from openpyxl import Workbook
from openpyxl.reader.excel import load_workbook
import mysql.connector
import operator
import math
import  datetime
from time import sleep
from tqdm import tqdm
import time
import random
from datetime import datetime, timedelta

def dbsql (selectsql):#数据库查询

    user="ning.wei15"
    pwd="wn3333"
    host="bizdb.zuolin.com"
    db="ehbiz"
    port="18306"

    cnx=mysql.connector.connect(user=user,password=pwd,database=db,host=host,port=port)
    cursor=cnx.cursor()
    select_sql=selectsql

    cursor.execute(select_sql)
    alldata=cursor.fetchall()
    #alldata.sort(key=operator.itemgetter(0))
    cursor.close()
    cnx.commit()
    cnx.close()
    return alldata

def shopsupprosale():

    select_sql='''SELECT DISTINCT
    c.`shop_name`,e.`supplier_name`,b.`prod_name`,SUM(b.`quantity`),COUNT(DISTINCT DATE_FORMAT(a.`payment_time`,'%Y-%m-%d'))
    FROM
    `tbl_order`a
    LEFT JOIN `tbl_shop_info`c
    ON a.`shop_no`=c.`shop_no`
    LEFT JOIN  `tbl_order_item`b
    ON b.`order_no`=a.`order_no`
    LEFT JOIN `tbl_model`d
    ON b.`prod_no`=d.`model_no`
    LEFT JOIN `tbl_commodity`e
    ON d.`commo_no`=e.`commo_no`
    WHERE
     DATE_SUB(CURDATE(), INTERVAL 60 DAY) <= a.`payment_time`
    AND a.`basic_state`!=1
    AND e.`cat_name`="盒饭"
    AND a.`shop_no`IN("14622779991551986272","14477417463124576784","14667588471072254231","14721933954264944406","14629620426739375504","14923963824806752750")
    GROUP BY c.`shop_name`,e.`supplier_name`,b.`prod_name`'''

    alldata=dbsql(select_sql)
    shopprodsale=[]
    shopsupdict={}
    shoplist=[t[0]for t in alldata]
    shoplist=set(shoplist)
    shoplist=list(shoplist)
    suplist=[]
    prolist=[]
    supnum=0
    pronum=0
    newalldata=[]
    for proddata in alldata:
         proddata=list(proddata)
         avgsale=int(proddata[3]/proddata[4])
         proddata.append(avgsale)
         newalldata.append(proddata)
    newalldata.sort(key=operator.itemgetter(0,1,5),reverse=True)
    for shopi in shoplist:
        suplist.append([])
        for proddata in newalldata:
            if  proddata[0]==shopi:
                  if proddata[1] not in suplist[supnum]:
                      suplist[supnum].append(proddata[1])
                      prolist.append([])
                      prolist[pronum].append(proddata[2])
                      pronum=pronum+1
                  else:
                      prolist[pronum-1].append(proddata[2])
        supnum=supnum+1
    suppronum=0
    shopsupnum=0
    for shopi in shoplist:
        shopsupdict[shopi]={}
        supi=suplist[shopsupnum]
        for supdata in supi:
            shopsupdict[shopi][supdata]={}
            shopsupdict[shopi][supdata]=prolist[suppronum]
            suppronum=suppronum+1
        shopsupnum=shopsupnum+1
    print(shopsupdict)
    return(shopsupdict)
def shopyesdaysale():
     select_sql='''
     SELECT DISTINCT
c.`shop_name`,e.`supplier_name`,SUM(b.`quantity`), DATE_FORMAT(a.`payment_time`,'%Y-%m-%d')
FROM
`tbl_order`a
LEFT JOIN `tbl_shop_info`c
ON a.`shop_no`=c.`shop_no`
LEFT JOIN  `tbl_order_item`b
ON b.`order_no`=a.`order_no`
LEFT JOIN `tbl_model`d
ON b.`prod_no`=d.`model_no`
LEFT JOIN `tbl_commodity`e
ON d.`commo_no`=e.`commo_no`
WHERE
 DATE_SUB(CURDATE(),INTERVAL 0 DAY)=DATE_FORMAT(a.`payment_time`,'%Y-%m-%d')
AND a.`basic_state`!=1
AND e.`cat_name`="盒饭"
AND a.`shop_no`IN("14622779991551986272","14477417463124576784","14667588471072254231","14721933954264944406","14629620426739375504","14923963824806752750")
GROUP BY c.`shop_name`,e.`supplier_name`, DATE_FORMAT(a.`payment_time`,'%Y-%m-%d')
    '''
     alldata=dbsql(select_sql)
     shopsupdict={}
     shoplist=[t[0]for t in alldata]
     superlist=[t[1]for t in alldata]
     shoplist=set(shoplist)
     shoplist=list(shoplist)
     superlist=set(superlist)
     superlist=list(superlist)
     for shopi in shoplist:
        shopsupdict[shopi]={}
        for supdata in alldata:
            if supdata[0]==shopi:
                supname=supdata[1]
                shopsupdict[shopi][supname]={}
                shopsupdict[shopi][supname]= int(supdata[2])
     #print(shopsupdict)
     wastage_sql='''
     SELECT
c.`shop_name`,d.`supplier_name`,SUM(a.`stock`)
FROM
`tbl_shop_commodity_log`a

LEFT JOIN `tbl_commodity`d
ON a.`commo_no`=d.`commo_no`
LEFT JOIN
`tbl_shop_info`c
ON a.`shop_no`=c.`shop_no`
WHERE
DATE_SUB(CURDATE(),INTERVAL 0 DAY)=DATE_FORMAT(a.`create_time`,'%Y-%m-%d')
AND  a.`reduce_stock_type` IN (2,4)
AND d.`cat_name`LIKE"%盒饭%"
GROUP BY c.`shop_name`,d.`supplier_name`
     '''
     wastagedata=dbsql(wastage_sql)
     shopwasdict={}
     shoplist=[t[0]for t in wastagedata]
     shoplist=set(shoplist)
     shoplist=list(shoplist)
     for shopi in shoplist:
        for supdata in  wastagedata:
            if supdata[0]==shopi:
                supname=supdata[1]
                shopwasdict.setdefault(shopi,[]).append(supname)
     #print(shopwasdict)
     return superlist,shopsupdict,shopwasdict
def menusort(consultedlist ,sortlist):
    new=[]
    for i in sortlist :
        if i not in consultedlist:
            new.append(i)
    for newi in new:
        sortlist.remove(newi)
    sorted(sortlist,key=consultedlist.index)
    for newi in  new:
        sortlist.append(newi)
    return sortlist
def randombooking(bookcount):
    countli=[]
    if 2<bookcount <6:
       endnum=math.ceil(bookcount/2)
    elif 6<bookcount<20 :
        endnum=math.ceil(bookcount/5)
    elif 20<bookcount<30:
        endnum=math.ceil(bookcount/7)
    elif 30<bookcount:
        endnum=math.ceil(bookcount/8)
    else:
        endnum=bookcount
    while bookcount>=2:
      a=random.randint(2,endnum)
      bookcount=bookcount-a
      countli.append(a)
      if bookcount==1:
          countli.append(bookcount)
          break
    return len(countli),sorted(countli,reverse=True)
def menulenbooking(bookcount,menulen):
    countli=[]
    allcount=bookcount
    amount=0
    endnum=math.ceil(bookcount/menulen)
    while bookcount>2:
      a=random.randint(2,endnum)
      bookcount=bookcount-a
      amount=amount+a
      countli.append(a)
      if len(countli)==menulen:
          countli[-1]=countli[-1]+allcount-amount
          break
      if len(countli)+1==menulen:
          if  allcount-amount>10:
              countli=[item+2 for item in countli ]
              countli.append(allcount-sum(countli))
          elif  allcount-amount>5:
              countli=[item+1 for item in countli ]
              countli.append(allcount-sum(countli))
          break
      if bookcount==1:
          countli.append(bookcount)
          break


    return len(countli),sorted(countli,reverse=True)

if __name__ == '__main__':
    wb = load_workbook(filename = r'H:\upupup\电商后台\自动化统计\订餐计划\menu.xlsx' )
    sheetnames = wb.get_sheet_names()
    ws = wb.get_sheet_by_name(sheetnames[0])
    wout=wb.create_sheet(0)
    supermenu=[]
    wout.cell(row=1,column=1).value="站点"
    wout.cell(row=1,column=2).value="供应商"
    wout.cell(row=1,column=3).value="菜名"
    wout.cell(row=1,column=4).value="数量"
    supermenudict={}

    for menurow in range(2,ws.max_row+1):
        supername=ws.cell(row=menurow,column=1).value
        if supername==None:
            continue
        supermenu.append(supername)
    supermenu=set(supermenu)
    supermenuli=list(supermenu)
    for superi in supermenuli:
       #supermenudict[superi]={}
       for rowi in range(2,ws.max_row+1):
           if ws.cell(row=rowi,column=1).value==superi:
               prodname=ws.cell(row=rowi,column=2).value
               supermenudict.setdefault(superi,[]).append(prodname)#供应商菜单字典
    superlist,shopsupdict,shopwasdict=shopyesdaysale()
    shopprodict=shopsupprosale()
    tomorrow= datetime.today() + timedelta(days=1)
    weekday=datetime.weekday(tomorrow)
    weekday=int(weekday)+1
    #weekday=3
    if weekday in [2,5]:#供应商每天预定量
        for shopkey in shopsupdict:
            for supk in shopsupdict[shopkey]:
                if shopkey in shopwasdict:
                    if supk in shopwasdict[shopkey]:
                        #if shopsupdict[shopkey][supk]>6:
                           # shopsupdict[shopkey][supk]=shopsupdict[shopkey][supk]-3
                       # if shopsupdict[shopkey][supk]>4:
                            #shopsupdict[shopkey][supk]=shopsupdict[shopkey][supk]-2
                        #else:
                            shopsupdict[shopkey][supk]=shopsupdict[shopkey][supk]-1
                else:
                    shopsupdict[shopkey][supk]=shopsupdict[shopkey][supk]
                if shopsupdict[shopkey][supk]<=2:
                    shopsupdict[shopkey][supk]=shopsupdict[shopkey][supk]+1

    else:
        for shopkey in shopsupdict:
            for supk in shopsupdict[shopkey]:
                if shopkey in shopwasdict:
                    if supk in shopwasdict[shopkey]:
                        shopsupdict[shopkey][supk]=shopsupdict[shopkey][supk]+1
                else:
                    shopsupdict[shopkey][supk]=shopsupdict[shopkey][supk]+1
                if shopsupdict[shopkey][supk]<=2:
                    shopsupdict[shopkey][supk]=shopsupdict[shopkey][supk]+2
    lunchscheduledict={}
    if set(supermenuli).issubset(set(superlist))==True:#是否有新供应商
        for shopkey in shopsupdict:
            lunchscheduledict[shopkey]={}
            for superi in supermenuli:
                lunchscheduledict[shopkey][superi]={}
                if superi in shopsupdict[shopkey]:
                    bookcount=shopsupdict[shopkey][superi]
                    ssprodli=shopprodict[shopkey][superi]
                    menuprodli=supermenudict[superi]
                    sortprodli=menusort(ssprodli,menuprodli)
                    prodmenucount,prodbookcountli=randombooking(bookcount)#菜品种类数量，每样菜品分配数量
                    if prodmenucount>len(menuprodli):
                       prodmenucount,prodbookcountli=menulenbooking(bookcount,len(menuprodli))
                    for prodi in range(0,prodmenucount):
                        boproname=sortprodli[prodi]
                        lunchscheduledict[shopkey][superi][boproname]=prodbookcountli[prodi]
    else:
        supermenuli=menusort(superlist,supermenuli)
        for shopkey in shopsupdict:
            lunchscheduledict[shopkey]={}
            newsbookc=0
            newsupi=len(supermenuli)-1
            for superi in supermenuli:

                lunchscheduledict[shopkey][superi]={}
                if superi in superlist:
                    newsupi=newsupi-1
                    if superi in shopsupdict[shopkey]:
                        bookcount=shopsupdict[shopkey][superi]
                        if bookcount>50:
                            bookcount=bookcount-10
                            newsbookc=newsbookc+10
                        elif bookcount>20:
                            bookcount=bookcount-5
                            newsbookc=newsbookc+5
                        elif bookcount>5:
                            bookcount=bookcount-2
                            newsbookc=newsbookc+2
                        ssprodli=shopprodict[shopkey][superi]
                        menuprodli=supermenudict[superi]
                        sortprodli=menusort(ssprodli,menuprodli)
                        prodmenucount,prodbookcountli=randombooking(bookcount)#菜品种类数量，每样菜品分配数量
                        if prodmenucount>len(sortprodli):
                           prodmenucount,prodbookcountli=menulenbooking(bookcount,len(sortprodli))
                        try:
                            for prodi in range(0,prodmenucount):
                                  boproname=sortprodli[prodi]
                                  lunchscheduledict[shopkey][superi][boproname]=prodbookcountli[prodi]
                        except  Exception:
                            print(len(sortprodli),shopkey,superi,ssprodli,sortprodli,prodmenucount,prodbookcountli)
                else:
                    menuprodli=supermenudict[superi]
                    if newsupi>0:
                        prodmenucount,prodbookcountli=randombooking(newsbookc/newsupi)#菜品种类数量，每样菜品分配数量
                    else:
                        prodmenucount,prodbookcountli=randombooking(newsbookc)
                    if prodmenucount>len(menuprodli):
                       if newsbookc<2:
                           newsbookc=newsbookc+2
                       prodmenucount,prodbookcountli=menulenbooking(newsbookc,len(menuprodli))
                    try:
                        for prodi in range(0,prodmenucount):
                            boproname=menuprodli[prodi]
                            lunchscheduledict[shopkey][superi][boproname]=prodbookcountli[prodi]
                    except  Exception:
                            print(len(sortprodli),shopkey,superi,sortprodli,prodmenucount,prodbookcountli)

    outrow=1
    for shopkey in lunchscheduledict:
        for superkey in lunchscheduledict[shopkey]:
            for prodkey in lunchscheduledict[shopkey][superkey]:
                outrow=outrow+1
                bookcount=lunchscheduledict[shopkey][superkey][prodkey]
                wout.cell(row=outrow,column=1).value=shopkey
                wout.cell(row=outrow,column=2).value=superkey
                wout.cell(row=outrow,column=3).value=prodkey
                wout.cell(row=outrow,column=4).value=bookcount
    wb.save(r'H:\upupup\电商后台\自动化统计\订餐计划\menu.xlsx')

