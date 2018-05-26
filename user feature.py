__author__ = 'vincent'
from openpyxl import Workbook
from openpyxl.reader.excel import load_workbook
import mysql.connector
import operator
import  datetime
from time import sleep
from tqdm import tqdm
import time
from datetime import datetime, timedelta




w= Workbook()
#创建excel工作薄
sheetnames = w.get_sheet_names()
ws = w.get_sheet_by_name(sheetnames[0])
ws.title="原始数据"
ws.cell(row=1,column=1).value="订单号"
ws.cell(row=1,column=2).value="用户ID"
ws.cell(row=1,column=3).value="用户名"
ws.cell(row=1,column=4).value="支付金额"
ws.cell(row=1,column=5).value="购买时间"


user="ning.wei15"
pwd="wn3333"
host="bizdb.zuolin.com"
db="ehbiz"
port="18306"

cnx=mysql.connector.connect(user=user,password=pwd,database=db,host=host,port=port)
cursor=cnx.cursor()

select_sql='''SELECT
a.`order_no`,a.`buyer_no`,a.`buyer_nick_name`,a.`paid_total`,a.`payment_time`
FROM
`tbl_order`a
WHERE
a.`payment_time`>"2017-09-01"
and a.`payment_time`<"2017-10-01"
AND a.`shop_no`IN("14622779991551986272","15021953740771020139","15021019137671284344")'''

cursor.execute(select_sql)
#数据库查询所有购买记录
alldata=cursor.fetchall()
alldata.sort(key=operator.itemgetter(4))
#根据购买金额排序
Arows=1
print("读取所有购买记录数据 ……")
#for userdata in alldata:
for Puserdata in tqdm(range(0,len(alldata))):

     userdata=alldata[Puserdata]
     Arows=Arows+1
     for uvalue in range(0,5):
        uservalue=userdata[uvalue]
        ws.cell(row=Arows,column=uvalue+1).value=uservalue
        time.sleep(0.0001)
#写入Excel表

wusb = w.create_sheet(0)
wusb.title="用户基础特征"
wusb.cell(row=1,column=1).value="用户ID"
wusb.cell(row=1,column=2).value="用户名"
wusb.cell(row=1,column=3).value="手机号"
wusb.cell(row=1,column=4).value="购买次数"
wusb.cell(row=1,column=5).value="购买总额"
wusb.cell(row=1,column=6).value="客单价"
wusb.cell(row=1,column=7).value="单次最高金额"
wusb.cell(row=1,column=8).value="最近购买时间"
wusb.cell(row=1,column=9).value="首次支付金额 "
wusb.cell(row=1,column=10).value="首次支付时间 "
wusb.cell(row=1,column=11).value="最近消费间隔"
wusb.cell(row=1,column=12).value="用户价值类型"
wusb.cell(row=1,column=13).value="用户综合评分"
print("计算用户基础特征 ……")

userbase_sql='''SELECT
 a.`buyer_no`,a.`buyer_nick_name`,a.`buyer_phone`,COUNT(a.`buyer_no`),SUM(a.`paid_total`),AVG(a.`paid_total`), MAX(a.`paid_total`),MAX(a.`payment_time`),
 (SELECT b.`paid_total` FROM `tbl_order`b
 WHERE b.`buyer_no`=a.`buyer_no`AND b.`payment_time`=MIN(a.`payment_time`))AS firstpaid,MIN(a.`payment_time`)
 FROM
 `tbl_order`a
 WHERE
 a.`payment_time`>"2017-09-01"
 and a.`payment_time`<"2017-10-01"
 AND a.`shop_no`IN("14622779991551986272","15021953740771020139","15021019137671284344")

 GROUP BY a.`buyer_no`
 ORDER BY COUNT(a.`buyer_no`)DESC'''

cursor.execute(userbase_sql)
usbdata=cursor.fetchall()
usbrow=1
usblist2=[]
usblist3=[]
usblist4=[]
usblist5=[]
usblist6=[]
usblist7=[]
usblist8=[]
usblist9=[]
usblist10=[]
maxMapaid=max(int(x[6]) for x  in usbdata)
minMipaid=min(int(x[6]) for x  in usbdata)
maxfre=max(int(x[3]) for x  in usbdata)
minfre=min(int(x[3]) for x  in usbdata)
maxavg=max(int(x[5]) for x  in usbdata)
minavg=min(int(x[5]) for x  in usbdata)
maxinterval=""
mininterval=""
'''contraday=datetime.strptime("2016-8-01",'%Y-%m-%d')
#for usb in usbdata:
for realdata in usbdata:
    if realdata[9]>contraday:
        usbdata.remove(realdata)'''
for Pusb in tqdm(range(0,len(usbdata))):
    usb=usbdata[Pusb]
    usbrow=usbrow+1
    for usbvalue in range(0,10):
        userbasevalue=usb[usbvalue]
        wusb.cell(row=usbrow,column=usbvalue+1).value=userbasevalue
        if usbvalue==3:
            if userbasevalue>1:
                usblist2.append((userbasevalue,int(usb[usbvalue+2])))
                #加入购买次数大于1的用户表
                if userbasevalue>2:
                    usblist3.append((userbasevalue,int(usb[usbvalue+2])))
                    if userbasevalue>3:
                       usblist4.append((userbasevalue,int(usb[usbvalue+2])))
                       if userbasevalue>4:
                           usblist5.append((userbasevalue,int(usb[usbvalue+2])))
                           if userbasevalue>5:
                               usblist6.append((userbasevalue,int(usb[usbvalue+2])))
                               if userbasevalue>6:
                                   usblist7.append((userbasevalue,int(usb[usbvalue+2])))
                                   if userbasevalue>7:
                                       usblist8.append((userbasevalue,int(usb[usbvalue+2])))
                                       if userbasevalue>8:
                                           usblist9.append((userbasevalue,int(usb[usbvalue+2])))
                                           if userbasevalue>9:
                                                usblist10.append((userbasevalue,int(usb[usbvalue+2])))
        if usbvalue==7:
            lasttime=userbasevalue
            recentday=(datetime.now()-lasttime).days

            wusb.cell(row=usbrow,column=11).value=recentday
            if maxinterval=="":
                maxinterval=recentday
                mininterval=recentday
            else:
                if maxinterval<recentday:
                    maxinterval=recentday
                if mininterval>recentday:
                    mininterval=recentday

    #print(usb)

for userm in range(2,wusb.max_row):
    oncepaid=int(wusb.cell(row=userm,column=7).value)
    interval=int(wusb.cell(row=userm,column=11).value)
    frequence=int(wusb.cell(row=userm,column=4).value)
    avgpaid=int(wusb.cell(row=userm,column=6).value)
    oncepaidmark=(oncepaid-minMipaid)/(maxMapaid-minMipaid)
    intervalmark=(maxinterval-interval)/(maxinterval-mininterval)
    frequencymark=(frequence-minfre)/(maxfre-minfre)
    avgpaidmark=(avgpaid-minavg)/(maxavg-minavg)
    usermark=10*(oncepaidmark*0.11+intervalmark*0.22+avgpaidmark*0.22+frequencymark*0.45)
    wusb.cell(row=userm,column=13).value=usermark
wusact = w.create_sheet(0)
wusact.title="用户购买行为"
wusact.cell(row=1,column=1).value="指标"
wusact.cell(row=1,column=2).value="客户数"
wusact.cell(row=1,column=3).value="占比"
wusact.cell(row=1,column=4).value="复购率"
wusact.cell(row=1,column=5).value="客单价"

wusact.cell(row=2,column=1).value="一次购买"
wusact.cell(row=2,column=2).value=len(usbdata)
#购买过的用户总数
wusact.cell(row=2,column=3).value=format(len(usbdata)/len(usbdata),".0%")
#至少一次购买用户占比
wusact.cell(row=2,column=4).value=format(len(usblist2)/len(usbdata),".0%")
#至少一次购买用户复购率
wusact.cell(row=2,column=5).value=(sum(int(x[5]) for x  in usbdata))/len(usbdata)
#至少一次购买过用户客单价

wusact.cell(row=3,column=1).value="二次购买"
wusact.cell(row=3,column=2).value=len(usblist2)
#至少三次购买过的用户总数
wusact.cell(row=3,column=3).value=format(len(usblist2)/len(usbdata),".0%")
#至少二次购买用户占比
wusact.cell(row=3,column=4).value=format(len(usblist3)/len(usblist2),".0%")
#至少二次购买用户复购率
wusact.cell(row=3,column=5).value=(sum(int(x[1]) for x  in usblist2))/len(usblist2)
#至少二次购买过用户客单价

wusact.cell(row=4,column=1).value="三次购买"
wusact.cell(row=4,column=2).value=len(usblist3)
#至少三次购买过的用户总数
wusact.cell(row=4,column=3).value=format(len(usblist3)/len(usbdata),".0%")
#至少三次购买用户占比
wusact.cell(row=4,column=4).value=format(len(usblist4)/len(usblist3),".0%")
#至少三次购买用户复购率
wusact.cell(row=4,column=5).value=(sum(int(x[1]) for x  in usblist3))/len(usblist3)
#至少三次购买过用户客单价


wusact.cell(row=5,column=1).value="四次购买"
wusact.cell(row=5,column=2).value=len(usblist4)
#至少四次购买过的用户总数
wusact.cell(row=5,column=3).value=format(len(usblist4)/len(usbdata),".0%")
#至少四次购买用户占比
if len(usblist7)!=0:
    wusact.cell(row=5,column=4).value=format(len(usblist5)/len(usblist4),".0%")
    #至少四次购买用户复购率
    wusact.cell(row=5,column=5).value=(sum(int(x[1]) for x  in usblist4))/len(usblist4)
    #至少四次购买过用户客单价

    wusact.cell(row=6,column=1).value="五次购买"
    wusact.cell(row=6,column=2).value=len(usblist5)
    #至少五次购买过的用户总数
    wusact.cell(row=6,column=3).value=format(len(usblist5)/len(usbdata),".0%")
    #至少五次购买用户占比
    wusact.cell(row=6,column=4).value=format(len(usblist6)/len(usblist5),".0%")
    #至少五次购买用户复购率

    wusact.cell(row=6,column=5).value=(sum(int(x[1]) for x  in usblist5))/len(usblist5)
    #至少五次购买过用户客单价

    wusact.cell(row=7,column=1).value="六次购买"
    wusact.cell(row=7,column=2).value=len(usblist6)
    #至少五次购买过的用户总数
    wusact.cell(row=7,column=3).value=format(len(usblist6)/len(usbdata),".0%")
    #至少五次购买用户占比
    wusact.cell(row=7,column=4).value=format(len(usblist7)/len(usblist6),".0%")
    #至少五次购买用户复购率
    wusact.cell(row=7,column=5).value=(sum(int(x[1]) for x  in usblist6))/len(usblist6)
    #至少五次购买过用户客单价

    wusact.cell(row=8,column=1).value="七次购买"
    wusact.cell(row=8,column=2).value=len(usblist7)
    #至少五次购买过的用户总数
    wusact.cell(row=8,column=3).value=format(len(usblist7)/len(usbdata),".0%")
        #至少五次购买用户占比
if len(usblist7)!=0:
        wusact.cell(row=8,column=4).value=format(len(usblist8)/len(usblist7),".0%")
        #至少五次购买用户复购率
        wusact.cell(row=8,column=5).value=(sum(int(x[1]) for x  in usblist7))/len(usblist7)
        #至少五次购买过用户客单价

        wusact.cell(row=9,column=1).value="8次购买"
        wusact.cell(row=9,column=2).value=len(usblist8)
        #至少五次购买过的用户总数
        wusact.cell(row=9,column=3).value=format(len(usblist8)/len(usbdata),".0%")
        #至少五次购买用户占比
if len(usblist8)!=0:
        wusact.cell(row=9,column=4).value=format(len(usblist9)/len(usblist8),".0%")
        #至少五次购买用户复购率
        wusact.cell(row=9,column=5).value=(sum(int(x[1]) for x  in usblist8))/len(usblist8)
        #至少五次购买过用户客单价

        wusact.cell(row=10,column=1).value="9次购买"
        wusact.cell(row=10,column=2).value=len(usblist9)
        #至少五次购买过的用户总数
        wusact.cell(row=10,column=3).value=format(len(usblist9)/len(usbdata),".0%")
        #至少五次购买用户占比
if len(usblist9)!=0:
        wusact.cell(row=10,column=4).value=format(len(usblist10)/len(usblist9),".0%")
        #至少五次购买用户复购率
        wusact.cell(row=10,column=5).value=(sum(int(x[1]) for x  in usblist9))/len(usblist9)
        #至少五次购买过用户客单价

        wusact.cell(row=11,column=1).value="10次以上"
        wusact.cell(row=11,column=2).value=len(usblist10)
        #至少五次购买过的用户总数
        wusact.cell(row=11,column=3).value=format(len(usblist10)/len(usbdata),".0%")
        #至少五次购买用户占比
        wusact.cell(row=11,column=4).value=""
        #至少五次购买用户复购率
        wusact.cell(row=11,column=5).value=(sum(int(x[1]) for x  in usblist10))/len(usblist10)
        #至少五次购买过用户客单价

wusrfm = w.create_sheet(0)
wusrfm.title="用户价值分层"
monetary=(sum(int(x[4]) for x  in usbdata))/len(usbdata)
frequency=(sum(int(x[3]) for x  in usbdata))/len(usbdata)
recency=14
keyvalue=[]
keydevelop=[]
keykeep=[]
keyretain=[]
genvalue=[]
gendevelop=[]
genkeep=[]
genretain=[]
for rfmuser in tqdm(range(2,wusb.max_row+1)):
    if wusb.cell(row=rfmuser,column=5).value>monetary:
        if wusb.cell(row=rfmuser,column=4).value>frequency:
            if wusb.cell(row=rfmuser,column=11).value<recency:
                wusb.cell(row=rfmuser,column=12).value="重要价值用户"
                keyvalue.append((wusb.cell(row=rfmuser,column=2).value,wusb.cell(row=rfmuser,column=3).value,int(wusb.cell(row=rfmuser,column=9).value)))
            else:
                wusb.cell(row=rfmuser,column=12).value="重要保持用户"
                keykeep.append((wusb.cell(row=rfmuser,column=2).value,wusb.cell(row=rfmuser,column=3).value,int(wusb.cell(row=rfmuser,column=9).value)))
        else:
            if wusb.cell(row=rfmuser,column=11).value<recency:
                wusb.cell(row=rfmuser,column=12).value="重要发展用户"
                keydevelop.append((wusb.cell(row=rfmuser,column=2).value,wusb.cell(row=rfmuser,column=3).value,int(wusb.cell(row=rfmuser,column=9).value)))
            else:
                wusb.cell(row=rfmuser,column=12).value="重要挽回用户"
                keyretain.append((wusb.cell(row=rfmuser,column=2).value,wusb.cell(row=rfmuser,column=3).value,int(wusb.cell(row=rfmuser,column=9).value)))
    else:
        if wusb.cell(row=rfmuser,column=4).value>frequency:
            if wusb.cell(row=rfmuser,column=11).value<recency:
                wusb.cell(row=rfmuser,column=12).value="一般价值用户"
                genvalue.append((wusb.cell(row=rfmuser,column=2).value,wusb.cell(row=rfmuser,column=3).value,int(wusb.cell(row=rfmuser,column=9).value)))
            else:
                wusb.cell(row=rfmuser,column=12).value="一般保持用户"
                genkeep.append((wusb.cell(row=rfmuser,column=2).value,wusb.cell(row=rfmuser,column=3).value,int(wusb.cell(row=rfmuser,column=9).value)))
        else:
            if wusb.cell(row=rfmuser,column=11).value<recency:
                wusb.cell(row=rfmuser,column=12).value="一般发展用户"
                gendevelop.append((wusb.cell(row=rfmuser,column=2).value,wusb.cell(row=rfmuser,column=3).value,int(wusb.cell(row=rfmuser,column=9).value)))
            else:
                wusb.cell(row=rfmuser,column=12).value="一般挽回用户"
                genretain.append((wusb.cell(row=rfmuser,column=2).value,wusb.cell(row=rfmuser,column=3).value,int(wusb.cell(row=rfmuser,column=9).value)))
wusrfm.cell(row=1,column=1).value="客户类型"
wusrfm.cell(row=1,column=2).value="客户数"
wusrfm.cell(row=1,column=3).value="客户占比"
wusrfm.cell(row=1,column=4).value="平均首次消费金额"
wusrfm.cell(row=2,column=1).value="重要价值用户"
wusrfm.cell(row=2,column=2).value=len(keyvalue)
wusrfm.cell(row=2,column=3).value=format(len(keyvalue)/len(usbdata),".0%")
wusrfm.cell(row=2,column=4).value=(sum(int(x[2]) for x  in keyvalue))/len(keyvalue)
if len(keykeep)!=0:
   wusrfm.cell(row=3,column=1).value="重要保持用户"
   wusrfm.cell(row=3,column=2).value=len(keykeep)
   wusrfm.cell(row=3,column=3).value=format(len(keykeep)/len(usbdata),".0%")
   wusrfm.cell(row=3,column=4).value=(sum(int(x[2]) for x  in keykeep))/len(keykeep)
if len(keydevelop)!=0:
    wusrfm.cell(row=4,column=1).value="重要发展用户"
    wusrfm.cell(row=4,column=2).value=len(keydevelop)
    wusrfm.cell(row=4,column=3).value=format(len(keydevelop)/len(usbdata),".0%")
    wusrfm.cell(row=4,column=4).value=(sum(int(x[2]) for x  in keydevelop))/len(keydevelop)
if len(keyretain)!=0:
    wusrfm.cell(row=5,column=1).value="重要挽回用户"
    wusrfm.cell(row=5,column=2).value=len(keyretain)
    wusrfm.cell(row=5,column=3).value=format(len(keyretain)/len(usbdata),".0%")
    wusrfm.cell(row=5,column=4).value=(sum(int(x[2]) for x  in keyretain))/len(keyretain)
if len(genvalue)!=0:
    wusrfm.cell(row=6,column=1).value="一般价值用户"
    wusrfm.cell(row=6,column=2).value=len(genvalue)
    wusrfm.cell(row=6,column=3).value=format(len(genvalue)/len(usbdata),".0%")
    wusrfm.cell(row=6,column=4).value=(sum(int(x[2]) for x  in genvalue))/len(genvalue)
if len(genkeep) !=0:
    wusrfm.cell(row=7,column=1).value="一般保持用户"
    wusrfm.cell(row=7,column=2).value=len(genkeep)
    wusrfm.cell(row=7,column=3).value=format(len(genkeep)/len(usbdata),".0%")
    wusrfm.cell(row=7,column=4).value=(sum(int(x[2]) for x  in genkeep))/len(genkeep)
wusrfm.cell(row=8,column=1).value="一般发展用户"
wusrfm.cell(row=8,column=2).value=len(gendevelop)
wusrfm.cell(row=8,column=3).value=format(len(gendevelop)/len(usbdata),".0%")
wusrfm.cell(row=8,column=4).value=(sum(int(x[2]) for x  in gendevelop))/len(gendevelop)
if len(genretain)!=0:
    wusrfm.cell(row=9,column=1).value="一般挽回用户"
    wusrfm.cell(row=9,column=2).value=len(genretain)
    wusrfm.cell(row=9,column=3).value=format(len(genretain)/len(usbdata),".0%")
    wusrfm.cell(row=9,column=4).value=(sum(int(x[2]) for x  in genretain))/len(genretain)




w.save(r"H:\upupup\电商后台\自动化统计\1009金融用户.csv")
print(maxinterval , mininterval , maxavg,minavg,maxfre,minfre,maxMapaid,minMipaid)
cursor.close()
cnx.commit()
cnx.close()