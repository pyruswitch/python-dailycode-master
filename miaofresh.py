__author__ = 'vincent'
# -*- coding:utf-8 -*-
from openpyxl import Workbook
import urllib
from pyquery import PyQuery as Pq
from selenium import webdriver
import time
import StrUtil

w= Workbook()
#创建excel工作薄
ws= w.create_sheet(0)
#创建Excel工作表

def get_page_content_str( url):
    time.sleep(1)

    try:
            print("现在开始抓取" + url)
            headers = {'User-Agent': 'Mozilla/5.0 (Windows NT 6.1; WOW64; rv:23.0) Gecko/20100101 Firefox/23.0'}
            #伪装浏览器
            request = urllib.request.Request(url=url, headers=headers)
            #构造请求
            m_fp = urllib.request.urlopen(request, timeout=500)
            #访问网站获取源码
            html_str = m_fp.read().decode('gbk')
            #读取源码，该网站使用的编码方式是utf-8
            if html_str == None:
                html_str= m_fp.read().decode('utf-8')
                if html_str == None:
                    html_str = m_fp.read().decode('gb2312')
            m_fp.close()
            return html_str
    except Exception:
            print(Exception)
#定义抓取网页源码函数

baseurl="https://list.tmall.com/search_product.htm?type=pc&q=%CB%AE%B9%FB&totalPage=100&sort=s&style=g&from=miao.index.pc_1_searchbutton&jumpto={}#J_Filter"
#目标网站URL
j=1
ws.cell(row=1,column=1).value="名称"
ws.cell(row=1,column=2).value="厂家"
ws.cell(row=1,column=3).value="厂家联系方式"
ws.cell(row=1,column=4).value="规格"
ws.cell(row=1,column=5).value="价格"
ws.cell(row=1,column=6).value="销量"
ws.cell(row=1,column=7).value="评价数"
#设置excel单元格表头
for i in range (1,30):
    url=baseurl.format(i)
    print("现在开始抓取第{}页".format(i))
    html_str=get_page_content_str(url)
    doc=Pq(html_str)
    fruits=doc("#J_ItemList > div")
    for fruit in fruits:

        fruitname=Pq(fruit)("div > p.productTitle > a").attr("title")
        fruiturl="https:"+Pq(fruit)("a").attr("href")
        fruitpri=Pq(fruit)("div > p.productPrice > em").attr("title")
        fruitsale=Pq(fruit)(" div > p.productStatus > span:nth-child(1) > em").text()
        fruitfavr=Pq(fruit)(" div > p.productStatus > span:nth-child(2) > a").text()
        detailhtml=get_page_content_str(fruiturl)
        if detailhtml==None:
         try:
            headers = {'User-Agent': 'Mozilla/5.0 (Windows NT 6.1; WOW64; rv:23.0) Gecko/20100101 Firefox/23.0'}
            #伪装浏览器
            request = urllib.request.Request(url=fruiturl, headers=headers)
            #构造请求
            m_fp = urllib.request.urlopen(request, timeout=500)
            #访问网站获取源码
            detailhtml = m_fp.read().decode('utf-8')
         except Exception:
            print(Exception)
            if detailhtml==None:
             try:
              headers = {'User-Agent': 'Mozilla/5.0 (Windows NT 6.1; WOW64; rv:23.0) Gecko/20100101 Firefox/23.0'}
              #伪装浏览器
              request = urllib.request.Request(url=fruiturl, headers=headers)
              #构造请求
              m_fp = urllib.request.urlopen(request, timeout=500)
              #访问网站获取源码
              detailhtml = m_fp.read().decode('gbk')
             except Exception:
              print(Exception)
        detail=Pq(detailhtml)
        fruitdetail=detail("#J_AttrUL").text()
        fruitbrand=StrUtil.substr(fruitdetail,"厂名："," ")
        fruitphone=StrUtil.substr(fruitdetail,"联系方式："," ")
        fruitspec=StrUtil.substr(fruitdetail,"净含量:"," ")
        #fruitsale=detail("#J_DetailMeta > div.tm-clear > div.tb-property > div ").text()
        #fruitfavr=detail("#J_Reviews > div > div.rate-header > div.rate-score > strong").text()
        j=j+1
        ws.cell(row=j,column=1).value=fruitname
        ws.cell(row=j,column=2).value=fruitbrand
        ws.cell(row=j,column=3).value=fruitphone
        ws.cell(row=j,column=4).value=fruitspec
        ws.cell(row=j,column=5).value=fruitpri
        ws.cell(row=j,column=6).value=fruitsale
        ws.cell(row=j,column=7).value=fruitfavr
        #写入excel单元格
        try:
         print(fruitname+"-"+"-"+fruitbrand+"-"+fruitphone+"-"+fruitspec+"-"+fruitpri+"-"+fruitsale+"-"+fruitfavr)
         #在控制台输出结果
        except Exception:
            print(Exception)
w.save(r'd:\pythontest\电商\onefruits.xlsx')
#保存到本地excel文件