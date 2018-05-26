__author__ = 'vincent'

from openpyxl import Workbook
from openpyxl.reader.excel import load_workbook
import mysql.connector
import operator
import  datetime





w= Workbook()
#创建excel工作薄
sheetnames = w.get_sheet_names()
ws = w.get_sheet_by_name(sheetnames[0])
#ws= w.create_sheet(0)
ws.cell(row=1,column=1).value="店铺"
ws.cell(row=1,column=2).value="订单号"
ws.cell(row=1,column=3).value="支付方式"
ws.cell(row=1,column=4).value="品类"
ws.cell(row=1,column=5).value="供应商 "
ws.cell(row=1,column=6).value="商品名"
ws.cell(row=1,column=7).value="规格"
ws.cell(row=1,column=8).value="成本价"
ws.cell(row=1,column=9).value="售价"
ws.cell(row=1,column=10).value="数量 "
ws.cell(row=1,column=11).value="总金额 "
ws.cell(row=1,column=12).value="折扣金额 "
ws.cell(row=1,column=13).value="支付金额 "
ws.cell(row=1,column=14).value="活动优惠 "
ws.cell(row=1,column=15).value="活动描述 "
ws.cell(row=1,column=16).value="支付时间"
ws.cell(row=1,column=17).value="下单时间段"

user="ning.wei12"
pwd="wn3633"
host="bizdb.zuolin.com"
db="ehbiz"
port="18306"

cnx=mysql.connector.connect(user=user,password=pwd,database=db,host=host,port=port)
cursor=cnx.cursor()

select_sql='''SELECT DISTINCT
c.`shop_name`,a.`order_no`,a.`pay_type`,e.`cat_name`,e.`supplier_name`,b.`prod_name`,e.`model`,d.`current_prime_price`,b.`price`,b.`quantity`,a.`price_total`,a.`discount_total`,a.`paid_total`,a.`activity_benefit_amount`,f.`activity_name`,a.`payment_time`
FROM
`tbl_order`a
LEFT JOIN `tbl_shop_info`c
ON a.`shop_no`=c.`shop_no`
LEFT JOIN  `tbl_order_item`b
ON b.`order_no`=a.`order_no`
LEFT JOIN `tbl_model`d
ON b.`prod_no`=d.`model_no`
LEFT JOIN `tbl_commodity`e
ON d.`commo_no`=e.`commo_no`
LEFT JOIN `tbl_order_activity_contact`f
ON a.`order_no`=f.`order_no`
LEFT JOIN `tbl_activity`g
ON g.activity_no=f.activity_no
WHERE
 a.`payment_time`>"2016-10-01 00:00:00"
AND a.`payment_time`<"2016-11-01 00:00:00"
AND a.`basic_state`!=1'''
cursor.execute(select_sql)
alldata=cursor.fetchall()
alldata.sort(key=operator.itemgetter(1))
rows=1
for proddata in alldata:
    rows=rows+1
    for Provalue in range(0,16):
        provalue=proddata[Provalue]
        ws.cell(row=rows,column=Provalue+1).value=provalue
        if Provalue==15:
            ostime=provalue.strftime("%H")
            times=str(ostime)+"—"+str(int(ostime)+1)
            ws.cell(row=rows,column=17).value=times
    print(proddata)
for i in range(2,ws.max_row+1):
    if "+"in ws.cell(row=i,column=6).value:
        ws.cell(row=i,column=4).value="套餐"
wg=w.create_sheet(0)
wg.title="购买行为"
wg.cell(row=1,column=1).value="购买行为"
wg.cell(row=1,column=2).value="次数"
wg.cell(row=1,column=3).value="组合详情"
wg.cell(row=1,column=4).value="次数"
groupli=[]
portion=0
group=0
package=0
j=2
maxrow=ws.max_row+1
while( j <maxrow):
    if ws.cell(row=j,column=2).value==ws.cell(row=j+1,column=2).value:
        groups=str(ws.cell(row=j,column=4).value)+"&"+str(ws.cell(row=j+1,column=4).value)
        regroups=str(ws.cell(row=j+1,column=4).value)+"&"+str(ws.cell(row=j,column=4).value)
        if regroups in groupli:
            groups=regroups
        if j+1<maxrow-1:
            for j1 in range (j+2,maxrow):
                if ws.cell(row=j,column=2).value==ws.cell(row=j1,column=2).value:
                    catname=ws.cell(row=j1,column=4).value
                    if str(catname)  in groups:
                       j=j1
                       if j==maxrow-1:
                          j=maxrow
                       else:
                        continue
                    else:
                        groups=groups+"&"+ws.cell(row=j1,column=4).value

                else:
                    if "套餐"in groups:
                        package=package+1
                    else:
                        groupli.append(groups)
                        group=group+1
                    j=j1
                    break
        else :
            if "套餐"in groups:
                package=package+1
                j=j+1
            else:
                groupli.append(groups)
                group=group+1
                j=j+1
        continue
    else:
        if  ws.cell(row=j,column=4).value=="套餐":
             package=package+1
        else:portion=portion+1
        j=j+1
        #print(j)
        #print("分类执行到第"+str(j)+"行啦")
        continue
print("分类结束")
wg.cell(row=2,column=1).value="单独"
wg.cell(row=2,column=2).value=portion
wg.cell(row=3,column=1).value="组合"
wg.cell(row=3,column=2).value=group
wg.cell(row=4,column=1).value="套餐"
wg.cell(row=4,column=2).value=package
grouprow=2
groups=set(groupli)
for item in groups:

    wg.cell(row=grouprow,column=3).value=item
    wg.cell(row=grouprow,column=4).value=groupli.count(item)
    grouprow=grouprow+1
    #print("组合详情执行到第"+str(grouprow)+"行啦")
wsale=w.create_sheet(0)
wsale.title="总体销量统计"
salelist=[]
supplierlist=["套餐-其他商品"]
wsale.cell(row=1,column=1).value="商品名"
wsale.cell(row=1,column=2).value="销量"
wsale.cell(row=1,column=3).value="毛利"
wsale.cell(row=1,column=4).value="供应商"
wsale.cell(row=1,column=5).value="结算价"
wsale.cell(row=1,column=6).value="销售额"
wsale.cell(row=1,column=7).value="销量"
togoldrow=2
tosupperow=3
wsale.cell(row=2,column=4).value="套餐-其他商品"
wsale.cell(row=tosupperow,column=5).value=0
wsale.cell(row=2,column=5).value=0
wsale.cell(row=2,column=7).value=0
onlinelunch_dict={"房咚":"14","乐盒饭":"12.82","鲸鱼":"14","星宝外卖":"12.5","船说":"14.5","鼎丰圣麦":"13","佳味外卖":"13.5"}
linelunch_dict={"房咚":"14","乐盒饭":"13.5","鲸鱼":"14","星宝外卖":"12.5","船说":"14.5","鼎丰圣麦":"13","佳味外卖":"13.5"}
for goldrow in range(2,maxrow):
    goldname=ws.cell(row=goldrow,column=6).value
    goldsale=ws.cell(row=goldrow,column=10).value
    goldcost=ws.cell(row=goldrow,column=8).value
    goldprice=ws.cell(row=goldrow,column=9).value
    suppliername=ws.cell(row=goldrow,column=5).value
    goldsamount=goldprice*goldsale
    if ws.cell(row=goldrow,column=4).value=="套餐":
       if suppliername in onlinelunch_dict:
          if ws.cell(row=goldrow,column=3).value==1:
              lungoldcost=float(onlinelunch_dict[suppliername])
          else:
              lungoldcost=float(linelunch_dict[suppliername])
          if suppliername not in supplierlist:
              wsale.cell(row=tosupperow,column=4).value=suppliername
              wsale.cell(row=tosupperow,column=5).value=lungoldcost*goldsale
              wsale.cell(row=tosupperow,column=6).value=goldsamount
              wsale.cell(row=tosupperow,column=7).value=goldsale
              tosupperow=tosupperow+1
              supplierlist.append(suppliername)
          else:
              indexsupplier=supplierlist.index(suppliername)+2
              wsale.cell(row=indexsupplier,column=5).value=wsale.cell(row=indexsupplier,column=5).value+lungoldcost*goldsale
              wsale.cell(row=indexsupplier,column=6).value=wsale.cell(row=indexsupplier,column=6).value+goldsamount
              wsale.cell(row=indexsupplier,column=7).value=wsale.cell(row=indexsupplier,column=7).value+goldsale
          wsale.cell(row=2,column=5).value=wsale.cell(row=2,column=5).value+(goldcost-lungoldcost)*goldsale
          wsale.cell(row=2,column=7).value=wsale.cell(row=2,column=7).value+goldsale
       else:
           if suppliername not in supplierlist:
              wsale.cell(row=tosupperow,column=4).value=suppliername
              wsale.cell(row=tosupperow,column=5).value=goldcost*goldsale
              wsale.cell(row=tosupperow,column=6).value=goldsamount
              wsale.cell(row=tosupperow,column=7).value=goldsale
              tosupperow=tosupperow+1
              supplierlist.append(suppliername)
           else:
              indexsupplier=supplierlist.index(suppliername)+2
              wsale.cell(row=indexsupplier,column=5).value=wsale.cell(row=indexsupplier,column=5).value+goldcost*goldsale
              wsale.cell(row=indexsupplier,column=6).value=wsale.cell(row=indexsupplier,column=6).value+goldsamount
              wsale.cell(row=indexsupplier,column=7).value=wsale.cell(row=indexsupplier,column=7).value+goldsale
    else:
           if suppliername not in supplierlist:
              wsale.cell(row=tosupperow,column=4).value=suppliername
              wsale.cell(row=tosupperow,column=5).value=goldcost*goldsale
              wsale.cell(row=tosupperow,column=6).value=goldsamount
              wsale.cell(row=tosupperow,column=7).value=goldsale
              tosupperow=tosupperow+1
              supplierlist.append(suppliername)
           else:
              indexsupplier=supplierlist.index(suppliername)+2
              wsale.cell(row=indexsupplier,column=5).value=wsale.cell(row=indexsupplier,column=5).value+goldcost*goldsale
              wsale.cell(row=indexsupplier,column=6).value=wsale.cell(row=indexsupplier,column=6).value+goldsamount
              wsale.cell(row=indexsupplier,column=7).value=wsale.cell(row=indexsupplier,column=7).value+goldsale
    if goldname not in salelist:

        wsale.cell(row=togoldrow,column=1).value=goldname
        wsale.cell(row=togoldrow,column=2).value=goldsale
        salelist.append(goldname)
        profit=goldsale*(goldprice-goldcost)
        wsale.cell(row=togoldrow,column=3).value=profit
        togoldrow=togoldrow+1
    else:
        indexrow=salelist.index(goldname)+2
        goldsale=wsale.cell(row=indexrow,column=2).value+ws.cell(row=goldrow,column=10).value
        wsale.cell(row=indexrow,column=2).value=goldsale
        profit=wsale.cell(row=indexrow,column=3).value+ws.cell(row=goldrow,column=10).value*(goldprice-goldcost)
        wsale.cell(row=indexrow,column=3).value=profit
    #print("结算执行到第"+str(togoldrow)+"行啦")
ssortlist=[]
for salesort in range(2,wsale.max_row+1):
    ssortlist.append((wsale.cell(row=salesort,column=1).value,wsale.cell(row=salesort,column=2).value,wsale.cell(row=salesort,column=3).value))
ssortlist.sort(key=operator.itemgetter(1),reverse=True)
sortrow=1
for ssort in ssortlist:
    sortrow=sortrow+1
    for ssortele in range(0,3):
        selement=ssort[ssortele]
        wsale.cell(row=sortrow,column=ssortele+1).value=selement
        #print("销量执行到第"+str(sortrow)+"行啦")
wt=w.create_sheet(0)
wt.title="午餐下单时间段"
wt.cell(row=1,column=1).value="时间段"
wt.cell(row=1,column=2).value="午餐下单量"
orderli=[]
timesli=[]
timerow=1
for timele in range(2,ws.max_row+1):
    if ws.cell(row=timele,column=4).value=="盒饭":
        if ws.cell(row=timele,column=2)not in orderli:
            timesli.append(ws.cell(row=timele,column=17).value)
            orderli.append(ws.cell(row=timele,column=2).value)
        else:continue
    else:continue
    #print("时间段执行到第"+str(timerow)+"行啦")
timegroup=set(timesli)
for itemtime in timegroup:
   timerow=timerow+1
   wt.cell(row=timerow,column=1).value=itemtime
   wt.cell(row=timerow,column=2).value=timesli.count(itemtime)

w.save(r"H:\upupup\电商后台\自动化统计\0825.xlsx")

cursor.close()
cnx.commit()
cnx.close()