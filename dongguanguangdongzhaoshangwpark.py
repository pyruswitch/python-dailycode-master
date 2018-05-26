__author__ = 'vincent'
from openpyxl import Workbook
import urllib
from pyquery import PyQuery as Pq
from selenium import webdriver
import time
import StrUtil

w= Workbook()
#创建excel工作薄
sheetnames = w.get_sheet_names()
ws= w.get_sheet_by_name(sheetnames[0])
#创建Excel工作表

def get_page_content_str( url):
    time.sleep(1)

    try:
            print("现在开始抓取" + url)
            headers = {'User-Agent': 'Mozilla/5.0 (Windows NT 6.1; WOW64; rv:23.0) Gecko/20100101 Firefox/23.0'}
            #伪装浏览器
            request = urllib.request.Request(url=url, headers=headers)
            #构造请求
            m_fp = urllib.request.urlopen(request, timeout=500)
            #访问网站获取源码
            html_str = m_fp.read().decode('utf-8')
            #读取源码，该网站使用的编码方式是utf-8

            return html_str
    except Exception:
            print(Exception)
            try:
                 request = urllib.request.Request(url=url, headers=headers)
                 m_fp = urllib.request.urlopen(request, timeout=500)
                 html_str= m_fp.read().decode('gbk')
                 return html_str
            except Exception:
             print(Exception)
             try:
                 request = urllib.request.Request(url=url, headers=headers)
                 m_fp = urllib.request.urlopen(request, timeout=500)
                 html_str= m_fp.read().decode('gb2312')
                 return html_str
             except Exception:
                      print(Exception)
            m_fp.close()
#定义抓取网页源码函数
baseurl="http://gd.zhaoshang.net/dongguan/yuanqu/0-0/{}"
#目标网站URL
j=1
ws.cell(row=1,column=1).value="园区名"
ws.cell(row=1,column=2).value="产业集群"
ws.cell(row=1,column=3).value="简介"
#设置excel单元格表头
for i in range (1,6):
    url=baseurl.format(i)
    print("现在开始抓取第{}页".format(i))
    html_str=get_page_content_str(url)
    doc=Pq(html_str)
    parks=doc("#content_lazy > div:nth-child(13) > div.carry_left > div > div.searchlist_item_1 > ul > li")
    for park in parks:
        j=j+1
        parkname=Pq(park)("dl > dt > a").text()
        parkhref=Pq(park)("dl > dt > a").attr("href")+"/intro"
        kind=Pq(park)(" dl > dd.parkinduscluster").text()
        parkdetail=get_page_content_str(parkhref)
        parkdoc=Pq(parkdetail)
        abs=parkdoc("body > div> div.container > div.cc16 > div.cc16_2 ").text()
        ws.cell(row=j,column=1).value=parkname
        ws.cell(row=j,column=2).value=kind
        ws.cell(row=j,column=3).value=abs
        #写入excel单元格
        print(parkname+"-"+kind)
        #在控制台输出结果
w.save(r'D:\pythontest\园区\东莞招商网3.xlsx')
#保存到本地excel文件

