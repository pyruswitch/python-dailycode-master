__author__ = 'vincent'
# -*- coding:utf-8 -*-
from openpyxl import Workbook
from openpyxl.reader.excel import load_workbook
from openpyxl.writer.excel import ExcelWriter
from openpyxl.cell import get_column_letter

wb = load_workbook(r'H:\upupup\电商后台\结算管理\乐盒饭\5月\5月午餐核对.xlsx')
sheetnames = wb.get_sheet_names()
wo = wb.get_sheet_by_name(sheetnames[0])
w= Workbook()
ws= w.create_sheet(0)
ws.cell(row=1,column=1).value="时间"
ws.cell(row=1,column=2).value="金额"
ws.cell(row=1,column=3).value="乐盒饭份数"
ws.cell(row=1,column=4).value="星宝外卖份数"
ws.cell(row=1,column=5).value="食点一刻份数"
ws.cell(row=1,column=6).value="讯美份数"
ws.cell(row=1,column=7).value="支付方式"

print("please wait a……")
for i in range(2,wo.get_highest_row()+1):
     data=wo.cell(row=i,column=3).value
     ws.cell(row=i,column=2).value=data
     date=wo.cell(row=i,column=2).value
     ws.cell(row=i,column=1).value=date
     pay=wo.cell(row=i,column=1).value
     ws.cell(row=i,column=7).value=pay
     if data%18.8==0:
         ws.cell(row=i,column=5).value=data//18.8
     elif data%18==0:
        ws.cell(row=i,column=6).value=data//18
     elif data%16==0:
        ws.cell(row=i,column=3).value=0
        ws.cell(row=i,column=4).value=data//16
     elif data%15==0:
         ws.cell(row=i,column=3).value=data//15
         ws.cell(row=i,column=4).value=0
     elif data/15<1:
            ws.cell(row=i,column=3).value=0
            ws.cell(row=i,column=4).value=0
     elif data/15>1and data/15<2:
           ws.cell(row=i,column=3).value=0
           ws.cell(row=i,column=4).value=1
     elif data/15>2and data/15<3:
           ws.cell(row=i,column=3).value=1
           ws.cell(row=i,column=4).value=1
     else:
           ws.cell(row=i,column=3).value=None
           ws.cell(row=i,column=4).value=None
w.save(r'H:\upupup\电商后台\结算管理\乐盒饭\5月\5月午餐核对结果.xlsx')
