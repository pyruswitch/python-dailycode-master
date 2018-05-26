__author__ = 'vincent'
# -*- coding:utf-8 -*-
import urllib
from pyquery import PyQuery as Pq
import StrUtil
import time
import os
from openpyxl import Workbook



def get_page_content_str( url):
    time.sleep(1)

    try:
            print("现在开始抓取" + url)
            headers = {'User-Agent': 'Mozilla/5.0 (Windows NT 6.1; WOW64; rv:23.0) Gecko/20100101 Firefox/23.0'}
            #伪装浏览器
            request = urllib.request.Request(url=url, headers=headers)
            #构造请求
            m_fp = urllib.request.urlopen(request, timeout=500)
            #访问网站获取源码
            html_str = m_fp.read().decode('utf-8')
            #读取源码，该网站使用的编码方式是utf-8

            return html_str
    except Exception:
            print(Exception)
            try:
                 request = urllib.request.Request(url=url, headers=headers)
                 m_fp = urllib.request.urlopen(request, timeout=500)
                 html_str= m_fp.read().decode('gbk')
                 return html_str
            except Exception:
             print(Exception)
             try:
                 request = urllib.request.Request(url=url, headers=headers)
                 m_fp = urllib.request.urlopen(request, timeout=500)
                 html_str= m_fp.read().decode('gb2312')
                 return html_str
             except Exception:
                      print(Exception)
            m_fp.close()
#定义抓取网页源码函数

baseurl="http://www.kkguan.com/list-1978-{}-sale-desc.html"
w= Workbook()
#创建excel工作薄
sheetnames = w.get_sheet_names()
ws= w.get_sheet_by_name(sheetnames[0])
#创建Excel工作表
ws.cell(row=1,column=1).value="商品名称"
ws.cell(row=1,column=2).value="售价"
ws.cell(row=1,column=3).value="评价数"
ws.cell(row=1,column=4).value="销量"
ws.cell(row=1,column=5).value="评分"
j=1
for i in range(1,9):
    url=baseurl.format(i)
    print("现在开始抓取第{}页".format(i))
    html_str=get_page_content_str(url)
    doc=Pq(html_str)
    prods=doc("body > div.wrapper1200 > div:nth-child(4) > ul > li > a.pic ")
    for prod  in prods:
        j=j+1
        prodname=Pq(prod).attr("title")
        prodhref="http://www.kkguan.com"+Pq(prod).attr("href")


        proddetail_str=get_page_content_str(prodhref)
        detail=Pq(proddetail_str)
        prodprice=detail("body > div.wrapper1200.clx > div.boder01.product_details_b2c.clx > div.info > ul > li.price > dl > p > em").text()
        feedrate=detail("body > div.wrapper1200.clx > div.boder01.product_details_b2c.clx > div.info > ul > li.mt10 > p:nth-child(1)").text()
        prodevacount=detail("body > div.wrapper1200.clx > div.boder01.product_details_b2c.clx > div.info > ul > li.mt10 > p:nth-child(1) > tt").text()
        salecount=detail("body > div.wrapper1200.clx > div.boder01.product_details_b2c.clx > div.info > ul > li.mt10 > p.mt10 > span > strong").text()
        ws.cell(row=j,column=1).value=prodname
        ws.cell(row=j,column=2).value=prodprice
        ws.cell(row=j,column=3).value=prodevacount
        ws.cell(row=j,column=4).value=salecount
        ws.cell(row=j,column=5).value=feedrate
        print(prodname+"-"+prodprice)
w.save(r'H:\upupup\电商后台\经营例会\日常运营需求\商品\KK馆.xlsx')