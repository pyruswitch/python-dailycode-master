__author__ = 'vincent'
# -*- coding:utf-8 -*-
from openpyxl import Workbook
import urllib
from pyquery import PyQuery as Pq
from selenium import webdriver
import time
import StrUtil

w= Workbook()
#创建excel工作薄
ws= w.create_sheet()
#创建Excel工作表

def get_page_content_str( url):
    time.sleep(1)

    try:
            print("现在开始抓取" + url)
            headers = {'User-Agent': 'Mozilla/5.0 (Windows NT 6.1; WOW64; rv:23.0) Gecko/20100101 Firefox/23.0'}
            #伪装浏览器
            request = urllib.request.Request(url=url, headers=headers)
            #构造请求
            m_fp = urllib.request.urlopen(request, timeout=500)
            #访问网站获取源码
            html_str = m_fp.read().decode('gb2312')
            #读取源码，该网站使用的编码方式是utf-8
            if html_str == None:
                html_str= m_fp.read().decode('gbk')
                if html_str == None:
                    html_str = m_fp.read().decode('utf-8')
            m_fp.close()
            return html_str
    except Exception:
            print(Exception)
#定义抓取网页源码函数

baseurl="http://www.xzlou.com/overall/index.html"
#目标网站URL
j=1
ws.cell(row=1,column=1).value="城市"
ws.cell(row=1,column=2).value="园区名"
ws.cell(row=1,column=3).value="面积"
ws.cell(row=1,column=4).value="简介"
#设置excel单元格表头
for i in range (1,2):
    url=baseurl
    print("现在开始抓取第{}页".format(i))
    html_str=get_page_content_str(url)
    if html_str==None:
         try:
            headers = {'User-Agent': 'Mozilla/5.0 (Windows NT 6.1; WOW64; rv:23.0) Gecko/20100101 Firefox/23.0'}
            #伪装浏览器
            request = urllib.request.Request(url=url, headers=headers)
            #构造请求
            m_fp = urllib.request.urlopen(request, timeout=500)
            #访问网站获取源码
            html_str = m_fp.read().decode('utf-8')
         except Exception:
            print(Exception)
            if html_str==None:
             try:
              headers = {'User-Agent': 'Mozilla/5.0 (Windows NT 6.1; WOW64; rv:23.0) Gecko/20100101 Firefox/23.0'}
              #伪装浏览器
              request = urllib.request.Request(url=url, headers=headers)
              #构造请求
              m_fp = urllib.request.urlopen(request, timeout=500)
              #访问网站获取源码
              html_str = m_fp.read().decode('gbk')
             except Exception:
              print(Exception)
    doc=Pq(html_str)
    parks=doc("body > div.main-posts > div > div > div > div")
    for park in parks:
        j=j+1
        city="北京"
        parkname=Pq(park)(" div > div.post-hover.text-center > div > h4 > a").text()
        parkurl=Pq(park)("  div > div.post-hover.text-center > div > h4 > a").attr("href")
        abs=Pq(park)(" div > div.post-hover.text-center > div > p > a").text()
        detailhtml=get_page_content_str(parkurl)

        if detailhtml==None:
         try:
            headers = {'User-Agent': 'Mozilla/5.0 (Windows NT 6.1; WOW64; rv:23.0) Gecko/20100101 Firefox/23.0'}
            #伪装浏览器
            request = urllib.request.Request(url=parkurl, headers=headers)
            #构造请求
            m_fp = urllib.request.urlopen(request, timeout=500)
            #访问网站获取源码
            detailhtml = m_fp.read().decode('utf-8')
         except Exception:
            print(Exception)
            if detailhtml==None:
             try:
              headers = {'User-Agent': 'Mozilla/5.0 (Windows NT 6.1; WOW64; rv:23.0) Gecko/20100101 Firefox/23.0'}
              #伪装浏览器
              request = urllib.request.Request(url=parkurl, headers=headers)
              #构造请求
              m_fp = urllib.request.urlopen(request, timeout=500)
              #访问网站获取源码
              detailhtml = m_fp.read().decode('gbk').encode("iso8859-1")
             except Exception:
              print(Exception)
        detail=Pq(detailhtml)
        area=detail("#projectID").text()
        area=StrUtil.substr(area,"：","米")
        ws.cell(row=j,column=1).value=city
        ws.cell(row=j,column=2).value=parkname
        ws.cell(row=j,column=3).value=area
        ws.cell(row=j,column=4).value=abs
        #写入excel单元格
        print(city+"-"+"-"+parkname+"-"+area+"-"+"-"+abs)
        #在控制台输出结果
w.save(r'D:\pythontest\园区\北京文化产业网.xlsx')
#保存到本地excel文件