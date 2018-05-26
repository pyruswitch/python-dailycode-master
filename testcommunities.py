__author__ = 'vincent'
# -*- coding:utf-8 -*-
import urllib
from pyquery import PyQuery as Pq
import time
from baidumap import xBaiduMap
from openpyxl import Workbook
from openpyxl.reader.excel import load_workbook
from openpyxl.writer.excel import ExcelWriter
from openpyxl.cell import get_column_letter

wb = load_workbook('E:/pythontest/南通小区.xlsx')
sheetnames = wb.get_sheet_names()
wo = wb.get_sheet_by_name(sheetnames[0])
w= Workbook()
ws= w.create_sheet(0)
def substr(mainstr, startstr, endstr):
    try:
        start_index = mainstr.index(startstr) + len(startstr)
        try:
            end_index = mainstr.index(endstr, start_index + 1)
            return mainstr[start_index:end_index]
        except:
            return mainstr[start_index:]
    except:
        return mainstr
ws.cell(row=1,column=1).value ="城市名"
ws.cell(row=1,column=2).value ="行政区"
ws.cell(row=1,column=3).value ="小区名"
ws.cell(row=1,column=4).value ="地址"
ws.cell(row=1,column=5).value ="纬度"
ws.cell(row=1,column=6).value ="经度"
ws.cell(row=1,column=7).value ="爬虫小区名"
if __name__ == '__main__':
    bm = xBaiduMap()
    print("please wait a minute....")
    for i in range(2,wo.get_highest_row()+1):
     city=wo.cell(row = i,column =1).value
     #print(city)
     name=wo.cell(row = i,column =2).value
     #print(name)
     ws.cell(row=i,column=3).value =name
     zuobiaos = bm.getSuggestion(name, city)
     if len(zuobiaos):
        zuobiao=str(zuobiaos[0])
        print(zuobiao)
        infor=str(zuobiao)
        inforcity=substr(infor,"'city': '","'")
        ws.cell(row=i,column=1).value =inforcity
        inforarea=substr(infor,"'district': '","'")
        ws.cell(row=i,column=2).value = inforarea
        inforname=substr(infor,"'name': '","'")
        ws.cell(row=i,column=7).value =inforname
        inforloca=substr(infor,": {","}")
        lat=substr(inforloca,"'lat': ",",")
        ws.cell(row=i,column=5).value =lat
        lng=substr(inforloca,"'lng': ",",")
        ws.cell(row=i,column=6).value =lng
        address=bm.getAddress(lat,lng)
        ws.cell(row=i,column=4).value =address
     else: pass
w.save('E:/pythontest/zuobiaomingdan.xlsx')