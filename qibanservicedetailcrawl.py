__author__ = 'vincent'
# -*- coding:utf-8 -*-
import urllib
from pyquery import PyQuery as Pq
import StrUtil
import time
import os
from openpyxl import Workbook



def get_page_content_str( url,i):
    time.sleep(1)

    try:
            print("现在开始抓取" + url)
            headers = {'User-Agent': 'Mozilla/5.0 (Windows NT 6.1; WOW64; rv:23.0) Gecko/20100101 Firefox/23.0'}
            #伪装浏览器
            values = {"page":i}
            data = urllib.parse.urlencode(values).encode('utf-8')
            request = urllib.request.Request(url=url, headers=headers,data=data)
            #构造请求
            m_fp = urllib.request.urlopen(request, timeout=500)
            #访问网站获取源码
            html_str = m_fp.read().decode('utf-8')
            #读取源码，该网站使用的编码方式是utf-8

            return html_str
    except Exception:
            print(Exception)
            try:
                 request = urllib.request.Request(url=url, headers=headers,data=data)
                 m_fp = urllib.request.urlopen(request, timeout=500)
                 html_str= m_fp.read().decode('gbk')
                 return html_str
            except Exception:
             print(Exception)
             try:
                 request = urllib.request.Request(url=url, headers=headers,data=data)
                 m_fp = urllib.request.urlopen(request, timeout=500)
                 html_str= m_fp.read().decode('gb2312')
                 return html_str
             except Exception:
                      print(Exception)
            m_fp.close()
#定义抓取网页源码函数

baseurl="http://www.qiban.com/bp/bpindex.jhtml"
w= Workbook()
#创建excel工作薄
sheetnames = w.get_sheet_names()
ws= w.get_sheet_by_name(sheetnames[0])
#创建Excel工作表
ws.cell(row=1,column=1).value="微信名称"
ws.cell(row=1,column=2).value="微信号"
ws.cell(row=1,column=3).value="分类"
ws.cell(row=1,column=4).value="粉丝数"
ws.cell(row=1,column=5).value="价格"
j=1
for i in range(0,48):
    url=baseurl
    print("现在开始抓取第{}页".format(i))
    html_str=get_page_content_str(url,i)
    doc=Pq(html_str)
    prods=doc("#myTable > tbody > tr")
    for prod  in prods:
        j=j+1
        wecname=Pq(prod)(" td:nth-child(2)").text()
        wecnum=Pq(prod)(" td:nth-child(3)").text()
        wecadre=Pq(prod)("td:nth-child(4)").text()
        wecgroup=Pq(prod)("td:nth-child(5)").text()
        fanstotal=Pq(prod)(" td:nth-child(6)").text()
        price=Pq(prod)("td:nth-child(7)").text()
        ws.cell(row=j,column=1).value=wecname
        ws.cell(row=j,column=2).value=wecnum
        ws.cell(row=j,column=3).value=wecadre
        ws.cell(row=j,column=4).value=fanstotal
        ws.cell(row=j,column=5).value=price
        print(wecname+"-"+wecnum)
w.save(r'H:\upupup\电商后台\品牌需求\企业推广.xlsx')