__author__ = 'vincent'
# -*- coding:utf-8 -*-
from openpyxl import Workbook
import urllib
from pyquery import PyQuery as Pq
from selenium import webdriver
import time
import StrUtil

w= Workbook()
#创建excel工作薄
sheetnames = w.get_sheet_names()
ws= w.get_sheet_by_name(sheetnames[0])
#创建Excel工作表

def get_page_content_str( url):
    time.sleep(1)

    try:
            print("现在开始抓取" + url)
            headers = {'User-Agent': 'Mozilla/5.0 (Windows NT 6.1; WOW64; rv:23.0) Gecko/20100101 Firefox/23.0'}
            #伪装浏览器
            request = urllib.request.Request(url=url, headers=headers)
            #构造请求
            m_fp = urllib.request.urlopen(request, timeout=500)
            #访问网站获取源码
            html_str = m_fp.read().decode('utf-8')
            #读取源码，该网站使用的编码方式是utf-8

            return html_str
    except Exception:
            print(Exception)
            try:
                 request = urllib.request.Request(url=url, headers=headers)
                 m_fp = urllib.request.urlopen(request, timeout=500)
                 html_str= m_fp.read().decode('gbk')
                 return html_str
            except Exception:
             print(Exception)
             try:
                 request = urllib.request.Request(url=url, headers=headers)
                 m_fp = urllib.request.urlopen(request, timeout=500)
                 html_str= m_fp.read().decode('gb2312')
                 return html_str
             except Exception:
                      print(Exception)
            m_fp.close()
#定义抓取网页源码函数

baseurl="http://y.zhaoshang800.com/dqlist-1-5-0--0-0-csname--p-{}.html"
#目标网站URL
j=1
ws.cell(row=1,column=1).value="区域"
ws.cell(row=1,column=2).value="行业"
ws.cell(row=1,column=3).value="类型"
ws.cell(row=1,column=4).value="级别"
ws.cell(row=1,column=5).value="园区名"
ws.cell(row=1,column=6).value="简介"
#设置excel单元格表头
for i in range (1,21):
    url=baseurl.format(i)
    print("现在开始抓取第{}页".format(i))
    html_str=get_page_content_str(url)
    if html_str==None:
         try:
            headers = {'User-Agent': 'Mozilla/5.0 (Windows NT 6.1; WOW64; rv:23.0) Gecko/20100101 Firefox/23.0'}
            #伪装浏览器
            request = urllib.request.Request(url=url, headers=headers)
            #构造请求
            m_fp = urllib.request.urlopen(request, timeout=500)
            #访问网站获取源码
            html_str = m_fp.read().decode('utf-8')
         except Exception:
            print(Exception)
            if html_str==None:
             try:
              headers = {'User-Agent': 'Mozilla/5.0 (Windows NT 6.1; WOW64; rv:23.0) Gecko/20100101 Firefox/23.0'}
              #伪装浏览器
              request = urllib.request.Request(url=url, headers=headers)
              #构造请求
              m_fp = urllib.request.urlopen(request, timeout=500)
              #访问网站获取源码
              html_str = m_fp.read().decode('gbk')
             except Exception:
              print(Exception)
    doc=Pq(html_str)
    parks=doc("body > div.W1000.recs > div.FL > ul > li")
    for park in parks:
        j=j+1
        parkname=Pq(park)("a").attr("title")
        kind=Pq(park)("dl > dd:eq(1) > span:nth-child(3)").text()
        kind=StrUtil.substr(kind,"：","'")
        pcb=Pq(park)("dl > dd:eq(1) > span:nth-child(2)").text()
        pcb=StrUtil.substr(pcb,"：","'")
        detail=Pq(park)(" dl > dd:eq(0)").text()
        detail=StrUtil.substr(detail,'"',"'")
        area=Pq(park)("dl > dd:eq(1) > span:nth-child(1)").text()
        area=StrUtil.substr(area,"：","'")
        level=Pq(park)("dl > dd:eq(1) > span:nth-child(4)").text()
        level=StrUtil.substr(level,"：","'")
        ws.cell(row=j,column=1).value=area
        ws.cell(row=j,column=2).value=pcb
        ws.cell(row=j,column=3).value=kind
        ws.cell(row=j,column=4).value=level
        ws.cell(row=j,column=5).value=parkname
        ws.cell(row=j,column=6).value=detail
        #写入excel单元格
        print(area+"-"+pcb+"-"+kind+"-"+level+"-"+parkname)
        #在控制台输出结果
w.save(r'D:\pythontest\园区\广州-招商园区.xlsx')
#保存到本地excel文件