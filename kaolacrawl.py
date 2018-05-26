__author__ = 'vincent'
# -*- coding:utf-8 -*-
import urllib
from pyquery import PyQuery as Pq
import StrUtil
import time
import os
from openpyxl import Workbook



def get_page_content_str( url):
    time.sleep(1)

    try:
            print("现在开始抓取" + url)
            headers = {'User-Agent': 'Mozilla/5.0 (Windows NT 6.1; WOW64; rv:23.0) Gecko/20100101 Firefox/23.0'}
            #伪装浏览器
            request = urllib.request.Request(url=url, headers=headers)
            #构造请求
            m_fp = urllib.request.urlopen(request, timeout=500)
            #访问网站获取源码
            html_str = m_fp.read().decode('utf-8')
            #读取源码，该网站使用的编码方式是utf-8

            return html_str
    except Exception:
            print(Exception)
            try:
                 request = urllib.request.Request(url=url, headers=headers)
                 m_fp = urllib.request.urlopen(request, timeout=500)
                 html_str= m_fp.read().decode('gbk')
                 return html_str
            except Exception:
             print(Exception)
             try:
                 request = urllib.request.Request(url=url, headers=headers)
                 m_fp = urllib.request.urlopen(request, timeout=500)
                 html_str= m_fp.read().decode('gb2312')
                 return html_str
             except Exception:
                      print(Exception)
            m_fp.close()
#定义抓取网页源码函数

baseurl="http://www.kaola.com/category/5905.html?key=&pageSize=60&pageNo={}&sortfield=2&isStock=false&isSelfProduct=false&isPromote=false&isDesc=true&b=&proIds=&source=false&country=&needBrandDirect=false&isNavigation=0&lowerPrice=-1&upperPrice=-1&backCategory=&headCategoryId=&changeContent=type#topTab"
w= Workbook()
#创建excel工作薄
sheetnames = w.get_sheet_names()
ws= w.get_sheet_by_name(sheetnames[0])
#创建Excel工作表
ws.cell(row=1,column=1).value="商品名称"
ws.cell(row=1,column=2).value="售价"
ws.cell(row=1,column=3).value="评价数"
ws.cell(row=1,column=4).value="好评率"
ws.cell(row=1,column=5).value="供货商"
j=1
for i in range(1,18):
    url=baseurl.format(i)
    print("现在开始抓取第{}页".format(i))
    html_str=get_page_content_str(url)
    doc=Pq(html_str)
    prods=doc("#result > li > div ")
    for prod  in prods:
        j=j+1
        prodname=Pq(prod)("a").attr("title")
        prodhref="http://www.kaola.com"+Pq(prod)("a").attr("href")
        prodevacount=Pq(prod)("div > p.goodsinfo.clearfix > a").text()
        prodprice=Pq(prod)("div > p.price > span.cur").text()
        proddetail_str=get_page_content_str(prodhref)
        detail=Pq(proddetail_str)
        supplier=detail("#js_skuBox > div > form > span").text()
        feedrate=detail("#j-producthead > div.PInfoWrap.clearfix > dl > div > dd.m-service.f-cb > span.send").text()
        ws.cell(row=j,column=1).value=prodname
        ws.cell(row=j,column=2).value=prodprice
        ws.cell(row=j,column=3).value=prodevacount
        ws.cell(row=j,column=4).value=feedrate
        ws.cell(row=j,column=5).value=supplier
        print(prodname+"-"+prodprice)
w.save(r'H:\upupup\电商后台\经营例会\日常运营需求\商品\网易考拉-饼干糕点.xlsx')