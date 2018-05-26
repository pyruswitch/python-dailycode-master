__author__ = 'vincent'
# -*- coding:utf-8 -*-
from openpyxl import Workbook
import urllib
from pyquery import PyQuery as Pq
from selenium import webdriver
import time
import StrUtil

w= Workbook()
#创建excel工作薄
ws= w.create_sheet(0)
#创建Excel工作表

def get_page_content_str( url):
    time.sleep(1)

    try:
            print("现在开始抓取" + url)
            headers = {'User-Agent': 'Mozilla/5.0 (Windows NT 6.1; WOW64; rv:23.0) Gecko/20100101 Firefox/23.0'}
            #伪装浏览器
            request = urllib.request.Request(url=url, headers=headers)
            #构造请求
            m_fp = urllib.request.urlopen(request, timeout=500)
            #访问网站获取源码
            html_str = m_fp.read().decode('utf-8')
            #读取源码，该网站使用的编码方式是utf-8
            if html_str == None:
                html_str= m_fp.read().decode('gbk')
                if html_str == None:
                    html_str = m_fp.read().decode('gb2312')
            m_fp.close()
            return html_str
    except Exception:
            print(Exception)
#定义抓取网页源码函数

baseurl="http://www.sfbest.com/fresh/52-0-0-0-{}-2-0-0-0-0-0.html#trackref=sfbest_channel_fresh_floor1-more"
#目标网站URL
j=1
ws.cell(row=1,column=1).value="名称"
ws.cell(row=1,column=2).value="品牌"
ws.cell(row=1,column=3).value="产地"
ws.cell(row=1,column=4).value="规格"
ws.cell(row=1,column=5).value="价格"
#设置excel单元格表头
for i in range (0,16):
    url=baseurl.format(i)
    print("现在开始抓取第{}页".format(i+1))
    html_str=get_page_content_str(url)
    doc=Pq(html_str)
    fruits=doc("body > div.content_list > div.main-box > div.p-list > ul>li")
    for fruit in fruits:
        j=j+1
        fruitname=Pq(fruit)("div > div:eq(2)>a").text()
        fruiturl=Pq(fruit)("div > div:eq(2)>a").attr("href")
        fruitpri=Pq(fruit)("div > div:eq(1)>span>span>strong").text()
        detailhtml=get_page_content_str(fruiturl)
        detail=Pq(detailhtml)
        fruitbrand=detail("body > div:eq(7)>div>div:eq(1)>div:eq(1)>ul>li:eq(0)>span>a").text()
        if "产地" in detail("body > div:eq(7)>div>div:eq(1)>div:eq(1)>ul>li:eq(0)").text():
            fruitbrand=""
            fruitarea=detail("body > div:eq(7)>div>div:eq(1)>div:eq(1)>ul>li:eq(0)>span>a").text()
            fruitspec=detail("body > div:eq(7)> div > div.pItemsSide > div.pdetail > ul > li:nth-child(2)").text()
        else:
            fruitarea=detail("body > div:eq(7)>div>div:eq(1)>div:eq(1)>ul>li:eq(1)>span>a").text()
            fruitspec=detail("body > div:eq(7) > div > div.pItemsSide > div.pdetail > ul > li:eq(2)").text()
        fruitspec=StrUtil.substr(fruitspec,'：',"'")
        ws.cell(row=j,column=1).value=fruitname
        ws.cell(row=j,column=2).value=fruitbrand
        ws.cell(row=j,column=3).value=fruitarea
        ws.cell(row=j,column=4).value=fruitspec
        ws.cell(row=j,column=5).value=fruitpri
        #写入excel单元格
        print(fruitname+"-"+"-"+fruitbrand+"-"+fruitarea+"-"+fruitspec+"-"+fruitpri)
        #在控制台输出结果
w.save(r'E:\pythontest\电商\sffruits.xlsx')
#保存到本地excel文件