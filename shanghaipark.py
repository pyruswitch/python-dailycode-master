__author__ = 'vincent'
# -*- coding:utf-8 -*-
from openpyxl import Workbook
import urllib
from pyquery import PyQuery as Pq
from selenium import webdriver
import time
import StrUtil

w= Workbook()
#创建excel工作薄
ws= w.create_sheet()
#创建Excel工作表

def get_page_content_str( url):
    time.sleep(1)

    try:
            print("现在开始抓取" + url)
            headers = {'User-Agent': 'Mozilla/5.0 (Windows NT 6.1; WOW64; rv:23.0) Gecko/20100101 Firefox/23.0'}
            #伪装浏览器
            request = urllib.request.Request(url=url, headers=headers)
            #构造请求
            m_fp = urllib.request.urlopen(request, timeout=500)
            #访问网站获取源码
            html_str = m_fp.read().decode('gb2312')
            #读取源码，该网站使用的编码方式是utf-8
            if html_str == None:
                html_str= m_fp.read().decode('gbk')
                if html_str == None:
                    html_str = m_fp.read().decode('utf-8')
            m_fp.close()
            return html_str
    except Exception:
            print(Exception)
#定义抓取网页源码函数

baseurl="http://shcci.eastday.com/eastday/shcci/node735815/node735871/index.html"
#目标网站URL
j=1
ws.cell(row=1,column=1).value="行政区"
ws.cell(row=1,column=2).value="园区名"
ws.cell(row=1,column=3).value="简介"
#设置excel单元格表头
for i in range (1,2):
    url=baseurl.format(i)
    print("现在开始抓取第{}页".format(i))
    html_str=get_page_content_str(url)
    if html_str==None:
         try:
            headers = {'User-Agent': 'Mozilla/5.0 (Windows NT 6.1; WOW64; rv:23.0) Gecko/20100101 Firefox/23.0'}
            #伪装浏览器
            request = urllib.request.Request(url=url, headers=headers)
            #构造请求
            m_fp = urllib.request.urlopen(request, timeout=500)
            #访问网站获取源码
            html_str = m_fp.read().decode('utf-8')
         except Exception:
            print(Exception)
            if html_str==None:
             try:
              headers = {'User-Agent': 'Mozilla/5.0 (Windows NT 6.1; WOW64; rv:23.0) Gecko/20100101 Firefox/23.0'}
              #伪装浏览器
              request = urllib.request.Request(url=url, headers=headers)
              #构造请求
              m_fp = urllib.request.urlopen(request, timeout=500)
              #访问网站获取源码
              html_str = m_fp.read().decode('gbk')
             except Exception:
              print(Exception)
    doc=Pq(html_str)
    parks=doc("body > div.main > div.piclb14")
    for park in parks:

      pro=Pq(park)(" h1 > a").text()
      names=Pq(park)("div > ul.lb16 > li")
      for name in names:
        j=j+1
        parkname=Pq(name)("a").text()
        parkurl=Pq(name)("a").attr("href")
        detailhtml=get_page_content_str(parkurl)
        if detailhtml==None:
         try:
            headers = {'User-Agent': 'Mozilla/5.0 (Windows NT 6.1; WOW64; rv:23.0) Gecko/20100101 Firefox/23.0'}
            #伪装浏览器
            request = urllib.request.Request(url=parkurl, headers=headers)
            #构造请求
            m_fp = urllib.request.urlopen(request, timeout=500)
            #访问网站获取源码
            detailhtml = m_fp.read().decode('utf-8')
         except Exception:
            print(Exception)
            if detailhtml==None:
             try:
              headers = {'User-Agent': 'Mozilla/5.0 (Windows NT 6.1; WOW64; rv:23.0) Gecko/20100101 Firefox/23.0'}
              #伪装浏览器
              request = urllib.request.Request(url=parkurl, headers=headers)
              #构造请求
              m_fp = urllib.request.urlopen(request, timeout=500)
              #访问网站获取源码
              detailhtml = m_fp.read().decode('iso8859-1').encode("iso8859-1")
             except Exception:
              print(Exception)
              if detailhtml==None:
               try:
                headers = {'User-Agent': 'Mozilla/5.0 (Windows NT 6.1; WOW64; rv:23.0) Gecko/20100101 Firefox/23.0'}
                #伪装浏览器
                request = urllib.request.Request(url="http://shcci.eastday.com/"+parkurl, headers=headers)
                #构造请求
                m_fp = urllib.request.urlopen(request, timeout=500)
                #访问网站获取源码
                detailhtml = m_fp.read().decode('iso8859-1').encode("iso8859-1")
               except Exception:
                 print(Exception)
        detail=Pq(detailhtml)
        abs=detail("body > div.main > div.piclb13 > div.piclb13k > div > ul > a").text()
        ws.cell(row=j,column=1).value=pro
        ws.cell(row=j,column=2).value= parkname
        ws.cell(row=j,column=3).value=abs
        print(pro+"-"+"-"+parkname+"-"+abs)
        #在控制台输出结果
w.save(r'D:\pythontest\园区\上海-shcci.xlsx')
#保存到本地excel文件