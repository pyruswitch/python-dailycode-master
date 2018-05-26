__author__ = 'vincent'
# -*- coding:utf-8 -*-
from openpyxl import Workbook
import urllib
from pyquery import PyQuery as Pq
from selenium import webdriver
import time
import StrUtil

w= Workbook()
#创建excel工作薄
ws= w.create_sheet(0)
#创建Excel工作表

def get_page_content_str( url):
    time.sleep(1)

    try:
            print("现在开始抓取" + url)
            headers = {'User-Agent': 'Mozilla/5.0 (Windows NT 6.1; WOW64; rv:23.0) Gecko/20100101 Firefox/23.0'}
            #伪装浏览器
            request = urllib.request.Request(url=url, headers=headers)
            #构造请求
            m_fp = urllib.request.urlopen(request, timeout=500)
            #访问网站获取源码
            html_str = m_fp.read().decode('utf-8')
            #读取源码，该网站使用的编码方式是utf-8
            if html_str == None:
                html_str= m_fp.read().decode('gbk')
                if html_str == None:
                    html_str = m_fp.read().decode('gb2312')
            m_fp.close()
            return html_str
    except Exception:
            print(Exception)
#定义抓取网页源码函数

baseurl="http://www.tootoo.cn/list-s1-13135-0-0-0-0-0-{}-0-0-0-1,2,3,0-zh_cn.html"
#目标网站URL
j=1
ws.cell(row=1,column=1).value="名称"
ws.cell(row=1,column=2).value="品牌"
ws.cell(row=1,column=3).value="产地"
ws.cell(row=1,column=4).value="规格"
ws.cell(row=1,column=5).value="价格"
ws.cell(row=1,column=6).value="评论"

#设置excel单元格表头
for i in range (1,9):
    url=baseurl.format(i)
    print("现在开始抓取第{}页".format(i))
    html_str=get_page_content_str(url)
    doc=Pq(html_str)
    fruits=doc("#list_goodslist > li")
    for fruit in fruits:
        j=j+1
        fruitname=Pq(fruit)("div.pro_title > a").text()
        fruiturl=Pq(fruit)("div.pro_title > a").attr("href")
        fruitpri=Pq(fruit)(" div.pro_price > b").text()
        fruitsale=Pq(fruit)(" div.pro_span > div.s_r > a").text()
        detailhtml=get_page_content_str(fruiturl)
        detail=Pq(detailhtml)
        fruitbrand=detail("#detailInfo_list > div:nth-child(1) > ul").text()
        fruitbrand=StrUtil.substr(fruitbrand,'品牌：',' ')
        #fruitarea=detail("#prodDetailCotentDiv > dl > dd").attr("title","产地").text()
        fruitspec=detail("#detailInfo_list > div:nth-child(1) > ul").text()
        fruitspec=StrUtil.substr(fruitspec,'规格： '," ")

        ws.cell(row=j,column=1).value=fruitname
        ws.cell(row=j,column=2).value=fruitbrand

        #ws.cell(row=j,column=3).value=fruitarea
        ws.cell(row=j,column=4).value=fruitspec
        ws.cell(row=j,column=5).value=fruitpri
        ws.cell(row=j,column=6).value=fruitsale

        #写入excel单元格
        print(fruitname+"-"+"-"+fruitbrand+"-"+fruitspec+"-"+fruitpri+"-"+fruitsale)
        #在控制台输出结果
w.save(r'd:\pythontest\电商\tuosock.xlsx')
#保存到本地excel文件