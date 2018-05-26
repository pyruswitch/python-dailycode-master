__author__ = 'vincent'
# -*- coding:utf-8 -*-
from openpyxl import Workbook
import urllib
from pyquery import PyQuery as Pq
from selenium import webdriver
import time
import StrUtil

w= Workbook()
#创建excel工作薄
ws= w.create_sheet(0)
#创建Excel工作表



baseurl="http://gz.womai.com/Sort-31359-497036.htm"
#目标网站URL
j=1
ws.cell(row=1,column=1).value="名称"
ws.cell(row=1,column=2).value="价格"
ws.cell(row=1,column=3).value="评价"
#设置excel单元格表头
driver=webdriver.Chrome("H:\下载\chromedriver_win32\chromedriver.exe")
driver.get(baseurl)
print("现在开始抓取"+baseurl)
driver.find_element_by_css_selector("#orderSellCount").click()
time.sleep(4)
for page in range(1,5):
    time.sleep(2)
    w.save(r'H:\upupup\电商后台\经营例会\日常运营需求\商品\我买网-地方特产.xlsx')
    for i in range(1,41):
        j=j+1
        try:
         name=driver.find_element_by_css_selector("#ajaxrightshow > div.product-list > ul>li:nth-child({}) >div>div:nth-child(3)>p ".format(i)).get_attribute("title")
        except  Exception:
            print(Exception)
            name=""

        try:
         price=driver.find_element_by_css_selector("#ajaxrightshow > div.product-list > ul>li:nth-child({}) >div>div:nth-child(2)".format(i)).text
         if price==""or "￥"not in price:
             price=driver.find_element_by_css_selector("#ajaxrightshow > div.product-list > ul>li:nth-child({}) >div:nth-child(2)>div:nth-child(3)".format(i)).text
         if price==""or "￥"not in price:
             price=driver.find_element_by_css_selector("#ajaxrightshow > div.product-list > ul>li:nth-child({}) >div:nth-child(2)>div:nth-child(2)".format(i)).text
         if price==""or "￥"not in price:
             price=driver.find_element_by_css_selector("#ajaxrightshow > div.product-list > ul>li:nth-child({}) >div>div:eq(1)".format(i)).text
        except  Exception:
            print(Exception)
            price=""
        #detailurl=driver.find_element_by_css_selector("body > div.wrapper > div.page-wrap > div > div > div.rest-list > ul > li:nth-child({}) > div > div > a.rest-atag".format(i))
        #detailurl.click()
        try:
         eva=driver.find_element_by_css_selector("#ajaxrightshow > div.product-list > ul>li:nth-child({}) >div>div:nth-child(4)>span>a>em".format(i)).text
        except  Exception:
            print(Exception)
            eva=""
        #adress=driver.find_element_by_css_selector("body > div.wrapper > div.page-wrap > div > div.rest-info > div.details > div.rest-info-down-wrap > div.location.fl > span.fl.info-detail").text
        #phone=driver.find_element_by_css_selector("body > div.wrapper > div.page-wrap > div > div.rest-info > div.details > div.rest-info-down-wrap > div.telephone.fl > span.fl.info-detail").text
        #driver.close()
        ws.cell(row=j,column=1).value=name
        ws.cell(row=j,column=2).value=price
        ws.cell(row=j,column=3).value=eva

        #写入excel单元格
        print(name+"-"+"-"+ price)

    if page==12:
        break
    driver.find_element_by_css_selector("#ajaxrightshow > div.product-list > div.pglist_page > div > div.page_l > a.next").click()

    time.sleep(1)
        #在控制台输出结果

#保存到本地excel文件